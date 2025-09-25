import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


def calculate_excel_data(file_path, sheet_name, calc_type, source_range, target_cell):
    """
    读取Excel中指定区域的数据进行计算并写入目标单元格

    参数:
    file_path (str): Excel文件路径
    sheet_name (str): 工作表名称
    calc_type (str): 计算类型 - 'column_sum'（列求和）或 'row_sum'（行求和）
    source_range (str): 数据源范围，格式：
        - 列求和："列字母:起始行-结束行" 如 "C:2-8"
        - 行求和："行号:起始列字母-结束列字母" 如 "3:B-H"
    target_cell (str): 目标单元格地址（如"A1"）
    """
    # 加载工作簿和工作表
    wb = load_workbook(file_path)
    sheet = wb[sheet_name]

    total = 0  # 初始化计算结果
    valid_values = []  # 存储有效数值

    # 解析源数据范围
    if calc_type == 'column_sum':
        # 列求和模式：格式 "列字母:起始行-结束行"
        col_letter, rows_range = source_range.split(':')
        start_row, end_row = map(int, rows_range.split('-'))

        # 遍历指定列的行范围
        for row in range(start_row, end_row + 1):
            cell_address = f"{col_letter}{row}"
            cell = sheet[cell_address]

            # 检查单元格值是否为数字
            if isinstance(cell.value, (int, float)):
                total += cell.value
                valid_values.append(cell.value)

    elif calc_type == 'row_sum':
        # 行求和模式：格式 "行号:起始列-结束列"
        row_num, cols_range = source_range.split(':')
        start_col, end_col = cols_range.split('-')

        # 将列字母转换为数字索引
        start_col_idx = column_index_from_string(start_col)
        end_col_idx = column_index_from_string(end_col)
        row_num = int(row_num)

        # 遍历指定行的列范围
        for col_idx in range(start_col_idx, end_col_idx + 1):
            col_letter = get_column_letter(col_idx)
            cell_address = f"{col_letter}{row_num}"
            cell = sheet[cell_address]

            # 检查单元格值是否为数字
            if isinstance(cell.value, (int, float)):
                total += cell.value
                valid_values.append(cell.value)

    # # 在函数中添加新计算类型
    # elif calc_type == 'average':
    #     result = total / len(valid_values) if valid_values else 0


    # # 同时计算多个区域
    # ranges = ["C:2-8", "D:2-8", "E:2-8"]
    # for i, range_str in enumerate(ranges, start=1):
    #     total = calculate_range(sheet, 'column_sum', range_str)
    #     sheet[f"F{i}"] = total

    else:
        raise ValueError(f"无效的计算类型: {calc_type}. 请使用 'column_sum' 或 'row_sum'")

    # 输出计算信息（调试用）
    print(f"计算类型: {calc_type}")
    print(f"源数据范围: {source_range}")
    print(f"有效值: {valid_values}")
    print(f"求和结果: {total}")

    # 将结果写入目标单元格
    sheet[target_cell] = total

    # 保存修改后的工作簿
    wb.save(file_path)
    print(f"结果已写入单元格 {target_cell}，文件保存成功！")


# 示例使用
if __name__ == "__main__":
    # # 示例1：列求和（第3列2-8行求和到C1）
    # calculate_excel_data(
    #     file_path="test2.xlsx",
    #     sheet_name="Sheet1",
    #     calc_type="column_sum",
    #     source_range="C:2-8",  # C列2-8行
    #     target_cell="C1"
    # )

    # 示例2：行求和（第3行B-H列求和到A3）
    calculate_excel_data(
        file_path="test2.xlsx",
        sheet_name="Sheet1",
        calc_type="row_sum",
        source_range="3:D-Q",  # 3行B-H列
        target_cell="C3"
    )
    #
    # # 示例3：行求和（第5行C-F列求和到G5）
    # calculate_excel_data(
    #     file_path="example.xlsx",
    #     sheet_name="Sheet1",
    #     calc_type="row_sum",
    #     source_range="5:C-F",  # 5行C-F列
    #     target_cell="G5"
    # )