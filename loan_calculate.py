import pandas as pd
import datetime
from datetime import datetime as dt
import warnings
import os

from circulate_formula import ExcelFormulaGenerator
from findAndSet import find_all_cells, find_cell
from loan_assignment import write_loan, copy_and_delete_sheets, loan_summary
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl_vba import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
"""

表c填充主要文件
需要表A、表C.1、表C.2
目前借款总结之后的 “计算指标”、“利息备付率”、“偿债备付率”，没有添加，因为这一行的值取自表b，目前还没输出。

"""

# 忽略特定警告
warnings.filterwarnings("ignore", category=FutureWarning)


def excel_serial_to_year(serial):
    """
    将Excel序列日期转换为年份
    参数:
    serial: Excel日期序列数
    返回:
    年份（整数）
    """
    try:
        if isinstance(serial, (int, float)):
            base_date = dt(1899, 12, 30)  # Excel的基准日期（1900-01-01为1）
            date = base_date + datetime.timedelta(days=serial)
            return date.year
        elif isinstance(serial, str) and "年" in serial:
            return int(serial[:4])
        else:
            return dt.now().year
    except:
        return dt.now().year


class Loan:
    def __init__(self, loan_id, name, loan_type, amount, start_year, end_year, term,
                 interest_rate, repayment_method, first_interest_month,
                 bond_issue_fee, bond_registration_fee, bond_repayment_fee):
        """
        初始化借款对象
        参数:
        loan_id: 借款序号
        name: 借款名称
        loan_type: 借款类别
        amount: 借款金额(万元)
        start_year: 开始年份
        end_year: 结束年份
        term: 借款周期(年)
        interest_rate: 借款利率
        repayment_method: 还款方式
        first_interest_month: 首年计息月份
        bond_issue_fee: 债券发行费(万元)
        bond_registration_fee: 债券发行登记服务费(万元)
        bond_repayment_fee: 债券还本付息兑付手续费(万元)
        """
        self.loan_id = loan_id
        self.name = name
        self.loan_type = loan_type
        self.amount = amount
        self.start_year = start_year
        self.end_year = end_year
        self.term = term
        self.interest_rate = interest_rate
        self.repayment_method = repayment_method
        self.first_interest_month = first_interest_month
        self.bond_issue_fee = bond_issue_fee
        self.bond_registration_fee = bond_registration_fee
        self.bond_repayment_fee = bond_repayment_fee
        self.schedule = None

    def generate_schedule(self, all_years):
        """
        生成还款计划表
        参数:
        all_years: 所有年份列表
        返回:
        包含还款计划的DataFrame
        """
        # 创建空的还款计划DataFrame，初始化为0.0
        schedule = pd.DataFrame(
            index=all_years,
            columns=['期初借款余额', '还本', '付息', '期末借款余额', '当期还本付息', '还本付息兑付手续费',
                     '债券发行及服务费'
                , '应付利息'],
            dtype=float
        )
        schedule = schedule.fillna(0.0)
        #
        # # 设置提款计划
        # if self.start_year in schedule.index:
        #     schedule.loc[self.start_year, '期初借款余额'] = 0.0
        #     schedule.loc[self.start_year, '期末借款余额'] = self.amount
        #
        # # 初始化余额
        # prev_balance = self.amount if self.start_year in all_years else 0.0
        #
        # # 计算每一年数据
        # for year in all_years:
        #     if year < self.start_year:
        #         schedule.loc[year, '期初借款余额'] = 0.0
        #         schedule.loc[year, '期末借款余额'] = 0.0
        #         continue
        #
        #     # 设置期初余额
        #     schedule.loc[year, '期初借款余额'] = prev_balance
        #
        #     # 计算利息
        #     interest = prev_balance * self.interest_rate
        #     schedule.loc[year, '付息'] = interest
        #
        #     # 计算还本
        #     principal_payment = 0.0
        #     if year == self.end_year:  # 最后一年
        #         principal_payment = prev_balance
        #     elif self.repayment_method in ["到期前逐年还本", "贷款期内等额本金还款", "等额本金"]:
        #         # 等额本金还款
        #         if self.term > 0:
        #             # 计算每年应还本金
        #             principal_payment = self.amount / self.term
        #             # 确保不会偿还超过剩余本金
        #             if principal_payment > prev_balance:
        #                 principal_payment = prev_balance
        #
        #     schedule.loc[year, '还本'] = principal_payment
        #
        #     # 计算期末余额
        #     ending_balance = prev_balance - principal_payment
        #     schedule.loc[year, '期末借款余额'] = ending_balance
        #     prev_balance = ending_balance
        #
        #     # 计算当期还本付息
        #     schedule.loc[year, '当期还本付息'] = principal_payment + interest
        #
        #     # 计算手续费
        #     if self.term > 0:
        #         schedule.loc[year, '还本付息兑付手续费'] = self.bond_repayment_fee / self.term
        #
        # # 债券发行及服务费只在开始年份计入
        # if self.start_year in schedule.index:
        #     schedule.loc[self.start_year, '债券发行及服务费'] = self.bond_issue_fee + self.bond_registration_fee

        self.schedule = schedule
        return schedule

    def get_output_table(self, all_years):
        """
        生成符合模板格式的输出表格
        参数:
        all_years: 所有年份列表
        返回:
        按照模板格式的DataFrame
        """
        if self.schedule is None:
            self.generate_schedule(all_years)

        # 创建输出表格
        output_df = pd.DataFrame(
            columns=['序号', '项目', '合计'] + [year for year in all_years]
        )

        # 添加借款标题行
        output_df.loc[0] = [self.loan_id, self.name, ""] + ["" for _ in all_years]

        # 添加详细数据行
        output_df.loc[1] = [f"{self.loan_id}.1", "期初借款余额", self.schedule['期初借款余额'].sum()] + list(
            self.schedule['期初借款余额'].values)
        output_df.loc[2] = [f"{self.loan_id}.2", "当期还本付息", self.schedule['当期还本付息'].sum()] + list(
            self.schedule['当期还本付息'].values)
        output_df.loc[3] = [f"{self.loan_id}.2.1", "其中：还本", self.schedule['还本'].sum()] + list(
            self.schedule['还本'].values)
        output_df.loc[4] = [f"{self.loan_id}.2.2", "付息", self.schedule['付息'].sum()] + list(
            self.schedule['付息'].values)
        output_df.loc[5] = [f"{self.loan_id}.3", "期末借款余额", self.schedule['期末借款余额'].sum()] + list(
            self.schedule['期末借款余额'].values)
        output_df.loc[6] = [f"{self.loan_id}.4", "还本付息兑付手续费",self.schedule['还本付息兑付手续费'].sum()] + list(
            self.schedule['还本付息兑付手续费'].values)
        output_df.loc[7] = [f"{self.loan_id}.5", "债券发行及服务费", self.schedule['债券发行及服务费'].sum()] + list(
            self.schedule['债券发行及服务费'].values)
        output_df.loc[8] = [f"{self.loan_id}.6", "应付利息", self.schedule['应付利息'].sum()] + list(
            self.schedule['应付利息'].values)

        # 添加详细数据行

        return output_df


class LoanRepaymentSystem:
    def __init__(self):
        self.loans = []
        self.all_years = []

    def add_loan(self, loan_data):
        """添加借款到系统"""
        # 转换Excel序列日期为年份
        start_year = excel_serial_to_year(loan_data['开始时间'])
        end_year = excel_serial_to_year(loan_data['结束时间'])

        loan = Loan(
            loan_id=loan_data['序号'],
            name=loan_data['借款名称'],
            loan_type=loan_data['借款类别'],
            amount=loan_data['借款金额（万元）'],
            start_year=start_year,
            end_year=end_year,
            term=loan_data['借款周期（年）'],
            interest_rate=loan_data['借款利率'],
            repayment_method=loan_data['还款方式'],
            first_interest_month=loan_data['首年计息月份'],
            bond_issue_fee=loan_data['债券发行费（万元）'],
            bond_registration_fee=loan_data['债券发行登记服务费（万元）'],
            bond_repayment_fee=loan_data['债券还本付息兑付手续费（万元）']
        )
        self.loans.append(loan)

    def calculate_years_range(self):
        """计算所有年份范围"""
        start_years = [loan.start_year for loan in self.loans]
        end_years = [loan.end_year for loan in self.loans]

        if not start_years or not end_years:
            return []

        min_year = min(start_years)
        max_year = max(end_years)
        return list(range(min_year, max_year + 1))

    # def export_to_excel(self, file_path):
    #     """
    #     导出结果到Excel文件
    #     参数:
    #     file_path: 输出文件路径
    #     """
    #     if not self.loans:
    #         print("没有可导出的借款数据")
    #         return
    #
    #     # 计算所有年份
    #     self.all_years = self.calculate_years_range()
    #     print("年份为：", self.all_years)
    #     print("一共有:", len(self.all_years))
    #
    #     # 创建Excel写入器
    #     with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
    #
    #         # 为每笔借款创建单独的工作表
    #         for loan in self.loans:
    #             # 生成借款计划表
    #             loan_table = loan.get_output_table(self.all_years)
    #
    #             # 设置工作表名称
    #             sheet_name = f"借款{loan.loan_id}"
    #
    #             # 写入Excel
    #             loan_table.to_excel(writer, sheet_name=sheet_name, index=False)
    #         print(f"结果已成功导出到: {os.path.abspath(file_path)}")
    #         print(f"包含以下工作表: {[f'借款{loan.loan_id}' for loan in self.loans]}")
    #
    #         loan_all = loan
    #         loan_all.loan_id = loan.loan_id + 1
    #         loan_all.name = "借款总结"
    #         loan_table = loan_all.get_output_table(self.all_years)
    #         # 设置工作表名称
    #         sheet_name = f"借款{loan_all.loan_id}"
    #         # 写入Excel
    #         loan_table.to_excel(writer, sheet_name=sheet_name, index=False)
    #
    #         print("已添加借款总结项目")
    #
    #     return len(self.all_years), self.all_years[-1]

    def export_to_excel(self, file_path):
        """
        导出结果到Excel文件，保留VBA宏
        参数:
        file_path: 输出文件路径
        """
        if not self.loans:
            print("没有可导出的借款数据")
            return

        # 计算所有年份
        self.all_years = self.calculate_years_range()
        print("年份为：", self.all_years)
        print("一共有:", len(self.all_years))

        # 检查文件是否存在，如果存在则加载，否则创建新工作簿
        if os.path.exists(file_path):
            try:
                # 加载现有工作簿，保留VBA宏
                book = load_workbook(file_path)
                print("已加载现有工作簿并保留VBA宏")
            except Exception as e:
                print(f"加载工作簿时出错: {e}")
                # 创建新工作簿作为备选
                book = Workbook()
                # 删除默认工作表
                if book.sheetnames:
                    default_sheet = book.active
                    book.remove(default_sheet)
                print("创建了新工作簿")
        else:
            # 文件不存在，创建新工作簿
            book = Workbook()
            # 删除默认工作表
            if book.sheetnames:
                default_sheet = book.active
                book.remove(default_sheet)
            print("创建了新工作簿")

        # 为每笔借款创建单独的工作表
        for loan in self.loans:
            # 生成借款计划表
            loan_table = loan.get_output_table(self.all_years)

            # 设置工作表名称
            sheet_name = f"借款{loan.loan_id}"

            # 检查工作表是否已存在，如果存在则删除
            if sheet_name in book.sheetnames:
                book.remove(book[sheet_name])
                print(f"已删除现有工作表: {sheet_name}")

            # 创建新工作表
            ws = book.create_sheet(sheet_name)

            # 将DataFrame写入工作表
            for r in dataframe_to_rows(loan_table, index=False, header=True):
                ws.append(r)

            print(f"已创建工作表: {sheet_name}")

        # 添加借款总结工作表
        if self.loans:
            loan_all = self.loans[-1]  # 获取最后一笔借款
            loan_all.loan_id = loan_all.loan_id + 1
            loan_all.name = "借款总结"
            loan_table = loan_all.get_output_table(self.all_years)

            # 设置工作表名称
            sheet_name = f"借款{loan_all.loan_id}"

            # 检查工作表是否已存在，如果存在则删除
            if sheet_name in book.sheetnames:
                book.remove(book[sheet_name])

            # 创建新工作表
            ws = book.create_sheet(sheet_name)

            # 将DataFrame写入工作表
            for r in dataframe_to_rows(loan_table, index=False, header=True):
                ws.append(r)

            print(f"已创建借款总结工作表: {sheet_name}")

        # 保存工作簿
        try:
            book.save(file_path)
            print(f"结果已成功导出到: {os.path.abspath(file_path)}")
            print(f"包含以下工作表: {[f'借款{loan.loan_id}' for loan in self.loans]} 和 借款{loan_all.loan_id}")
        except Exception as e:
            print(f"保存工作簿时出错: {e}")
            return

        return len(self.all_years), self.all_years[-1]



def read_financing_info(file_path, sheet_name):
    """
    从 Excel 读取融资信息，自动识别表头所在行
    参数:
        file_path: Excel 文件路径
        sheet_name: 工作表名称
    返回:
        字典列表，每个字典代表一笔借款
    """
    required_columns = [
        '序号', '借款名称', '借款类别', '借款金额\n（万元）', '开始时间', '结束时间',
        '借款周期（年）', '借款利率', '还款方式', '首年计息月份',
        '债券发行费（万元）', '债券发行登记服务费\n（万元）',
        '债券还本付息兑付手续费\n（万元）'
    ]

    try:
        # 1）先把整张表读进来，不指定 header
        raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

        # 2）逐行查找"序号"所在行（忽略 NaN）
        header_idx = None
        for idx, row in raw.iterrows():
            # 把 NaN 过滤掉，再转字符串，防止类型不一致
            row_vals = [str(v).strip() for v in row.dropna().values]
            if '序号' in row_vals:
                header_idx = idx
                break

        if header_idx is None:
            raise ValueError("未在工作表中找到'序号'列，无法确定表头位置")

        # 3）用这一行作为表头重新读取
        df = pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            header=header_idx  # 真正的表头行
        )

        # 4）列名统一去掉前后空格，防止隐藏空格导致匹配失败
        df.columns = df.columns.astype(str).str.strip()

        # 5）检查缺失列
        missing = [c for c in required_columns if c not in df.columns]
        if missing:
            raise ValueError(f"Excel 缺少必要列: {', '.join(missing)}")

        # 6）后续解析逻辑保持不变
        loans = []
        for _, row in df.iterrows():
            try:
                # 跳过空行（序号为NaN的行）
                if pd.isna(row['序号']):
                    continue

                # 处理首年计息月份
                first_interest_month = row['首年计息月份']
                if pd.isna(first_interest_month):
                    first_interest_month = 1
                else:
                    # 确保转换为整数
                    first_interest_month = int(float(str(first_interest_month).strip()))

                # 处理日期字段
                def parse_date(d):
                    if pd.isna(d):
                        # 返回默认日期或处理缺失日期的方式
                        return (dt.now().date() - dt(1899, 12, 30).date()).days

                    # 如果是字符串格式的日期
                    if isinstance(d, str) and "年" in d:
                        try:
                            # 处理"2023年1月"格式
                            date_obj = dt.strptime(d.strip(), "%Y年%m月")
                            return (date_obj.date() - dt(1899, 12, 30).date()).days
                        except ValueError:
                            # 如果格式不匹配，返回当前日期
                            return (dt.now().date() - dt(1899, 12, 30).date()).days

                    # 如果是Excel日期数字
                    try:
                        return float(d)
                    except (ValueError, TypeError):
                        # 如果转换失败，返回当前日期
                        return (dt.now().date() - dt(1899, 12, 30).date()).days

                # 创建贷款信息字典
                loan = {
                    '序号': int(float(row['序号'])),  # 先转浮点再转整，避免NaN问题
                    '借款名称': str(row['借款名称']).strip() if not pd.isna(row['借款名称']) else "",
                    '借款类别': str(row['借款类别']).strip() if not pd.isna(row['借款类别']) else "",
                    '借款金额（万元）': float(row['借款金额\n（万元）']) if not pd.isna(row['借款金额\n（万元）']) else 0.0,
                    '开始时间': parse_date(row['开始时间']),
                    '结束时间': parse_date(row['结束时间']),
                    '借款周期（年）': int(float(row['借款周期（年）'])) if not pd.isna(row['借款周期（年）']) else 0,
                    '借款利率': float(row['借款利率']) if not pd.isna(row['借款利率']) else 0.0,
                    '还款方式': str(row['还款方式']).strip() if not pd.isna(row['还款方式']) else "",
                    '首年计息月份': first_interest_month,
                    '债券发行费（万元）': float(row['债券发行费（万元）']) if not pd.isna(row['债券发行费（万元）']) else 0.0,
                    '债券发行登记服务费（万元）': float(row['债券发行登记服务费\n（万元）']) if not pd.isna(
                        row['债券发行登记服务费\n（万元）']) else 0.0,
                    '债券还本付息兑付手续费（万元）': float(row['债券还本付息兑付手续费\n（万元）']) if not pd.isna(
                        row['债券还本付息兑付手续费\n（万元）']) else 0.0
                }
                loans.append(loan)
            except Exception as e:
                print(f"行解析失败: {e}，跳过")
                continue

        return loans

    except Exception as e:
        print(f"读取融资信息失败: {e}")
        return []


def format_existing_excel(file_path, sheet_name, output_path=None):
    """
    修改现有Excel文件的格式：微软雅黑字体、居中对齐、全边框

    参数:
    file_path: 原始Excel文件路径
    output_path: 输出文件路径（如果为None，则覆盖原文件）
    """
    # 如果没有指定输出路径，则覆盖原文件
    if output_path is None:
        output_path = file_path

    # 加载工作簿
    workbook = load_workbook(file_path)
    # 处理所有工作表
    sheet = workbook[sheet_name]

    # 设置微软雅黑字体和居中对齐
    font = Font(name='微软雅黑', size=11)
    alignment = Alignment(horizontal='center', vertical='center')

    # 设置全边框
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 应用到所有有数据的单元格
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:  # 只处理有内容的单元格
                cell.font = font
                cell.alignment = alignment
                cell.border = border

    # 设置矢车菊蓝背景 (RGB: 100, 149, 237)
    cornflower_light_fill = PatternFill(
        start_color="ADC6E6",  # 矢车菊蓝浅色60%的近似十六进制代码
        end_color="ADC6E6",
        fill_type="solid"
    )

    # 设置红色字体
    red_font = Font(name='微软雅黑', size=11, color="FF0000", bold=True)

    # 设置居中对齐
    alignment = Alignment(horizontal='center', vertical='center')

    # 设置全边框
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 获取第二行中的所有单元格
    second_row = 2  # Excel行号从1开始，所以第二行是2

    # 检查第二行是否有数据
    if sheet.max_row >= second_row:
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=second_row, column=col)

            # 应用格式
            cell.fill = cornflower_light_fill
            cell.font = red_font
            cell.alignment = alignment
            cell.border = border

    # 可选：保持其他行的格式不变，或者应用基本格式
    # 这里我们保持其他行不变，只修改第二行

    # 自动调整列宽以适应内容
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        # # 设置列宽，留出一些边距
        # adjusted_width = min(max_length + 2, 50)  # 限制最大列宽为50
        # sheet.column_dimensions[column_letter].width = adjusted_width

    # 保存工作簿
    workbook.save(output_path)
    print(f"Excel格式已成功更新并保存到: {output_path}")


def table_C1(input_path, sheet_name):
    """
    把表c的值回填到表C.1中
    :param input_path:  文件路径
    :param sheet_name:  表C.1的名称
    :return:
    """

    # 获取表c的值，用于回填表C.1
    value = "还本付息兑付手续费"
    table_c_sheet = "c借款还本付息计划表"
    address = find_all_cells(input_path, value, table_c_sheet)  # 返回：[(row1, col1), (row2, col2), ...]
    loop_num = len(address) - 1
    start_address = address[0]
    start_address = get_column_letter(column_index_from_string(start_address[1])+1) + str(start_address[0])
    # 获取填入表C.1的位置
    value2 = "债券还本付息兑付手续费\n（万元）"
    r, c = find_cell(input_path, value2, sheet_name)
    target_address = c + str(r+2)

    formula_generator = ExcelFormulaGenerator(
        data_file=input_path,  # 包含所有数据的单一文件
        output_file=input_path  # 输出到同一个文件
    )

    config = [
        {
            "operation": "custom",
            "formula_template": "=({A_val})",
            "params": {
                "A_val": {"sheet": table_c_sheet, "cell": start_address},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": target_address  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 9, "col_shift": 0},  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 1, "col_shift": 0}  # 每次向右移动一列且向下移动一行
            }
        }
    ]

    formula_generator.generate_formulas(config)

def loan_fill(input_file):
    # 输入文件路径
    # input_file = input("请输入融资信息Excel文件路径: ").strip()
    # input_file = "财务分析套表自做模板编程用.xlsm"
    # input_file = "财务分析套表自做模板_test.xlsx"
    sheet_name = "C.1项目融资信息"

    # 验证文件存在
    if not os.path.exists(input_file):
        print(f"文件不存在: {input_file}")
        exit()

    # 输出文件路径
    # output_file = input("请输入输出文件路径（默认为: 借款还本付息计划表.xlsx）: ").strip()
    # if not output_file:
    #     output_file = "借款还本付息计划表.xlsx"

    output_file = input_file

    # 创建还款系统
    system = LoanRepaymentSystem()

    # 从Excel读取融资信息
    financing_data = read_financing_info(input_file, sheet_name)

    if not financing_data:
        print("未读取到融资信息，请检查文件格式和内容")
        exit()

    print(f"成功读取 {len(financing_data)} 笔借款信息")

    # 添加借款
    for loan_data in financing_data:
        print("借款信息loan_data为：", loan_data)
        system.add_loan(loan_data)
        print(f"添加借款: {loan_data['序号']} {loan_data['借款名称']} ({loan_data['借款金额（万元）']}万元)")

    # 生成并导出还款计划
    year_num, last_year = system.export_to_excel(output_file)

    print("借款还本付息计划表生成完成！")
    print("表头年份数量一共有：", year_num)
    print("最后一年的年份为：", last_year)

    for loan_data in financing_data:
        write_loan(loan_data, output_file)

    # 将生成的各模板表c’ 复制到一个sheet中,然后删除被复制的sheet
    source = output_file  # 替换为源Excel文件路径
    target = output_file  # 替换为目标Excel文件路径
    target_sheet = "c借款还本付息计划表"  # 替换为目标sheet名称
    sheets_to_copy = [f"借款{i + 1}" for i in range(len(financing_data) + 1)]
    print("sheets_to_copy:", sheets_to_copy)
    copy_and_delete_sheets(source, target, target_sheet, sheets_to_copy, True)

    # 填充借款总结的内容
    loan_summary(output_file, target_sheet, year_num)

    # 修改格式
    # format_existing_excel(output_file, target_sheet)

    table_C1_sheet = "C.1项目融资信息"
    # 回填C.1表
    table_C1(input_file, table_C1_sheet)

    return last_year


# 主程序
if __name__ == "__main__":
    # # 输入文件路径
    # # input_file = input("请输入融资信息Excel文件路径: ").strip()
    # input_file = "财务分析套表自做模板编程用.xlsm"
    # # input_file = "财务分析套表自做模板_test.xlsx"
    # sheet_name = "C.1项目融资信息"
    #
    # # 验证文件存在
    # if not os.path.exists(input_file):
    #     print(f"文件不存在: {input_file}")
    #     exit()
    #
    # # 输出文件路径
    # # output_file = input("请输入输出文件路径（默认为: 借款还本付息计划表.xlsx）: ").strip()
    # # if not output_file:
    # #     output_file = "借款还本付息计划表.xlsx"
    #
    # output_file = input_file
    #
    # # 创建还款系统
    # system = LoanRepaymentSystem()
    #
    # # 从Excel读取融资信息
    # financing_data = read_financing_info(input_file, sheet_name)
    #
    # if not financing_data:
    #     print("未读取到融资信息，请检查文件格式和内容")
    #     exit()
    #
    # print(f"成功读取 {len(financing_data)} 笔借款信息")
    #
    # # 添加借款
    # for loan_data in financing_data:
    #     print("借款信息loan_data为：", loan_data)
    #     system.add_loan(loan_data)
    #     print(f"添加借款: {loan_data['序号']} {loan_data['借款名称']} ({loan_data['借款金额（万元）']}万元)")
    #
    # # 生成并导出还款计划
    # year_num, last_year = system.export_to_excel(output_file)
    #
    # print("借款还本付息计划表生成完成！")
    # print("表头年份数量一共有：", year_num)
    # print("最后一年的年份为：", last_year)
    #
    # for loan_data in financing_data:
    #     write_loan(loan_data, input_file)
    #
    # # 将生成的各模板表c’ 复制到一个sheet中,然后删除被复制的sheet
    # source = input_file  # 替换为源Excel文件路径
    # target = input_file  # 替换为目标Excel文件路径
    # target_sheet = "c借款还本付息计划表"  # 替换为目标sheet名称
    # sheets_to_copy = [f"借款{i + 1}" for i in range(len(financing_data) + 1)]
    # print("sheets_to_copy:", sheets_to_copy)
    # copy_and_delete_sheets(source, target, target_sheet, sheets_to_copy, True)
    #
    #
    # # 填充借款总结的内容
    # loan_summary(input_file, target_sheet, year_num)
    #
    # # 修改格式
    # format_existing_excel(input_file, target_sheet)

    input_file = "财务分析套表自做模板编程用.xlsm"
    last_year = loan_fill(input_file)
    print(last_year)
    print("done")
