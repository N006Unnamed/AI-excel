from circulate_formula import ExcelFormulaGenerator
from copy_formula import expand_excel_header, find_last_used_column, ExcelAutoFiller, find_last_used_row, copy_cell_style
from openpyxl.utils import get_column_letter, column_index_from_string
from findAndSet import find_cell, find_all_cells
import openpyxl
from loan_assignment import right_n_cells, excel_date_to_year, down_n_cells, up_n_cells
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, Protection
from loan_assignment import struct_years, to_absolute_address
from openpyxl_vba import load_workbook



def struct_years_all(file_path, table_b_sheet):
    """
    获取所有建设期的开始年份和结束年份
    :param input_path:
    :param target_year:
    :return:
    """

    # 建设期的时间范围
    start_value = "建设开始年月"
    end_value = "建设完成年月"
    # 加载工作簿
    wb = load_workbook(file_path)
    ws = wb[table_b_sheet]

    start_years_address = find_all_cells(file_path, start_value, table_b_sheet)  # [(row1, col1), (row2, col2), ...]
    end_years_address = find_all_cells(file_path, end_value, table_b_sheet)  # [(row1, col1), (row2, col2), ...]
    # 建设期开始年月单元格地址列表
    start = []
    start_years = []
    # 建设期完成年月单元格地址列表
    end = []
    end_years = []
    for start_year_address in start_years_address:
        start.append(right_n_cells(start_year_address[1] + str(start_year_address[0]), 2))
    for end_year_address in end_years_address:
        end.append(right_n_cells(end_year_address[1] + str(end_year_address[0]), 2))

    # 求出每个 建设期 的 运营期(年)
    operating_period = []
    for i in range(len(end)):
        operating_period.append(down_n_cells(end[i], 3))
    # 根据单元格地址求出其值所代表的年份
    print("start:", start)
    print("end:", end)
    # 遍历所有单元格地址
    for cell_address in start:
        # 获取单元格值
        cell = ws[cell_address]
        start_years.append(cell.value)

    for cell_address in end:
        # 获取单元格值
        cell = ws[cell_address]
        end_years.append(cell.value)

    start_years = excel_date_to_year(start_years)
    end_years = excel_date_to_year(end_years)

    print("start_years:", start_years)
    print("end_years:", end_years)

    results = []

    for i in range(len(start_years)):
        result = {
            'start_year':start_years[i],
            'end_year':end_years[i]
        }
        results.append(result)

    return results, operating_period

def operation_years(file_path, table_B_sheet):
    """
    找出运营期最早开始的年份
    :param file_path:
    :param table_B_sheet:
    :return:
    """
    # 运营期开始的最早时间
    start_value = "运营开始年月"
    # 加载工作簿
    wb = load_workbook(file_path)
    ws = wb[table_B_sheet]

    start_years_address = find_all_cells(file_path, start_value, table_B_sheet)  # [(row1, col1), (row2, col2), ...]

    # 运营期开始年月单元格地址列表
    start = []
    start_years = []
    # 运营期开始年月单元格地址列表
    for start_year_address in start_years_address:
        start.append(right_n_cells(start_year_address[1] + str(start_year_address[0]), 2))
    # 根据单元格地址求出其值所代表的年份
    print("start:", start)
    # 遍历所有单元格地址
    for cell_address in start:
        # 获取单元格值
        cell = ws[cell_address]
        start_years.append(cell.value)

    start_years = excel_date_to_year(start_years)

    print("start_years:", start_years)

    # 找出最早的年份
    min_year = min(start_years)

    return min_year


def table_D2(input_path, sheet_name, last_col):
    """
    填充D.2表，1.1建设投资，对应D.3的“总计”行
             1.2建设期利息，只填充建设期，对应c表的n.5+n.6；只填充建设期，不要全部填充
             2.2债务资金，对应C.2的“4合计”行
    :param input_path: 文件路径
    :param sheet_name: D.2表的名称
    :param last_col: D.2表扩充表头后的最后一列
    :return:
    """

    # 1.1建设投资 第一个地址
    target_address = "D4"

    # D.3表 总计 的第一个地址
    table_D3_sheet = "D.3建设投资分年计划表"
    r = find_last_used_row(input_path, table_D3_sheet)
    start_address = "E" + str(r)

    # 循环次数
    loop_num = column_index_from_string(last_col) - column_index_from_string("D") + 1

    # 建设期期间
    table_B_sheet = "B项目信息"
    start_year, end_year = struct_years(input_path, table_B_sheet)

    # 建设期循环次数
    struct_loop = end_year - start_year + 1

    # 1.2建设期利息 第一个地址
    struct_target = "D5"

    # 找到表c的 n.5 和 n.6
    table_c_sheet = "c借款还本付息计划表"
    c_last_r = find_last_used_row(input_path, table_c_sheet)
    c_address1 = "D" + str(c_last_r)
    c_address2 = "D" + str(c_last_r - 1)

    # 2.2债务资金 第一个地址
    target_address2 = "D12"

    # C.2 表 合计 的第一个地址
    table_C2_sheet = "C.2项目每年借款信息"
    value = "合计"
    r3, c3 = find_cell(input_path, value, table_C2_sheet)
    c3 = get_column_letter(column_index_from_string(c3)+1)
    C2_start_address = c3 + str(r3)

    # D.2 合计
    sum_address = "C3"
    sum_address1 = "D3"
    sum_address2 = last_col + str(3)
    sum_loop = 14

    formula_generator = ExcelFormulaGenerator(
        data_file=input_path,  # 包含所有数据的单一文件
        output_file=input_path  # 输出到同一个文件
    )

    config = [
        # 1.1 建设投资
        {
            "operation": "custom",
            "formula_template": "=({A_val})",
            "params": {
                "A_val": {"sheet": table_D3_sheet, "cell": start_address},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": target_address  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        },
        # 1.2 建设期利息
        {
            "operation": "custom",
            "formula_template": "=({A_val}+{B_val})",
            "params": {
                "A_val": {"sheet": table_c_sheet, "cell": c_address1},  # A表的y (向右移动)
                "B_val": {"sheet": table_c_sheet, "cell": c_address2}  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": struct_target  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": struct_loop,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                    "B_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        },
        # 2.2债务资金
        {
            "operation": "custom",
            "formula_template": "=({A_val})",
            "params": {
                "A_val": {"sheet": table_C2_sheet, "cell": C2_start_address},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": target_address2  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        },

        # 合计
        {
            "operation": "custom",
            "formula_template": "=(SUM({A_val}:{B_val}))",
            "params": {
                "A_val": {"sheet": '', "cell": sum_address1},  # A表的y (向右移动)
                "B_val": {"sheet": '', "cell": sum_address2},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": sum_address  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": sum_loop,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 1, "col_shift": 0},  # 每次向下移动一行
                    "B_val": {"row_shift": 1, "col_shift": 0}  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 1, "col_shift": 0}  # 每次向右移动一列且向下移动一行
            }
        }
    ]

    formula_generator.generate_formulas(config)


def table_D4(input_path, sheet_name, last_col):
    """
    表D.4填充，从第一次运营时间开始填充，之前都留空
             1.1应收账款，对应F表的5年度合计(含税)
             1.2.3在产品，没有就填0
    :param input_path:
    :param sheet_name:
    :return:
    """

    # 最早开始的运营期
    # table_B_sheet = "B项目信息"
    # start_year = operation_years(input_path, table_B_sheet)
    #
    # _, c = find_cell(input_path, start_year, sheet_name)

    # 1.1 应收账款
    table_F_sheet = "F项目收入"
    table_F_value = "年度合计（含税）"
    r, c = find_cell(input_path, table_F_value, table_F_sheet)
    c = get_column_letter(column_index_from_string(c)+2)
    table_F_address = c + str(r)

    loop_num = column_index_from_string(last_col) - column_index_from_string("E")

    config = [
        # 营业收入（不含增值税）
        {
            "operation": "custom",
            "formula_template": "=({A_val}/$D$4)",
            "params": {
                "A_val": {"sheet": table_F_sheet, "cell": table_F_address},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": "E4"  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        }
    ]

    formula_generator = ExcelFormulaGenerator(
        data_file=input_path,  # 包含所有数据的单一文件
        output_file=input_path  # 输出到同一个文件
    )

    formula_generator.generate_formulas(config)


    print("done")


# "5.2 其他管理费用"，未完成
def table_E(input_path, sheet_name, last_col):
    """
    表E的填充，5.2，公式中对应c表n.4还本付息兑付手续费
             7折旧费，对应E.2的“折旧额合计”行
             8摊销费，对应E.3的“折旧额合计”行
             9利息支出，对应c的n.4+n.6，且从建设期结束开始填充(建设期的利息算在折旧里)
    :param input_path:
    :param sheet_name:
    :return:
    """

    # 7 折旧费
    table_E2_sheet = "E.2固定资产折旧费估算表"
    table_E2_value = "折旧额合计"
    r, _ = find_cell(input_path, table_E2_value, table_E2_sheet)
    table_E2_address = "F" + str(r)

    # 8 摊销费
    table_E3_sheet = "E.3无形资产和其他资产摊销费估算表"
    table_E3_value = "折旧额合计"
    r, _ = find_cell(input_path, table_E3_value, table_E3_sheet)
    table_E3_address = "F" + str(r)

    # 9 利息支出
    # 建设期 结束时间
    table_B_sheet = "B项目信息"
    table_c_sheet = "c借款还本付息计划表"
    _, end_year = struct_years(input_path, table_B_sheet)
    # 利息支出 起始位置
    _, c = find_cell(input_path, end_year + 1, sheet_name)
    target_address = c + str(14)
    # 表c 还本付息兑付手续费 和 应付利息 的起始位置
    _, c = find_cell(input_path, end_year + 1, table_c_sheet)
    r = find_last_used_row(input_path, table_c_sheet)
    table_c_address1 = c + str(r)
    table_c_address2 = c + str(r - 2)


    # 循环次数
    loop_num = column_index_from_string(last_col) - column_index_from_string("C") + 1


    config = [
        # 折旧费
        {
            "operation": "custom",
            "formula_template": "=({A_val})",
            "params": {
                "A_val": {"sheet": table_E2_sheet, "cell": table_E2_address},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": "C12"  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        },

        # 摊销费
        {
            "operation": "custom",
            "formula_template": "=({A_val})",
            "params": {
                "A_val": {"sheet": table_E3_sheet, "cell": table_E3_address},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": "C13"  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        },

        # 利息支出
        {
            "operation": "custom",
            "formula_template": "=({A_val}+{B_val})",
            "params": {
                "A_val": {"sheet": table_c_sheet, "cell": table_c_address1},  # A表的y (向右移动)
                "B_val": {"sheet": table_c_sheet, "cell": table_c_address2},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": target_address  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                    "B_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        },

    ]

    formula_generator = ExcelFormulaGenerator(
        data_file=input_path,  # 包含所有数据的单一文件
        output_file=input_path  # 输出到同一个文件
    )

    formula_generator.generate_formulas(config)


    print("done")



def table_E2(input_path, sheet_name, last_col):
    """
    表E.2填充，C列公式，对E.4的引用要包含对应的建设年份;
             同时，包含c表中对应建设年份的n.5+n.6
             D列，对应B表中，各建设期的“运营期(年)”，即1.5、2.5等
    :param input_path:
    :param sheet_name:
    :return:
    """
    # 加载现有工作簿
    wb = load_workbook(input_path)
    ws = wb[sheet_name]

    # 确定数据开始的行和列
    # 假设表头在第3行，数据从第4行开始
    start_row = 4
    year_columns = []  # 存储年份列的索引

    # # 查找年份列
    # for col in range(6, ws.max_column + 1):
    #     if ws.cell(row=3, column=col).value and isinstance(ws.cell(row=3, column=col).value, (int, str)):
    #         year_columns.append(col)
    #         # 只保留前4个年份列
    #         if len(year_columns) >= 4:
    #             break

    # 如果没有找到年份列，使用默认列
    if not year_columns:
        last_col_num = column_index_from_string(last_col)
        year_columns = list(range(6, last_col_num + 1))   # 生成 [6, 7, 8, 9, 10] F, G, H, I, J列

    # 清除现有数据行（从第4行开始）
    for row in range(start_row, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).value = None

    # 获取一共有多少 建设期 和各个建设期的始末年份
    value = "原值"
    table_B_sheet = "B项目信息"
    r, c = find_cell(input_path, value, sheet_name)
    results, operating_period = struct_years_all(input_path, table_B_sheet)

    c_next = get_column_letter(column_index_from_string(c) + 1)
    c_next_next = get_column_letter(column_index_from_string(c) + 2)

    r1 = 0
    r2 = r+2

    # 添加建设期行
    current_row = start_row
    for i in range(len(results)):
        r1 = r2 + i
        result = results[i]
        formula = generate_formula(input_path, result['start_year'], result['end_year'], tax=False)
        temp_r, temp_c = find_cell(input_path, result['end_year'], sheet_name)
        end_year_address = temp_c + str(temp_r)
        end_year_address = to_absolute_address(end_year_address)

        # 原值 目标单元格
        target_address = c + str(r1)
        # 年折旧率 目标单元格
        target_address2 = c_next + str(r1)

        ws.cell(row=current_row, column=1).value = f'建设期{i + 1}'  # 序号
        ws.cell(row=current_row, column=2).value = '固定资产投资'  # 项目
        ws.cell(row=current_row, column=3).value = formula  # 原值
        ws.cell(row=current_row, column=4).value = f"=1/{table_B_sheet}!{to_absolute_address(operating_period[i])}*100%"  # 年折旧率
        ws.cell(row=current_row, column=5).value = f"={target_address}*(1-A财务假设!$D$12)*{target_address2}"  # 年折旧额

        # 设置年份列为0
        for j, col_idx in enumerate(year_columns):
            current_year = get_column_letter(col_idx) + str(3)
            orign_value = get_column_letter(3) + str(current_row)
            orign_value = to_absolute_address(orign_value)
            annual_amount = get_column_letter(5) + str(current_row)
            annual_amount = to_absolute_address(annual_amount)
            ws.cell(row=current_row, column=col_idx).value = f"=IF({current_year}-{end_year_address}>0, {orign_value}-{annual_amount}*({current_year}-{end_year_address}), 0)"

        current_row += 1

    # 添加净值合计行
    ws.cell(row=current_row, column=1).value = '净值合计'  # 序号
    ws.cell(row=current_row, column=2).value = '固定资产投资'  # 项目
    ws.cell(row=current_row, column=3).value = f"=SUM({c + str(r2)}:{c + str(r1)})"  # 原值
    ws.cell(row=current_row, column=4).value = '/'  # 年折旧率
    ws.cell(row=current_row, column=5).value = f"=SUM({c_next_next + str(r2)}:{c_next_next + str(r1)})"  # 年折旧额


    # 设置年份列为0
    for j, col_idx in enumerate(year_columns):
        orign_c = get_column_letter(col_idx)
        ws.cell(row=current_row, column=col_idx).value = f"=SUM({orign_c + str(r2)}:{orign_c + str(r1)})"

    current_row += 2
    r1 += 2
    r2 += 2

    # 添加折旧额行
    for i in range(len(results)):

        result = results[i]
        temp_r, temp_c = find_cell(input_path, result['end_year'], sheet_name)
        end_year_address = temp_c + str(temp_r)
        end_year_address = to_absolute_address(end_year_address)

        ws.cell(row=current_row, column=1).value = f'建设期{i + 1}'  # 序号
        ws.cell(row=current_row, column=2).value = '固定资产投资'  # 项目
        ws.cell(row=current_row, column=3).value = ''  # 原值
        ws.cell(row=current_row, column=4).value = ''  # 年折旧率
        ws.cell(row=current_row, column=5).value = ''  # 年折旧额

        annual_r = 4 + i

        # 设置年份列为0
        for j, col_idx in enumerate(year_columns):
            address1 = to_absolute_address("E" + str(annual_r))
            address2 = get_column_letter(col_idx) + str(annual_r)
            current_year = get_column_letter(col_idx) + str(3)
            ws.cell(row=current_row, column=col_idx).value = f'=IF({current_year}-{end_year_address}>0, IF({address2}<>"",{address1}, 0),0 )'

        current_row += 1
        r1 += 1
        r2 += 1

    # 添加折旧额合计行
    ws.cell(row=current_row, column=1).value = '折旧额合计'  # 序号
    ws.cell(row=current_row, column=2).value = '固定资产投资'  # 项目
    ws.cell(row=current_row, column=3).value = ''  # 原值
    ws.cell(row=current_row, column=4).value = ''  # 年折旧率
    annual_c = get_column_letter(5)
    ws.cell(row=current_row, column=5).value = f"=SUM({annual_c + str(r2)}:{annual_c + str(r1)})"  # 年折旧额

    # 设置年份列为0
    for j, col_idx in enumerate(year_columns):
        annual_c = get_column_letter(col_idx)
        ws.cell(row=current_row, column=col_idx).value = f"=SUM({annual_c + str(r2)}:{annual_c + str(r1)})"

    # 保存修改后的文件
    wb.save(input_path)
    print(f"文件已更新: {input_path}")

    return len(results)



def generate_formula(input_path, start_year, end_year, tax=True):
    """
    生成Excel公式，引用建设期内所有年份的单元格。
    start_year: 建设期起始年份（例如2025）
    end_year: 建设期结束年份（例如2026或2027）
    返回: Excel公式字符串
    """
    if start_year < 2025:
        start_year = 2025  # 假设建设期从2025年开始
    if end_year < start_year:
        return "="  # 无效年份范围

    ref_list = []
    r_list_E4 = []
    r_list_c = []
    table_D3_sheet = "D.3建设投资分年计划表"
    table_E4_sheet = "E.4建设投资税后金额表"
    table_c_sheet = "c借款还本付息计划表"

    values = ["工程类费用", "硬件类", "工程建设其他费用", "预备费"]
    for value in values:
        r, _ = find_cell(input_path, value, table_D3_sheet)
        r_list_E4.append(r)

    r = find_last_used_row(input_path, table_c_sheet)
    r_list_c.append(r)
    r_list_c.append(r-1)

    # 处理E.4建设投资税后金额表
    for year in range(start_year, end_year + 1):

        if tax:
            year_value = year
        else:
            year_value = str(year)+"\n（税后）"
        _, c = find_cell(input_path, year_value, table_E4_sheet)
        for row in r_list_E4:
            ref_list.append(f"{table_E4_sheet}!{c}{row}")

    # 处理c借款还本付息计划表
    for year in range(start_year, end_year + 1):

        # year_value = str(year)
        _, c = find_cell(input_path, year, table_c_sheet)
        for row in r_list_c:  # 注意行顺序：先47后46
            ref_list.append(f"{table_c_sheet}!{c}{row}")

    formula = "=" + "+".join(ref_list)
    return formula


def generate_formula_E3(input_path, start_year, end_year, tax=True):
    """
    生成Excel公式，引用建设期内所有年份的单元格。
    start_year: 建设期起始年份（例如2025）
    end_year: 建设期结束年份（例如2026或2027）
    返回: Excel公式字符串
    """
    if start_year < 2025:
        start_year = 2025  # 假设建设期从2025年开始
    if end_year < start_year:
        return "="  # 无效年份范围

    ref_list = []

    r_list_E3 = []
    table_E4_sheet = "D.3建设投资分年计划表"

    values = ["软件类", "安全类", "数据类"]
    for value in values:
        r, _ = find_cell(input_path, value, table_E4_sheet)
        r_list_E3.append(r)

    # 处理E.4建设投资税后金额表
    for year in range(start_year, end_year + 1):
        table_E4_sheet = "E.4建设投资税后金额表"
        if tax:
            year_value = year
        else:
            year_value = str(year)+"\n（税后）"
        _, c = find_cell(input_path, year_value, table_E4_sheet)
        for row in r_list_E3:
            ref_list.append(f"{table_E4_sheet}!{c}{row}")


    formula = "=" + "+".join(ref_list)
    return formula


def table_E3(input_path, sheet_name, last_col):
    """
    清空并重新创建无形资产折旧表

    参数:
    filename: str, Excel文件名
    construction_periods: int, 建设期的数量
    """
    # 加载现有工作簿
    wb = load_workbook(input_path)
    ws = wb[sheet_name]

    # 保存表头和格式信息
    header_row = 3  # 表头在第3行
    header_values = []
    header_styles = []

    # 获取表头值和样式
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col)
        header_values.append(cell.value)
        header_styles.append({
            'fill': cell.fill,
            'font': cell.font,
            'alignment': cell.alignment,
            'border': cell.border
        })

    # 确定年份列
    last_col_num = column_index_from_string(last_col)
    year_columns = list(range(6, last_col_num + 1))  # 生成 [6, 7, 8, 9, 10] F, G, H, I, J列

    # for col in range(1, len(header_values) + 1):
    #     if header_values[col - 1] and isinstance(header_values[col - 1], (int, str)) and (
    #             isinstance(header_values[col - 1], int) or
    #             (isinstance(header_values[col - 1], str) and header_values[col - 1].isdigit())
    #     ):
    #         year_columns.append(col)

    # # 只保留前4个年份列
    # year_columns = year_columns[:4]

    # 清空工作表（保留表头）
    for row in range(header_row + 1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).value = None


    table_B_sheet = "B项目信息"
    results, operating_period = struct_years_all(input_path, table_B_sheet)


    # 重新创建表格内容
    current_row = header_row + 1

    temp = 0

    sum_list = []
    r_list = []
    # ********************************************* 净值合计 上半部分 ************************************************
    # 为每个建设期创建行
    for period in range(1, len(results) + 1):

        # 求出当前 建设期的完成年份
        result = results[period-1]
        temp_r, temp_c = find_cell(input_path, result['start_year'], sheet_name)
        start_year_address = temp_c + str(temp_r)
        start_year_address = to_absolute_address(start_year_address)

        temp_r, temp_c = find_cell(input_path, result['end_year'], sheet_name)
        end_year_address = temp_c + str(temp_r)
        end_year_address = to_absolute_address(end_year_address)

        operate_date = operating_period[period-1]
        operate_date = to_absolute_address(operate_date)

        formula = generate_formula_E3(input_path, result['start_year'], result['end_year'], tax=False)

        # 建设期合计行
        ws.cell(row=current_row, column=1).value = f'建设期{period}'
        ws.cell(row=current_row, column=2).value = f'合计{period}'

        address1 = get_column_letter(3) + str(current_row + 1)
        address2 = get_column_letter(3) + str(current_row + 5)
        ws.cell(row=current_row, column=3).value = f'={address1}+{address2}'  # 原值

        # 折旧额合计
        sum_list.append(current_row)

        ws.cell(row=current_row, column=4).value = '/'  # 年折旧率
        address1 = get_column_letter(5) + str(current_row + 1)
        address2 = get_column_letter(5) + str(current_row + 5)
        ws.cell(row=current_row, column=5).value = f'={address1}+{address2}'  # 年折旧额

        # 设置年份列为0
        for col_idx in year_columns:
            address1 = get_column_letter(col_idx) + str(current_row + 1)
            address2 = get_column_letter(col_idx) + str(current_row + 5)
            ws.cell(row=current_row, column=col_idx).value = f'={address1}+{address2}'

        current_row += 1

        # 无形资产行
        ws.cell(row=current_row, column=1).value = f'{period}.1'  # 空序号
        ws.cell(row=current_row, column=2).value = '无形资产'
        address1 = get_column_letter(3) + str(current_row + 1)
        address2 = get_column_letter(3) + str(current_row + 3)
        ws.cell(row=current_row, column=3).value = f'=SUM({address1}:{address2})'  # 原值
        ws.cell(row=current_row, column=4).value = '/'  # 年折旧率
        address1 = get_column_letter(5) + str(current_row + 1)
        address2 = get_column_letter(5) + str(current_row + 3)
        ws.cell(row=current_row, column=5).value = f'=SUM({address1}:{address2})'  # 年折旧额

        # 设置年份列为0
        for col_idx in year_columns:
            address1 = get_column_letter(col_idx) + str(current_row + 1)
            address2 = get_column_letter(col_idx) + str(current_row + 3)
            ws.cell(row=current_row, column=col_idx).value = f'=SUM({address1}:{address2})'

        current_row += 1

        # 土地使用权行
        r_list.append(current_row)
        ws.cell(row=current_row, column=1).value = f'{period}.1.1'  # 空序号
        ws.cell(row=current_row, column=2).value = '土地使用权'
        ws.cell(row=current_row, column=3).value = 0  # 原值

        ws.cell(row=current_row, column=4).value = f'=1/{table_B_sheet}!{operate_date}*100%'  # 年折旧率
        address1 = get_column_letter(3) + str(current_row)
        address2 = get_column_letter(4) + str(current_row)
        ws.cell(row=current_row, column=5).value = f'={address1}*{address2}'  # 年折旧额

        # 设置年份列为0
        for col_idx in year_columns:
            current_year = get_column_letter(col_idx) + str(3)
            address1 = get_column_letter(3) + str(current_row)
            address1 = to_absolute_address(address1)
            address2 = get_column_letter(5) + str(current_row)
            address2 = to_absolute_address(address2)
            ws.cell(row=current_row, column=col_idx).value = f'=IF({current_year}-{end_year_address}>0, {address1}-{address2}*({current_year}-{end_year_address}), {temp})'


        current_row += 1

        # 专利技术行
        ws.cell(row=current_row, column=1).value = f'{period}.1.2'  # 空序号
        ws.cell(row=current_row, column=2).value = '专利技术'
        ws.cell(row=current_row, column=3).value = 0  # 原值
        ws.cell(row=current_row, column=4).value = f'=1/{table_B_sheet}!{operate_date}*100%'  # 年折旧率
        address1 = get_column_letter(3) + str(current_row)
        address2 = get_column_letter(4) + str(current_row)
        ws.cell(row=current_row, column=5).value = f'={address1}*{address2}'  # 年折旧额

        # 设置年份列为0
        for col_idx in year_columns:
            current_year = get_column_letter(col_idx) + str(3)
            address1 = get_column_letter(3) + str(current_row)
            address1 = to_absolute_address(address1)
            address2 = get_column_letter(5) + str(current_row)
            address2 = to_absolute_address(address2)
            ws.cell(row=current_row, column=col_idx).value = f'=IF({current_year}-{end_year_address}>0, {address1}-{address2}*({current_year}-{end_year_address}), {temp})'

        current_row += 1

        # 软件及其他无形资产行
        ws.cell(row=current_row, column=1).value = f'{period}.1.3'  # 空序号
        ws.cell(row=current_row, column=2).value = '软件及其他无形资产'
        ws.cell(row=current_row, column=3).value = formula  # 原值
        ws.cell(row=current_row, column=4).value = f'=1/{table_B_sheet}!{operate_date}*100%'  # 年折旧率
        address1 = get_column_letter(3) + str(current_row)
        address2 = get_column_letter(4) + str(current_row)
        ws.cell(row=current_row, column=5).value = f'={address1}*{address2}'  # 年折旧额

        # 设置年份列为0
        for col_idx in year_columns:
            current_year = get_column_letter(col_idx) + str(3)
            address1 = get_column_letter(3) + str(current_row)
            address1 = to_absolute_address(address1)
            address2 = get_column_letter(5) + str(current_row)
            address2 = to_absolute_address(address2)
            ws.cell(row=current_row, column=col_idx).value = f'=IF({current_year}-{end_year_address}>0, {address1}-{address2}*({current_year}-{end_year_address}), {temp})'

        current_row += 1

        # 其他资产行
        ws.cell(row=current_row, column=1).value = f'{period}.2'  # 空序号
        ws.cell(row=current_row, column=2).value = '其他资产'
        address1 = get_column_letter(3) + str(current_row + 1)
        ws.cell(row=current_row, column=3).value = f'={address1}'  # 原值
        ws.cell(row=current_row, column=4).value = '/'  # 年折旧率
        address1 = get_column_letter(5) + str(current_row + 1)
        ws.cell(row=current_row, column=5).value = f'={address1}'  # 年折旧额

        # 设置年份列为0
        for col_idx in year_columns:
            address1 = get_column_letter(col_idx) + str(current_row + 1)
            ws.cell(row=current_row, column=col_idx).value = f'={address1}'

        current_row += 1

        # 项目开办费行
        ws.cell(row=current_row, column=1).value = f'{period}.2.1'  # 空序号
        ws.cell(row=current_row, column=2).value = '项目开办费'
        ws.cell(row=current_row, column=3).value = 0  # 原值
        ws.cell(row=current_row, column=4).value = f'=1/{table_B_sheet}!{operate_date}*100%'  # 年折旧率
        address1 = get_column_letter(3) + str(current_row)
        address2 = get_column_letter(4) + str(current_row)
        ws.cell(row=current_row, column=5).value = f'={address1}*{address2}'  # 年折旧额

        # 设置年份列为0
        for col_idx in year_columns:
            current_year = get_column_letter(col_idx) + str(3)
            address1 = get_column_letter(3) + str(current_row)
            address1 = to_absolute_address(address1)
            address2 = get_column_letter(5) + str(current_row)
            address2 = to_absolute_address(address2)
            ws.cell(row=current_row, column=col_idx).value = f'=IF({current_year}-{end_year_address}>0, {address1}-{address2}*({current_year}-{end_year_address}), {temp})'

        current_row += 1

    # 添加折旧额合计行

    ws.cell(row=current_row, column=2).value = '净值合计'
    sum_address_list = []
    for sum in sum_list:
        sum_address = get_column_letter(3)+str(sum)
        sum_address_list.append(sum_address)
    ws.cell(row=current_row, column=3).value = f'={'+'.join(sum_address_list)}'  # 原值
    ws.cell(row=current_row, column=4).value = '/'  # 年折旧率
    sum_address_list = []
    for sum in sum_list:
        sum_address = get_column_letter(5)+str(sum)
        sum_address_list.append(sum_address)
    ws.cell(row=current_row, column=5).value = f'={'+'.join(sum_address_list)}'  # 年折旧额

    # 设置年份列为0
    for col_idx in year_columns:
        sum_address_list = []
        for sum in sum_list:
            sum_address = get_column_letter(col_idx) + str(sum)
            sum_address_list.append(sum_address)
        ws.cell(row=current_row, column=col_idx).value = f'={'+'.join(sum_address_list)}'

    current_row += 2

    sum_list = []

    # ********************************************* 折旧额合计 下半部分 ***************************************
    # 为每个建设期创建行
    for period in range(1, len(results) + 1):

        # 求出当前 建设期的完成年份
        result = results[period-1]
        temp_r, temp_c = find_cell(input_path, result['start_year'], sheet_name)
        start_year_address = temp_c + str(temp_r)
        start_year_address = to_absolute_address(start_year_address)

        temp_r, temp_c = find_cell(input_path, result['end_year'], sheet_name)
        end_year_address = temp_c + str(temp_r)
        end_year_address = to_absolute_address(end_year_address)

        operate_date = operating_period[period-1]
        operate_date = to_absolute_address(operate_date)

        formula = generate_formula_E3(input_path, result['start_year'], result['end_year'], tax=False)

        # 建设期合计行
        ws.cell(row=current_row, column=1).value = f'建设期{period}'
        ws.cell(row=current_row, column=2).value = f'合计{period}'

        # address1 = get_column_letter(3) + str(current_row + 1)
        # address2 = get_column_letter(3) + str(current_row + 5)
        # ws.cell(row=current_row, column=3).value = f'={address1}+{address2}'  # 原值
        ws.cell(row=current_row, column=3).value = ''  # 原值

        # 折旧额合计
        sum_list.append(current_row)

        # ws.cell(row=current_row, column=4).value = '/'  # 年折旧率
        ws.cell(row=current_row, column=4).value = ''  # 年折旧率
        # address1 = get_column_letter(5) + str(current_row + 1)
        # address2 = get_column_letter(5) + str(current_row + 5)
        # ws.cell(row=current_row, column=5).value = f'={address1}+{address2}'  # 年折旧额
        ws.cell(row=current_row, column=5).value = ''  # 年折旧额

        # 设置年份列为0
        for col_idx in year_columns:
            address1 = get_column_letter(col_idx) + str(current_row + 1)
            address2 = get_column_letter(col_idx) + str(current_row + 5)
            ws.cell(row=current_row, column=col_idx).value = f'={address1}+{address2}'

        current_row += 1

        # 无形资产行
        ws.cell(row=current_row, column=1).value = f'{period}.1'  # 空序号
        ws.cell(row=current_row, column=2).value = '无形资产'
        # address1 = get_column_letter(3) + str(current_row + 1)
        # address2 = get_column_letter(3) + str(current_row + 3)
        # ws.cell(row=current_row, column=3).value = f'=SUM({address1}:{address2})'  # 原值
        ws.cell(row=current_row, column=3).value = ''  # 原值
        # ws.cell(row=current_row, column=4).value = '/'  # 年折旧率
        ws.cell(row=current_row, column=4).value = ''  # 年折旧率
        # address1 = get_column_letter(5) + str(current_row + 1)
        # address2 = get_column_letter(5) + str(current_row + 3)
        # ws.cell(row=current_row, column=5).value = f'=SUM({address1}:{address2})'  # 年折旧额
        ws.cell(row=current_row, column=5).value = ''  # 年折旧额

        # 设置年份列为0
        for col_idx in year_columns:
            address1 = get_column_letter(col_idx) + str(current_row + 1)
            address2 = get_column_letter(col_idx) + str(current_row + 3)
            ws.cell(row=current_row, column=col_idx).value = f'=SUM({address1}:{address2})'

        current_row += 1

        # 土地使用权行
        ws.cell(row=current_row, column=1).value = f'{period}.1.1'  # 空序号
        ws.cell(row=current_row, column=2).value = '土地使用权'
        # ws.cell(row=current_row, column=3).value = 0  # 原值
        ws.cell(row=current_row, column=3).value = ''  # 原值

        # ws.cell(row=current_row, column=4).value = f'=1/{table_B_sheet}!{operate_date}*100%'  # 年折旧率
        ws.cell(row=current_row, column=4).value = ''  # 年折旧率
        # address1 = get_column_letter(3) + str(current_row)
        # address2 = get_column_letter(4) + str(current_row)
        # ws.cell(row=current_row, column=5).value = f'={address1}*{address2}'  # 年折旧额
        ws.cell(row=current_row, column=5).value = ''  # 年折旧额

        r = r_list[period - 1]

        # 设置年份列为0
        for col_idx in year_columns:
            current_year = get_column_letter(col_idx) + str(3)
            address1 = get_column_letter(col_idx) + str(r)
            address2 = get_column_letter(5) + str(r)
            address2 = to_absolute_address(address2)
            ws.cell(row=current_row, column=col_idx).value = f'=IF({current_year}-{end_year_address}>0, IF({address1}<>"",{address2},0), {temp})'


        current_row += 1

        # 专利技术行
        ws.cell(row=current_row, column=1).value = f'{period}.1.2'  # 空序号
        ws.cell(row=current_row, column=2).value = '专利技术'
        # ws.cell(row=current_row, column=3).value = 0  # 原值
        ws.cell(row=current_row, column=3).value = ''  # 原值
        # ws.cell(row=current_row, column=4).value = f'=1/{table_B_sheet}!{operate_date}*100%'  # 年折旧率
        ws.cell(row=current_row, column=4).value = ''  # 年折旧率
        # address1 = get_column_letter(3) + str(current_row)
        # address2 = get_column_letter(4) + str(current_row)
        # ws.cell(row=current_row, column=5).value = f'={address1}*{address2}'  # 年折旧额
        ws.cell(row=current_row, column=5).value = ''  # 年折旧额

        r += 1

        # 设置年份列为0
        for col_idx in year_columns:
            current_year = get_column_letter(col_idx) + str(3)
            address1 = get_column_letter(col_idx) + str(r)
            address2 = get_column_letter(5) + str(r)
            address2 = to_absolute_address(address2)
            ws.cell(row=current_row, column=col_idx).value = f'=IF({current_year}-{end_year_address}>0, IF({address1}<>"",{address2},0), {temp})'

        current_row += 1

        # 软件及其他无形资产行
        ws.cell(row=current_row, column=1).value = f'{period}.1.3'  # 空序号
        ws.cell(row=current_row, column=2).value = '软件及其他无形资产'
        # ws.cell(row=current_row, column=3).value = formula  # 原值
        ws.cell(row=current_row, column=3).value = ''  # 原值
        # ws.cell(row=current_row, column=4).value = f'=1/{table_B_sheet}!{operate_date}*100%'  # 年折旧率
        ws.cell(row=current_row, column=4).value = ''  # 年折旧率
        # address1 = get_column_letter(3) + str(current_row)
        # address2 = get_column_letter(4) + str(current_row)
        # ws.cell(row=current_row, column=5).value = f'={address1}*{address2}'  # 年折旧额
        ws.cell(row=current_row, column=5).value = ''  # 年折旧额

        r += 1

        # 设置年份列为0
        for col_idx in year_columns:
            current_year = get_column_letter(col_idx) + str(3)
            address1 = get_column_letter(col_idx) + str(r)
            address2 = get_column_letter(5) + str(r)
            address2 = to_absolute_address(address2)
            ws.cell(row=current_row, column=col_idx).value = f'=IF({current_year}-{end_year_address}>0, IF({address1}<>"",{address2},0), {temp})'

        current_row += 1

        # 其他资产行
        ws.cell(row=current_row, column=1).value = f'{period}.2'  # 空序号
        ws.cell(row=current_row, column=2).value = '其他资产'
        # address1 = get_column_letter(3) + str(current_row + 1)
        # ws.cell(row=current_row, column=3).value = f'={address1}'  # 原值
        ws.cell(row=current_row, column=3).value = ''  # 原值
        # ws.cell(row=current_row, column=4).value = '/'  # 年折旧率
        ws.cell(row=current_row, column=4).value = ''  # 年折旧率
        # address1 = get_column_letter(5) + str(current_row + 1)
        # ws.cell(row=current_row, column=5).value = f'={address1}'  # 年折旧额
        ws.cell(row=current_row, column=5).value = ''  # 年折旧额

        # 设置年份列为0
        for col_idx in year_columns:
            address1 = get_column_letter(col_idx) + str(current_row + 1)
            ws.cell(row=current_row, column=col_idx).value = f'={address1}'

        current_row += 1

        # 项目开办费行
        ws.cell(row=current_row, column=1).value = f'{period}.2.1'  # 空序号
        ws.cell(row=current_row, column=2).value = '项目开办费'
        # ws.cell(row=current_row, column=3).value = 0  # 原值
        ws.cell(row=current_row, column=3).value = ''  # 原值
        # ws.cell(row=current_row, column=4).value = f'=1/{table_B_sheet}!{operate_date}*100%'  # 年折旧率
        ws.cell(row=current_row, column=4).value = ''  # 年折旧率
        # address1 = get_column_letter(3) + str(current_row)
        # address2 = get_column_letter(4) + str(current_row)
        # ws.cell(row=current_row, column=5).value = f'={address1}*{address2}'  # 年折旧额
        ws.cell(row=current_row, column=5).value = ''  # 年折旧额

        r += 2

        # 设置年份列为0
        for col_idx in year_columns:
            current_year = get_column_letter(col_idx) + str(3)
            address1 = get_column_letter(col_idx) + str(r)
            address2 = get_column_letter(5) + str(r)
            address2 = to_absolute_address(address2)
            ws.cell(row=current_row, column=col_idx).value = f'=IF({current_year}-{end_year_address}>0, IF({address1}<>"",{address2},0), {temp})'

        current_row += 1

    # 添加折旧额合计行

    ws.cell(row=current_row, column=2).value = '折旧额合计'
    sum_address_list = []
    # for sum in sum_list:
    #     sum_address = get_column_letter(3)+str(sum)
    #     sum_address_list.append(sum_address)
    # ws.cell(row=current_row, column=3).value = f'={'+'.join(sum_address_list)}'  # 原值
    ws.cell(row=current_row, column=3).value = ''  # 原值
    # ws.cell(row=current_row, column=4).value = '/'  # 年折旧率
    ws.cell(row=current_row, column=4).value = ''  # 年折旧率
    # sum_address_list = []
    # for sum in sum_list:
    #     sum_address = get_column_letter(5)+str(sum)
    #     sum_address_list.append(sum_address)
    # ws.cell(row=current_row, column=5).value = f'={'+'.join(sum_address_list)}'  # 年折旧额
    ws.cell(row=current_row, column=5).value = ''  # 年折旧额

    # 设置年份列为0
    for col_idx in year_columns:
        sum_address_list = []
        for sum in sum_list:
            sum_address = get_column_letter(col_idx) + str(sum)
            sum_address_list.append(sum_address)
        ws.cell(row=current_row, column=col_idx).value = f'={'+'.join(sum_address_list)}'


    # # 应用基本格式
    # apply_basic_formatting(ws, header_row, current_row, len(header_values))

    # 保存修改后的文件
    wb.save(input_path)
    print(f"文件已重新创建: {input_path}")



def apply_basic_formatting(ws, header_row, last_row, num_columns):
    """
    应用基本格式到工作表

    参数:
    ws: 工作表对象
    header_row: int, 表头行号
    last_row: int, 最后一行行号
    num_columns: int, 列数
    """
    # 设置表头样式
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_font = Font(bold=True)
    center_aligned = Alignment(horizontal="center")

    for col in range(1, num_columns + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_aligned

    # 设置边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in range(header_row, last_row + 1):
        for col in range(1, num_columns + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            if col in [3, 4, 5]:  # 原值、年折旧率、年折旧额列
                cell.alignment = center_aligned

    # 调整列宽
    for col in range(1, num_columns + 1):
        max_length = 0
        for row in range(header_row, last_row + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                try:
                    length = len(str(cell_value))
                    if length > max_length:
                        max_length = length
                except:
                    pass
        adjusted_width = min(max_length + 2, 15)
        ws.column_dimensions[get_column_letter(col)].width = adjusted_width


def table_F(input_path, sheet_name, last_col):
    """
    表F 的填充，5和6年度合计，分别累加上面分项的含税和不含税值。
             若有“补贴收入”，需单独加上。
    :param input_path:
    :param sheet_name:
    :return:
    """

    # 所有 含税 的单元格地址
    tax_value = "含税"
    # 年度合计（含税） 单元格地址
    tax_target_value = "年度合计（含税）"
    table_F_tool(input_path, sheet_name, last_col, tax_value, tax_target_value)
    print("年度合计（含税）计算完成")

    # 所有 不含税 的单元格地址
    no_tax_value = "不含税"
    # 年度合计（不含税） 单元格地址
    no_tax_target_value = "年度合计（不含税）"
    table_F_tool(input_path, sheet_name, last_col, no_tax_value, no_tax_target_value)
    print("年度合计（不含税）计算完成")


def table_F_tool(input_path, sheet_name,last_col, value, target_value):
    """
    表F填充的工具，用于生成单元格公式参数量不定的情况。
    :param input_path:
    :param sheet_name:
    :param last_col:
    :param value:
    :param target_value:
    :return:
    """

    tax_address = find_all_cells(input_path, value, sheet_name)
    all_tax_address = []
    for i in range(len(tax_address)):
        temp_address = tax_address[i]
        c = get_column_letter(column_index_from_string(temp_address[1]) + 1)
        r = temp_address[0]
        all_tax_address.append(c+str(r))

    r, c = find_cell(input_path, target_value, sheet_name)
    c = get_column_letter(column_index_from_string(c)+2)
    tax_target_address = c + str(r)

    loop_num = column_index_from_string(last_col) - column_index_from_string(c)


    # 1. 生成参数名 B_val, C_val, D_val ...
    param_names = [f"{chr(66 + i)}_val" for i in range(len(all_tax_address))]
    # 2. 生成 params 区块
    params = {
        name: {"sheet": "", "cell": cell}
        for name, cell in zip(param_names, all_tax_address)
    }
    # 3. 生成公式模板  =({B_val}+{C_val}+...)
    template = "=(" + "+".join([f"{{{p}}}" for p in param_names]) + ")"

    params_offsets = {
        name: {"row_shift": 0, "col_shift": 1}
        for name in param_names
    }

    loop_config = {
        "count": loop_num,  # 循环3次

        # 参数独立偏移设置
        "param_offsets": params_offsets,

        # 目标单元格偏移设置
        "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
    }

    summary_config = [{
        "operation": "custom",
        "formula_template": template,
        "params": params,
        "target": {"sheet": sheet_name, "cell": tax_target_address},
        "loop": loop_config
    }]

    print("summary_config:", summary_config)

    formula_generator = ExcelFormulaGenerator(
        data_file=input_path,  # 包含所有数据的单一文件
        output_file=input_path  # 输出到同一个文件
    )

    formula_generator.generate_formulas(summary_config)


def table_G(input_path, sheet_name, last_col):
    """
    表G填充， D3单元格，对应G.1中F列最后一行
            2当期销项税额，下面的子项，对应F表中的不含税价行(纵向序号不是连续，而是+2)
    :param input_path:
    :param sheet_name:
    :param last_col:
    :return:
    """

    table_G1 = "G.1进项增值税率-建设投资"
    value = "增值税（万元）"
    _, c = find_cell(input_path, value, table_G1)
    r = find_last_used_row(input_path, table_G1)
    G1_address = c + str(r)

    # 当期进项税额 第一个单元格地址
    value1 = "当期进项税额"
    r1, c1 = find_cell(input_path, value1, sheet_name)
    c1 = get_column_letter(column_index_from_string(c1)+2)
    address1 = c1 + str(r1)

    # 当期销项税额 第一个单元格地址
    value1 = "当期销项税额"
    r1, c1 = find_cell(input_path, value1, sheet_name)
    c1 = get_column_letter(column_index_from_string(c1)+2)
    address2 = c1 + str(r1)

    # 当期应缴增值税  第一个单元格地址
    value1 = "当期应缴增值税"
    r1, c1 = find_cell(input_path, value1, sheet_name)
    c1 = get_column_letter(column_index_from_string(c1)+2)
    address3 = c1 + str(r1)

    # 城建税及教育费附加  第一个单元格地址
    value2 = "城建税及教育费附加"
    r1, c1 = find_cell(input_path, value2, sheet_name)
    c1 = get_column_letter(column_index_from_string(c1)+2)
    address4 = c1 + str(r1)

    # 合计  第一个单元格地址
    value1 = "合计"
    r1, c1 = find_cell(input_path, value1, sheet_name)
    c1 = get_column_letter(column_index_from_string(c1) + 2)
    address5 = c1 + str(r1)

    loop_num = column_index_from_string(last_col) - column_index_from_string(c1)


    formula_generator = ExcelFormulaGenerator(
        data_file=input_path,  # 包含所有数据的单一文件
        output_file=input_path  # 输出到同一个文件
    )

    config = [
        # "D3"
        {
            "operation": "custom",
            "formula_template": "=({A_val})",
            "params": {
                "A_val": {"sheet": table_G1, "cell": G1_address},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": "D3"  # 目标起始位置 (向右移动)
            }
        },
        # "E3" 及之后
        {
            "operation": "custom",
            "formula_template": "=(IF({A_val}=0,ABS({B_val}-{C_val}-{D_val}),0))",
            "params": {
                "A_val": {"sheet": "", "cell": address3},  # A表的y (向右移动)
                "B_val": {"sheet": "", "cell": address2},  # A表的y (向右移动)
                "C_val": {"sheet": "", "cell": address1},  # A表的y (向右移动)
                "D_val": {"sheet": "", "cell": "D3"}  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": "E3"  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num-1,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                    "B_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                    "C_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                    "D_val": {"row_shift": 0, "col_shift": 1}  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        },

        # 当期进项税额
        {
            "operation": "custom",
            "formula_template": "=(SUM({A_val}:{B_val}))",
            "params": {
                "A_val": {"sheet": "", "cell": down_n_cells(address1, 1)},  # A表的y (向右移动)
                "B_val": {"sheet": "", "cell": up_n_cells(address2, 1)}  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": address1  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                    "B_val": {"row_shift": 0, "col_shift": 1}  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        },

        # 当期销项税额
        {
            "operation": "custom",
            "formula_template": "=(SUM({A_val}:{B_val}))",
            "params": {
                "A_val": {"sheet": "", "cell": down_n_cells(address2, 1)},  # A表的y (向右移动)
                "B_val": {"sheet": "", "cell": up_n_cells(address3, 1)}  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": address2  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                    "B_val": {"row_shift": 0, "col_shift": 1}  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        },

        # 当期应缴增值税
        {
            "operation": "custom",
            "formula_template": "=(MAX({A_val}-{B_val}-{C_val},0))",
            "params": {
                "A_val": {"sheet": "", "cell": right_n_cells(address2, 1)},  # A表的y (向右移动)
                "B_val": {"sheet": "", "cell": right_n_cells(address1, 1)},  # A表的y (向右移动)
                "C_val": {"sheet": "", "cell": "E3"}  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": right_n_cells(address3, 1)  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num - 1,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                    "B_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                    "C_val": {"row_shift": 0, "col_shift": 1}  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        },

        # 城建税及教育费附加
        {
            "operation": "custom",
            "formula_template": "=({A_val}*(A财务假设!$D$6+A财务假设!$D$7+A财务假设!$D$8))",
            "params": {
                "A_val": {"sheet": "", "cell": right_n_cells(address3, 1)}  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": right_n_cells(address4, 1)  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num - 1,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1}  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        },

        # 合计
        {
            "operation": "custom",
            "formula_template": "=({A_val}+{B_val})",
            "params": {
                "A_val": {"sheet": "", "cell": address3},  # A表的y (向右移动)
                "B_val": {"sheet": "", "cell": address3},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": address5  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                    "B_val": {"row_shift": 0, "col_shift": 1}  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        },


    ]

    formula_generator.generate_formulas(config)



def table_format(input_path, sheet_name):
    """
    整理excel表格的格式
    :param input_path:
    :param sheet_name:
    :return:
    """

    wb = load_workbook(input_path)
    ws = wb[sheet_name]

    # 创建居中对齐样式
    alignment_center = Alignment(horizontal='center', vertical='center')
    # 设置全边框
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    # 设置指定区域单元格居中对齐
    for row in ws:
        for cell in row:
            cell.alignment = alignment_center
            cell.border = border

    wb.save(input_path)
    wb.close()


def copy_one(input_path, target_year):

    # 用户输入
    output_path = input_path
    header_row = 2
    sheet_names = ["G.税金及附加测算表", "E总成本费用估算表", "D.4流动资金估算表", "D.2项目总投资使用计划与资金筹措表",
                    "F项目收入", "F.1项目单项收入信息"]
    # sheet_names = ["G.税金及附加测算表", "E总成本费用估算表", "D.4流动资金估算表", "D.2项目总投资使用计划与资金筹措表",
    #                 "F项目收入"]


    for sheet_name in sheet_names:
        _, start_col = find_last_used_column(input_path, sheet_name)
        start_col_plus = column_index_from_string(start_col) + 1
        start_col_plus = get_column_letter(start_col_plus)
        # start_col = "H"
        print(start_col)
        print(start_col_plus)

        # 执行扩展
        try:
            expand_excel_header(
                input_file=input_path,
                sheet_name=sheet_name,
                output_file=output_path,
                target_year=target_year,
                header_row=header_row,
                start_col=start_col
            )
        except Exception as e:
            print(f"\n错误: {str(e)}")
            print("操作未完成，请检查输入参数是否正确")


        # if sheet_name == "D.2项目总投资使用计划与资金筹措表":
        #     table_D2(input_path, sheet_name, start_col)
        if sheet_name == "F项目收入":
            table_F(input_path, sheet_name, start_col)
        elif sheet_name == "G.税金及附加测算表":
            table_G(input_path, sheet_name, start_col)
        elif sheet_name == "D.4流动资金估算表":
            table_D4(input_path, sheet_name, start_col)
        elif sheet_name == "E总成本费用估算表":
            table_E(input_path, sheet_name, start_col)

        r, c = find_cell(input_path, target_year, sheet_name)

        filler = ExcelAutoFiller(input_path)
        filler.set_active_sheet(sheet_name)


        last_row = find_last_used_row(input_path, sheet_name)

        for i in range(r + 1, last_row + 1):
            filler.auto_fill(f"{start_col}{i}", f"{start_col_plus}{i}:{c}{i}")

        if sheet_name == "D.2项目总投资使用计划与资金筹措表":
            table_D2(input_path, sheet_name, c)
        # elif sheet_name == "F项目收入":
        #     table_F(input_path, sheet_name, c)
        # elif sheet_name == "G.税金及附加测算表":
        #     table_G(input_path, sheet_name, c)
        # elif sheet_name == "D.4流动资金估算表":
        #     table_D4(input_path, sheet_name, c)
        # elif sheet_name == "E总成本费用估算表":
        #     table_E(input_path, sheet_name, c)

    # print("将 G6 的公式填充到 J6:K6...")
    # filler.auto_fill("G6", "J6:K6")



def copy_two(input_path, target_year):

    # 用户输入
    output_path = input_path
    header_row = 2
    sheet_names = ["E.1项目运营期费用", "E.1.1项目运营期费用(不含税）"]
    # sheet_names = ["E.1.1项目运营期费用(不含税）"]

    last_row = 0
    c = ""

    for sheet_name in sheet_names:
        _, start_col = find_last_used_column(input_path, sheet_name)
        start_col_plus = column_index_from_string(start_col) + 1
        start_col_plus = get_column_letter(start_col_plus)

        start_col_minus = column_index_from_string(start_col) - 1
        start_col_minus = get_column_letter(start_col_minus)
        # start_col = "H"
        print(start_col)
        print(start_col_plus)

        # 执行扩展
        try:
            expand_excel_header(
                input_file=input_path,
                sheet_name=sheet_name,
                output_file=output_path,
                target_year=target_year,
                header_row=header_row,
                start_col=start_col_minus
            )
        except Exception as e:
            print(f"\n错误: {str(e)}")
            print("操作未完成，请检查输入参数是否正确")

        r, c = find_cell(input_path, target_year, sheet_name)

        filler = ExcelAutoFiller(input_path)
        filler.set_active_sheet(sheet_name)

        last_row = find_last_used_row(input_path, sheet_name) + 2

        for i in range(r + 1, last_row):
            filler.auto_fill(f"{start_col_minus}{i}", f"{start_col}{i}:{c}{i}")

        # 增加 "合计" 项
        c2 = column_index_from_string(c) + 1
        wb = load_workbook(input_path)
        ws = wb[sheet_name]
        ws.cell(row=r, column=c2).value = "合计"
        ws.cell(row=r+1, column=c2).value = "/"
        wb.save(input_path)

        # "合计" 列的公式起始地址
        sum_start_row = r + 2
        sum_start_col = get_column_letter(c2)
        target_address = sum_start_col + str(sum_start_row)

        # SUM公式的起始与结束地址
        value = "负荷系数"
        _, c3 = find_cell(input_path, value, sheet_name)
        start_address = c3 + str(sum_start_row)
        end_address = c + str(sum_start_row)
        if sheet_name == "E.1.1项目运营期费用(不含税）":
            loop_num = last_row - sum_start_row
        else:
            loop_num = last_row - sum_start_row - 1


        formula_generator = ExcelFormulaGenerator(
            data_file=input_path,  # 包含所有数据的单一文件
            output_file=input_path  # 输出到同一个文件
        )

        config = [
            {
                "operation": "custom",
                "formula_template": "=(SUM({A_val}:{B_val}))",
                "params": {
                    "A_val": {"sheet": "", "cell": start_address},  # A表的y (向右移动)
                    "B_val": {"sheet": "", "cell": end_address},  # A表的y (向右移动)

                },
                "target": {
                    "sheet": sheet_name,  # 结果工作表
                    "cell": target_address  # 目标起始位置 (向右移动)
                },
                "loop": {
                    "count": loop_num,  # 循环3次

                    # 参数独立偏移设置
                    "param_offsets": {
                        "A_val": {"row_shift": 1, "col_shift": 0},  # 每次向下移动一行
                        "B_val": {"row_shift": 1, "col_shift": 0},  # 每次向下移动一行
                    },

                    # 目标单元格偏移设置
                    "target_offset": {"row_shift": 1, "col_shift": 0}  # 每次向右移动一列且向下移动一行
                }
            }
        ]

        formula_generator.generate_formulas(config)

        wb = load_workbook(input_path)
        ws = wb[sheet_name]


        # 更新格式
        for index in range(header_row, last_row + 1):
            source_cell_address = c + str(index)
            target_cell_address = sum_start_col + str(index)
            source_cell = ws[source_cell_address]
            target_cell = ws[target_cell_address]
            copy_cell_style(source_cell, target_cell)

        wb.save(input_path)
        wb.close()







        # if sheet_name == "b利润与利润分配表（损益和利润分配表）":
        #     table_b(input_path, sheet_name, c)
        # elif sheet_name == "d财务计划现金流量表":
        #     table_d(input_path, sheet_name, c)
        # else:
        #     table_e(input_path, sheet_name, c)

    # print("将 G6 的公式填充到 J6:K6...")
    # filler.auto_fill("G6", "J6:K6")



def copy_three(input_path, target_year):

    # 用户输入
    output_path = input_path
    header_row = 3
    sheet_names = ["E.2固定资产折旧费估算表", "E.3无形资产和其他资产摊销费估算表"]


    for sheet_name in sheet_names:
        _, start_col = find_last_used_column(input_path, sheet_name)
        start_col_plus = column_index_from_string(start_col) + 1
        start_col_plus = get_column_letter(start_col_plus)
        # start_col = "H"
        print(start_col)
        print(start_col_plus)
        loop = 0

        # 执行扩展
        try:
            expand_excel_header(
                input_file=input_path,
                sheet_name=sheet_name,
                output_file=output_path,
                target_year=target_year,
                header_row=header_row,
                start_col=start_col
            )
        except Exception as e:
            print(f"\n错误: {str(e)}")
            print("操作未完成，请检查输入参数是否正确")

        if sheet_name == "E.2固定资产折旧费估算表":
            loop = table_E2(input_path, sheet_name, start_col)
        elif sheet_name == "E.3无形资产和其他资产摊销费估算表":
            table_E3(input_path, sheet_name, start_col)

        r, c = find_cell(input_path, target_year, sheet_name)

        filler = ExcelAutoFiller(input_path)
        filler.set_active_sheet(sheet_name)


        last_row = find_last_used_row(input_path, sheet_name)

        for i in range(r + 1, last_row + 1):
            filler.auto_fill(f"{start_col}{i}", f"{start_col_plus}{i}:{c}{i}")

        # 补充 建设期 项
        if sheet_name == "E.2固定资产折旧费估算表":
            wb = load_workbook(input_path)
            ws = wb[sheet_name]

            last_row_E2 = find_last_used_row(input_path, sheet_name)
            target_address_E2 = "E" + str(last_row_E2)
            start_address_E2 = "F" + str(last_row_E2)
            end_addess_E2 = c + str(last_row_E2)

            for i in range(loop):
                target_address_E2 = up_n_cells(target_address_E2, 1)
                start_address_E2 = up_n_cells(start_address_E2, 1)
                end_addess_E2 = up_n_cells(end_addess_E2, 1)
                ws[target_address_E2] = f"=SUM({start_address_E2}:{end_addess_E2})"

            wb.save(input_path)







if __name__ == "__main__":

    input_path = "财务分析套表自做模板编程用.xlsm"
    # 年份行号在第二行
    sheet_names = ["D.2项目总投资使用计划与资金筹措表", "D.4流动资金估算表",
                   "E总成本费用估算表", "F项目收入", "G.税金及附加测算表"]
    # 年份行号在第三行 (后放)
    sheet_names2 = ["E.2固定资产折旧费估算表", "E.3无形资产和其他资产摊销费估算表"]
    # 不能直接复制
    sheet_name3 = ["E.1项目运营期费用", "E.1.1项目运营期费用(不含税）", "F.1项目单项收入信息"]

    sheet_name = ""
    target_year = 2047
    header_row = 2
    header_row2 = 3

    # start_col = ""

    # for sheet_name in sheet_names:
    #     print(sheet_name)
    #     start_col = find_last_used_column(input_path, sheet_name)
    #     expand_excel_header(input_path, sheet_name, input_path, target_year, header_row)

    # for sheet_name in sheet_names2:
    #     print(sheet_name)
    #     start_col = find_last_used_column(input_path, sheet_name)
    #     expand_excel_header(input_path, sheet_name, input_path, target_year, header_row2)
    #
    # copy_one(input_path, target_year)
    # copy_two(input_path, target_year)

    # results, _ = struct_years_all(input_path, "B项目信息")
    #
    # print(results)
    #
    # for result in results:
    #     print(result)
    #     print(result['start_year'])
    #     print(result['end_year'])
    #     formula = generate_formula(input_path, result['start_year'], result['end_year'])
    #     print(formula)
    #
    # table_E2(input_path, "E.2固定资产折旧费估算表")

    # copy_one(input_path, target_year)
    copy_two(input_path, target_year)
    # copy_three(input_path, target_year)

    # 假设从另一个表中获取到建设期数量为2
    construction_periods = 2

    # 更新现有的Excel文件
    # table_E2(input_path, "E.2固定资产折旧费估算表", 'K')
    # table_E3(input_path, "E.3无形资产和其他资产摊销费估算表", 'K')

    print("done")
