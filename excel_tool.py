import openpyxl
from openpyxl.utils import get_column_letter


def calculate_excel_data(file_path, sheet_name, source_col, start_row, end_row, target_cell):
    """
    读取Excel中指定列的数据，计算后写入目标单元格

    参数:
    file_path (str): Excel文件路径
    sheet_name (str): 工作表名称
    source_col (int): 数据来源列号（从1开始）
    start_row (int): 数据起始行号
    end_row (int): 数据结束行号
    target_cell (str): 目标单元格地址（如"A1"）
    """
    # 加载工作簿和工作表
    wb = load_workbook(file_path)
    sheet = wb[sheet_name]

    total = 0  # 初始化计算结果
    valid_values = []  # 存储有效数值

    # 遍历指定数据范围
    for row in range(start_row, end_row + 1):
        # 构建单元格地址（如"C2"）
        cell_address = f"{get_column_letter(source_col)}{row}"
        cell = sheet[cell_address]

        # 检查单元格值是否为数字
        if isinstance(cell.value, (int, float)):
            total += cell.value
            valid_values.append(cell.value)

    # 输出计算信息（调试用）
    print(f"计算列 {get_column_letter(source_col)} 行 {start_row}-{end_row}：")
    print(f"有效值: {valid_values}")
    print(f"求和结果: {total}")

    # 将结果写入目标单元格
    sheet[target_cell] = total

    # 保存修改后的工作簿（原始文件会被覆盖）
    wb.save(file_path)
    print(f"结果已写入单元格 {target_cell}，文件保存成功！")


# 示例使用
if __name__ == "__main__":
    # 参数配置（根据实际需求修改）
    EXCEL_PATH = "test2.xlsx"  # Excel文件路径
    SHEET_NAME = "Sheet1"  # 工作表名称
    SOURCE_COL = 3  # 数据来源列（第3列）
    START_ROW = 2  # 数据起始行（第2行）
    END_ROW = 8  # 数据结束行（第8行）
    TARGET_CELL = "C1"  # 目标单元格（第3列第1行）

    # 执行计算
    calculate_excel_data(
        file_path=EXCEL_PATH,
        sheet_name=SHEET_NAME,
        source_col=SOURCE_COL,
        start_row=START_ROW,
        end_row=END_ROW,
        target_cell=TARGET_CELL
    )
