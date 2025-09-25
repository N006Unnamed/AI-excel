from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.worksheet.datavalidation import DataValidationList

from circulate_formula import ExcelFormulaGenerator
import datetime
from findAndSet import find_cell, find_all_cells, find_all_coords_co
import re
from copy import copy
from openpyxl.utils import get_column_letter, column_index_from_string
import openpyxl
import os
from loan_function import repay_method_cal
import string
from typing import List, Dict, Any
from openpyxl_vba import load_workbook

def right_n_cells(addr: str, n: int) -> str:
    """
    addr: 形如 "D5" 的单元格地址
    n   : 向右偏移量
    return: 偏移后的地址，例如 "I5"
    """
    match = re.match(r"([A-Za-z]+)(\d+)", addr)
    if not match:
        raise ValueError(f"无效的单元格地址格式: {addr}")

    col_str = match.group(1)  # 列字母部分（可能包含多个字母）
    row_str = match.group(2)  # 行数字部分

    col_idx = column_index_from_string(col_str)  # 字母 → 数字
    new_col_idx = col_idx + n  # 向右移动 n 格
    new_col = get_column_letter(new_col_idx)  # 数字 → 字母
    return f"{new_col}{row_str}"

def left_n_cells(addr: str, n: int) -> str:
    """
    addr: 形如 "D5" 的单元格地址
    n   : 向左偏移量
    return: 偏移后的地址，例如 "C5"
    """
    match = re.match(r"([A-Za-z]+)(\d+)", addr)
    if not match:
        raise ValueError(f"无效的单元格地址格式: {addr}")

    col_str = match.group(1)  # 列字母部分（可能包含多个字母）
    row_str = match.group(2)  # 行数字部分

    col_idx = column_index_from_string(col_str)  # 字母 → 数字
    new_col_idx = col_idx - n  # 向右移动 n 格
    new_col = get_column_letter(new_col_idx)  # 数字 → 字母
    return f"{new_col}{row_str}"


def down_n_cells(addr: str, n: int) -> str:
    """
    改进版本：处理包含多字母列的 Excel 地址

    addr: 形如 "D5"、"AA24"、"AB123" 的单元格地址
    n   : 向下偏移量
    return: 偏移后的地址，例如 "AA25"
    """
    # 使用正则表达式分离列字母和行数字
    match = re.match(r"([A-Za-z]+)(\d+)", addr)
    if not match:
        raise ValueError(f"无效的单元格地址格式: {addr}")

    col_str = match.group(1)  # 列字母部分（可能包含多个字母）
    row_str = match.group(2)  # 行数字部分

    new_row_idx = int(row_str) + n  # 计算新行号

    # 确保行号不小于1
    if new_row_idx < 1:
        new_row_idx = 1

    return f"{col_str}{new_row_idx}"

def up_n_cells(addr: str, n: int) -> str:
    """
    改进版本：处理包含多字母列的 Excel 地址

    addr: 形如 "D5"、"AA24"、"AB123" 的单元格地址
    n   : 向上偏移量
    return: 偏移后的地址，例如 "AA25"
    """
    # 使用正则表达式分离列字母和行数字
    match = re.match(r"([A-Za-z]+)(\d+)", addr)
    if not match:
        raise ValueError(f"无效的单元格地址格式: {addr}")

    col_str = match.group(1)  # 列字母部分（可能包含多个字母）
    row_str = match.group(2)  # 行数字部分

    new_row_idx = int(row_str) - n  # 计算新行号

    # 确保行号不小于1
    if new_row_idx < 1:
        new_row_idx = 1

    return f"{col_str}{new_row_idx}"


def to_absolute_address(cell_ref):
    """
    将 Excel 单元格地址转换为绝对引用格式

    参数:
    cell_ref (str): 单元格地址（如 "A1", "B3", "C.2项目每年借款信息!D4"）

    返回:
    str: 绝对引用格式的地址（如 "$A$1", "C.2项目每年借款信息!$D$4"）
    """
    # 检查是否包含工作表名称
    if '!' in cell_ref:
        # 分离工作表名称和单元格地址
        sheet_name, cell_address = cell_ref.split('!', 1)

        # 处理带引号的工作表名（如 'Sheet Name'!A1）
        if sheet_name.startswith("'") and sheet_name.endswith("'"):
            sheet_name = f"'{sheet_name.strip("'")}'"

        # 转换单元格部分为绝对引用
        return f"{sheet_name}!{_convert_cell_part(cell_address)}"
    else:
        # 直接转换单元格地址
        return _convert_cell_part(cell_ref)


def _convert_cell_part(cell_address):
    """处理单元格地址部分（不含工作表名称）"""
    # 分离列字母和行数字
    col_part = ''.join(filter(str.isalpha, cell_address))
    row_part = ''.join(filter(str.isdigit, cell_address))

    # 添加$符号创建绝对引用
    return f"${col_part}${row_part}"




def excel_date_to_year(excel_dates):
    """
    将 Excel 日期数值转换为对应的年份

    参数:
    excel_dates: Excel 日期数值列表，如 [45901, 46966]

    返回:
    包含对应年份的列表
    """
    years = []

    for excel_date in excel_dates:
        try:
            # Excel 的日期系统是从 1900-01-01 开始计算的
            # # 注意：Excel 错误地将 1900 年视为闰年，所以需要调整
            # if excel_date >= 60:
            #     # 对于大于等于 60 的日期，减去 1 天来修正 Excel 的闰年错误
            #     excel_date -= 1
            # 将 Excel 日期转换为 Python datetime 对象
            date = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=excel_date)

            # 提取年份
            year = date.year
            years.append(year)

        except Exception as e:
            # 如果转换出错，记录错误信息
            years.append(f"错误: {str(e)}")

    return years


def struct_years(file_path, table_b_sheet):
    # 建设期的时间范围
    start_value = "建设开始年月"
    end_value = "建设完成年月"
    # 加载工作簿
    wb = load_workbook(file_path)
    ws = wb[table_b_sheet]

    start_years_address = find_all_cells(file_path, start_value, table_b_sheet)  # [(row1, col1), (row2, col2), ...]
    end_years_address = find_all_cells(file_path, end_value, table_b_sheet)  # [(row1, col1), (row2, col2), ...]
    # 建设期开始年月单元格地址列表
    start = []
    start_years = []
    # 建设期完成年月单元格地址列表
    end = []
    end_years = []
    for start_year_address in start_years_address:
        start.append(right_n_cells(start_year_address[1] + str(start_year_address[0]), 2))
    for end_year_address in end_years_address:
        end.append(right_n_cells(end_year_address[1] + str(end_year_address[0]), 2))
    # 根据单元格地址求出其值所代表的年份
    print("start:", start)
    print("end:", end)
    # 遍历所有单元格地址
    for cell_address in start:
        # 获取单元格值
        cell = ws[cell_address]
        start_years.append(cell.value)

    for cell_address in end:
        # 获取单元格值
        cell = ws[cell_address]
        end_years.append(cell.value)

    start_years = excel_date_to_year(start_years)
    end_years = excel_date_to_year(end_years)

    print("start_years:", start_years)
    print("end_years:", end_years)

    # 找出最大与最小的区间
    min_year = min(start_years)
    max_year = max(end_years)

    return min_year, max_year


def write_loan(loan, file_path):
    # Excel 日期序列号
    start_sn = loan['开始时间']
    end_sn = loan['结束时间']

    # 把序列号转成日期对象（注意 Excel 的 epoch 是 1899-12-30）
    epoch = datetime.datetime(1899, 12, 30)
    start_date = epoch + datetime.timedelta(days=start_sn)
    end_date = epoch + datetime.timedelta(days=end_sn)

    # 取年份
    start_year = start_date.year
    end_year = end_date.year
    delta_year = end_year - start_year  # 整数年差
    delta_year_precise = (end_sn - start_sn) / 365.25  # 带小数年差

    # 输出
    print(f"起始年份 : {start_year}")
    print(f"结束年份 : {end_year}")
    print(f"相差整数年: {delta_year}")
    print(f"相差精确年: {delta_year_precise:.2f}")

    # 表头年份数量（循环次数）
    year_num = delta_year + 1

    # 循环的表名,通过传入loan作为参数，然后从loan中获取
    sheet_name = loan['借款名称']

    # 还款方式,通过传入loan作为参数，然后从loan中获取
    repayment_method = loan['还款方式']

    # 通过起始年份寻找到第一个填充的单元格的地址
    row, col = find_cell(file_path, start_year, sheet_name)

    # C.1表的起始位置 ,本方法内部循环的参数，每次循环向下位移1
    shift = loan['序号']
    C_1_sheet_name = "C.1项目融资信息"
    r, _ = find_cell(file_path, "序号", C_1_sheet_name)
    # 此处的 2 表示在C.1表中，到第一笔借款需要向下偏移几行，本方法所使用的模板中去掉了第一行的表名行，因此向下偏移2
    C_1 = "D" + str(shift + r + 1)

    # C.2表的起始位置,本方法内部循环的参数，每次循环向下位移1
    C_2_sheet_name = "C.2项目每年借款信息"
    # **************************************** 此处修改 *********************************************
    # C_2_start_year = str(start_year) + "（万元）"
    C_2_start_year = start_year
    C_2_row, C_2_col = find_cell(file_path, C_2_start_year, C_2_sheet_name)
    # 此处的 2 与C.1表中的类似，表示在C.2表中，到第一笔借款需要向下偏移几行，本方法所使用的模板中未去掉了第一行的表名行，向下偏移2
    r, _ = find_cell(file_path, "序号", C_2_sheet_name)
    C_2 = C_2_col + str(shift + r)

    # 还本位移,通过传入loan作为参数，然后从loan中提取计算
    # delta_year = end_year - start_year        # 整数年差
    origin = col + str(5)
    target = right_n_cells(origin, delta_year)
    print(target)  # -> I5

    # 还本付息兑付手续费率
    repay_value = "债券还本付息兑付手续费率"
    repay_sheet = "A财务假设"
    repay_rates_row, repay_rates_col = find_cell(file_path, repay_value, repay_sheet)
    repay_rates_addr = right_n_cells(f"{repay_rates_col}{repay_rates_row}", 2)
    repay_rates_addr = to_absolute_address(repay_rates_addr)

    # 债券发行及服务费费率
    bond_value = "债券发行登记服务费率"
    # **************************************** 此处修改 *********************************************
    # bond_value2 = "5年期及以上债券发行手续费率"
    bond_value2 = "5年期及以上债券发行手续费为发行额的0.08%"
    bond_sheet = "A财务假设"
    bond_rates_row, bond_rates_col = find_cell(file_path, bond_value, bond_sheet)
    bond_rates_row2, bond_rates_col2 = find_cell(file_path, bond_value2, bond_sheet)
    bond_rates_addr = right_n_cells(f"{bond_rates_col}{bond_rates_row}", 2)
    # bond_rates_addr2 = right_n_cells(f"{bond_rates_col2}{bond_rates_row2}", 2)
    # **************************************** 此处修改 *********************************************
    bond_rates_addr2 = right_n_cells(f"{bond_rates_col2}{bond_rates_row2}", 1)
    bond_rates_addr = to_absolute_address(bond_rates_addr)
    bond_rates_addr2 = to_absolute_address(bond_rates_addr2)

    # 建设期应付利息 获得建设期的起始和结束年份
    table_B_sheet = "B项目信息"
    table_c_sheet = "c借款还本付息计划表"
    start_year, end_year = struct_years(file_path, table_B_sheet)
    print("建设期的区间为：", start_year, "-", end_year)
    _, c1 = find_cell(file_path, start_year, table_c_sheet)
    _, c2 = find_cell(file_path, end_year, table_c_sheet)

    # 初始化公式生成器（输入输出为同一个文件）
    formula_generator = ExcelFormulaGenerator(
        data_file=file_path,  # 包含所有数据的单一文件
        output_file=file_path  # 输出到同一个文件
    )

    # 还款方式：到期一次性还清
    one_time_repay = [
        # 期初借款余额 = 前一年的期末借款余额
        {
            "operation": "custom",
            "formula_template": "=({A_val})",
            "params": {
                "A_val": {"sheet": "", "cell": left_n_cells(col + str(7), 1)},  # B表的y (向右移动)
            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": col + str(3)  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": year_num,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向右移动一列
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        }

        # 当期还本付息 = 还本 + 付息
        , {
            "operation": "custom",
            "formula_template": "=({A_val}+{B_val})",
            "params": {
                "A_val": {"sheet": "", "cell": col + str(5)},  # A表的y (向右移动)
                "B_val": {"sheet": "", "cell": col + str(6)},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": col + str(4)  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": year_num,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                    "B_val": {"row_shift": 0, "col_shift": 1},  # 每次向右移动一列
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        }

        #  其中：还本, 当期还本，由选择的还款方式决定（手动填写）
        # , {
        #     "operation": "custom",
        #     "formula_template": "=({A_val})",
        #     "params": {
        #         "A_val": {"sheet": "C.1项目融资信息", "cell": C_1},  # A表的y (向右移动)
        #
        #     },
        #     "target": {
        #         "sheet": sheet_name,  # 结果工作表
        #         "cell": target  # 目标起始位置 (向右移动)
        #     }
        # }

        # # 付息 = (期末+还本）*利率，但是一些项目会乘以系数
        # , {
        #     "operation": "custom",
        #     "formula_template": "=(({A_val}+{C_val})*{B_val})",
        #     "params": {
        #         "A_val": {"sheet": "", "cell": col + str(7)},  # A表的y (向右移动)
        #         "C_val": {"sheet": "", "cell": col + str(5)},  # A表的y (向右移动)
        #         "B_val": {"sheet": "C.1项目融资信息", "cell": to_absolute_address(right_n_cells(C_1, 4))},  # A表的y (向右移动)
        #
        #     },
        #     "target": {
        #         "sheet": sheet_name,  # 结果工作表
        #         "cell": col + str(6)  # 目标起始位置 (向右移动)
        #     },
        #     "loop": {
        #         "count": year_num,  # 循环3次
        #
        #         # 参数独立偏移设置
        #         "param_offsets": {
        #             "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
        #             "C_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
        #             "B_val": {"row_shift": 0, "col_shift": 0},  # 每次向右移动一列
        #         },
        #
        #         # 目标单元格偏移设置
        #         "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
        #     }
        # }

        # 期末借款余额 = 期初 + 新增 - 还本
        , {
            "operation": "custom",
            "formula_template": "=({A_val}+{B_val}-{C_val})",
            "params": {
                "A_val": {"sheet": "", "cell": col + str(3)},  # A表的y (向右移动)
                "B_val": {"sheet": C_2_sheet_name, "cell": to_absolute_address(C_2)},  # A表的y (向右移动)
                "C_val": {"sheet": "", "cell": col + str(5)},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": col + str(7)  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": year_num,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                    "B_val": {"row_shift": 0, "col_shift": 1},  # 每次向右移动一列
                    "C_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        }

        # 还本付息兑付手续费 = 当期还本付息 * 费率
        , {
            "operation": "custom",
            "formula_template": "=({A_val}*{B_val})",
            "params": {
                "A_val": {"sheet": "", "cell": col + str(4)},  # A表的y (向右移动)
                "B_val": {"sheet": repay_sheet, "cell": repay_rates_addr},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": col + str(8)  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": year_num,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                    "B_val": {"row_shift": 0, "col_shift": 0},  # 每次向右移动一列
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        }

        # 债券发行及服务费 = 当期新增 * 两个费率和
        , {
            "operation": "custom",
            "formula_template": "=({A_val}*({B_val}+{C_val}))",
            "params": {
                "A_val": {"sheet": C_2_sheet_name, "cell": to_absolute_address(C_2)},  # A表的y (向右移动)
                "B_val": {"sheet": repay_sheet, "cell": bond_rates_addr2},  # A表的y (向右移动)
                "C_val": {"sheet": repay_sheet, "cell": bond_rates_addr},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": col + str(9)  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": year_num,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                    "B_val": {"row_shift": 0, "col_shift": 0},  # 每次向下移动一行
                    "C_val": {"row_shift": 0, "col_shift": 0},  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        }

        # 建设期应付利息 除最后一行
        , {
            "operation": "custom",
            "formula_template": "=(IF(({A_val}-YEAR({B_val}))=0,"
                                "(12-MONTH({B_val})+1)*{C_val}*{D_val}/12,"
                                "{C_val}*{D_val}))",
            "params": {
                "A_val": {"sheet": "", "cell": to_absolute_address(col + str(2))},  # A表的y (向右移动)
                "B_val": {"sheet": C_1_sheet_name, "cell": to_absolute_address(right_n_cells(C_1, 1))},  # A表的y (向右移动)
                "C_val": {"sheet": C_1_sheet_name, "cell": to_absolute_address(C_1)},  # A表的y (向右移动)
                "D_val": {"sheet": C_1_sheet_name, "cell": to_absolute_address(right_n_cells(C_1, 4))},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": col + str(10)  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": year_num - 1,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                    "B_val": {"row_shift": 0, "col_shift": 0},  # 每次向右移动一列
                    "C_val": {"row_shift": 0, "col_shift": 0},  # 每次向下移动一行
                    "D_val": {"row_shift": 0, "col_shift": 0},  # 每次向右移动一列
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        }

        # 建设期应付利息 最后一行
        , {
            "operation": "custom",
            "formula_template": "=(MONTH({B_val})-1)*{C_val}*{D_val}/12",
            "params": {
                "B_val": {"sheet": C_1_sheet_name, "cell": to_absolute_address(right_n_cells(C_1, 1))},  # A表的y (向右移动)
                "C_val": {"sheet": C_1_sheet_name, "cell": to_absolute_address(C_1)},  # A表的y (向右移动)
                "D_val": {"sheet": C_1_sheet_name, "cell": to_absolute_address(right_n_cells(C_1, 4))},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": right_n_cells(col + str(10), year_num-1)  # 目标起始位置 (向右移动)
            }
        }

        # 合计1
        , {
            "operation": "custom",
            "formula_template": "=(SUM({A_val}:{B_val}))",
            "params": {
                "A_val": {"sheet": "", "cell": col + str(4)},  # A表的y (向右移动)
                "B_val": {"sheet": "", "cell": right_n_cells(col + str(4), year_num-1)},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": "C4"  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": 3,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 1, "col_shift": 0},  # 每次向下移动一行
                    "B_val": {"row_shift": 1, "col_shift": 0},  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 1, "col_shift": 0}  # 每次向右移动一列且向下移动一行
            }
        }

        # 合计2
        , {
            "operation": "custom",
            "formula_template": "=(SUM({A_val}:{B_val}))",
            "params": {
                "A_val": {"sheet": "", "cell": col + str(8)},  # A表的y (向右移动)
                "B_val": {"sheet": "", "cell": right_n_cells(col + str(8), year_num-1)},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": "C8"  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": 2,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 1, "col_shift": 0},  # 每次向下移动一行
                    "B_val": {"row_shift": 1, "col_shift": 0},  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 1, "col_shift": 0}  # 每次向右移动一列且向下移动一行
            }
        }

        # # 合计3
        # , {
        #     "operation": "custom",
        #     "formula_template": "=(SUM({A_val}:{B_val}))",
        #     "params": {
        #         "A_val": {"sheet": "", "cell": c1 + str(10)},  # A表的y (向右移动)
        #         "B_val": {"sheet": "", "cell": c2 + str(10)},  # A表的y (向右移动)
        #
        #     },
        #     "target": {
        #         "sheet": sheet_name,  # 结果工作表
        #         "cell": "C10"  # 目标起始位置 (向右移动)
        #     },
        # }


    ]
    interest = {}
    # 借款周期小于10，利息一年一付
    if loan['借款周期（年）'] < 10:
        # 付息 = (期末+还本）*利率，但是一些项目会乘以系数
        interest = {
            "operation": "custom",
            "formula_template": "=(({A_val}+{C_val})*{B_val})",
            "params": {
                "A_val": {"sheet": "", "cell": col + str(7)},  # A表的y (向右移动)
                "C_val": {"sheet": "", "cell": col + str(5)},  # A表的y (向右移动)
                "B_val": {"sheet": C_1_sheet_name, "cell": to_absolute_address(right_n_cells(C_1, 4))},
                # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": col + str(6)  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": year_num,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                    "C_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                    "B_val": {"row_shift": 0, "col_shift": 0},  # 每次向右移动一列
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        }

    # 借款周期大于等于10年，半年一付，如果首次计息是在上半年，则当年只付一次息，最后一年也付一次；
    elif is_first_half_of_year(loan['开始时间']) == "上半年":
        print("开始时间为：", loan['开始时间'], "是上半年")
        # 付息 = (期末+还本）*利率，但是一些项目会乘以系数
        interest = {
            "operation": "custom",
            "formula_template": "=(({A_val}+{C_val})*{B_val}/2)",
            "params": {
                "A_val": {"sheet": "", "cell": col + str(7)},  # A表的y (向右移动)
                "C_val": {"sheet": "", "cell": col + str(5)},  # A表的y (向右移动)
                "B_val": {"sheet": "C.1项目融资信息", "cell": to_absolute_address(right_n_cells(C_1, 4))},
                # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": col + str(6)  # 目标起始位置 (向右移动)
            }
        }

        one_time_repay.append(interest)

        interest = {
            "operation": "custom",
            "formula_template": "=(({A_val}+{C_val})*{B_val})",
            "params": {
                "A_val": {"sheet": "", "cell": right_n_cells(col + str(7), 1)},  # A表的y (向右移动)
                "C_val": {"sheet": "", "cell": right_n_cells(col + str(5), 1)},  # A表的y (向右移动)
                "B_val": {"sheet": "C.1项目融资信息", "cell": to_absolute_address(right_n_cells(C_1, 4))},
                # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": right_n_cells(col + str(6), 1)  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": year_num - 2,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                    "C_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                    "B_val": {"row_shift": 0, "col_shift": 0},  # 每次向右移动一列
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        }

        one_time_repay.append(interest)

        interest = {
            "operation": "custom",
            "formula_template": "=(({A_val}+{C_val})*{B_val}/2)",
            "params": {
                "A_val": {"sheet": "", "cell": right_n_cells(col + str(7), year_num - 1)},  # A表的y (向右移动)
                "C_val": {"sheet": "", "cell": right_n_cells(col + str(5), year_num - 1)},  # A表的y (向右移动)
                "B_val": {"sheet": "C.1项目融资信息", "cell": to_absolute_address(right_n_cells(C_1, 4))},
                # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": right_n_cells(col + str(6), year_num - 1)  # 目标起始位置 (向右移动)
            }
        }


    # 如果首次计息是在下半年，则当年不用付，最后一年付两次。
    else:
        print("开始时间为：", loan['开始时间'], "是下半年")
        # 付息 = (期末+还本）*利率，但是一些项目会乘以系数
        interest = {
            "operation": "custom",
            "formula_template": "=(({A_val}+{C_val})*{B_val})",
            "params": {
                "A_val": {"sheet": "", "cell": right_n_cells(col + str(7), 1)},  # A表的y (向右移动)
                "C_val": {"sheet": "", "cell": right_n_cells(col + str(5), 1)},  # A表的y (向右移动)
                "B_val": {"sheet": "C.1项目融资信息", "cell": to_absolute_address(right_n_cells(C_1, 4))},
                # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": right_n_cells(col + str(6), 1)  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": year_num - 1,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                    "C_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                    "B_val": {"row_shift": 0, "col_shift": 0},  # 每次向右移动一列
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        }



    one_time_repay.append(interest)

    # 生成公式
    formula_generator.generate_formulas(one_time_repay)


    # 计算还本
    repay_config = repay_method_cal(file_path, sheet_name, origin, C_1_sheet_name, C_1, delta_year, repayment_method)
    formula_generator.generate_formulas(repay_config)


def copy_and_delete_sheets(source_file, target_file, target_sheet_name,
                           sheets_to_copy=None, preserve_formulas=True):
    """
    将源Excel文件中指定的sheet内容复制到目标Excel文件的特定sheet中
    支持公式的相对引用调整

    参数:
        preserve_formulas (bool): 是否保留并调整公式（默认为True）
    """
    # 检查源文件和目标文件是否相同
    same_file = os.path.abspath(source_file) == os.path.abspath(target_file)

    if same_file:
        print("注意：源文件和目标文件相同，将直接在同一个文件中操作")
        _copy_and_delete_in_same_file(
            source_file,
            target_sheet_name,
            sheets_to_copy,
            preserve_formulas
        )
    else:
        print("源文件和目标文件不同，将分别处理")
        _copy_and_delete_in_different_files(
            source_file,
            target_file,
            target_sheet_name,
            sheets_to_copy,
            preserve_formulas
        )


def _copy_and_delete_in_same_file(file_path, target_sheet_name,
                                  sheets_to_copy=None, preserve_formulas=True):
    """在同一个文件中复制并删除sheet"""
    wb = load_workbook(file_path)

    if target_sheet_name in wb.sheetnames:
        tgt_sheet = wb[target_sheet_name]
        merged_ranges = list(tgt_sheet.merged_cells.ranges)
        for merged_range in merged_ranges:
            try:
                tgt_sheet.unmerge_cells(str(merged_range))
            except KeyError as e:
                # 如果遇到不存在的单元格，跳过并继续
                print(f"跳过不存在的合并单元格: {merged_range}, 错误: {e}")
                continue
        tgt_sheet.delete_rows(1, tgt_sheet.max_row)

    else:
        tgt_sheet = wb.create_sheet(title=target_sheet_name)

    # 确定要复制的sheet列表
    sheets_to_process = _get_sheets_to_process(wb, sheets_to_copy, target_sheet_name)

    # 记录目标表的起始行
    start_target_row = tgt_sheet.max_row + 1

    # 处理第一个sheet
    first_sheet_name = sheets_to_process[0]
    first_sheet = wb[first_sheet_name]
    _copy_sheet_content(first_sheet, tgt_sheet, preserve_formulas, include_header=True)
    print(f"已复制sheet: {first_sheet_name} (包含表头)")

    # 处理其余sheet
    for sheet_name in sheets_to_process[1:]:
        sheet = wb[sheet_name]
        row_count = _copy_sheet_content(sheet, tgt_sheet, preserve_formulas, include_header=False)
        print(f"已复制sheet: {sheet_name} (跳过表头，共{row_count}行数据)")

    # 自动调整列宽
    # adjust_column_width(tgt_sheet)

    # 删除已复制的sheet
    delete_count = _delete_processed_sheets(wb, sheets_to_process)

    # 保存工作簿
    wb.save(file_path)
    print(f"\n操作完成！数据已复制到 {file_path} 的 [{target_sheet_name}] sheet")
    print(f"共处理 {len(sheets_to_process)} 个sheet，删除 {delete_count} 个sheet")



def _copy_and_delete_in_different_files(source_file, target_file, target_sheet_name,
                                        sheets_to_copy=None, preserve_formulas=True):
    """在不同文件中复制并删除sheet"""
    # 加载源工作簿
    src_wb = load_workbook(source_file)

    # 创建目标工作簿
    tgt_wb = openpyxl.Workbook()

    # 删除默认创建的sheet
    for sheet_name in tgt_wb.sheetnames:
        del tgt_wb[sheet_name]

    # 创建目标sheet
    tgt_sheet = tgt_wb.create_sheet(title=target_sheet_name)

    # 确定要复制的sheet列表
    sheets_to_process = _get_sheets_to_process(src_wb, sheets_to_copy, target_sheet_name)

    # 处理第一个sheet
    first_sheet_name = sheets_to_process[0]
    first_sheet = src_wb[first_sheet_name]
    _copy_sheet_content(first_sheet, tgt_sheet, preserve_formulas, include_header=True)
    print(f"已复制sheet: {first_sheet_name} (包含表头)")

    # 处理其余sheet
    for sheet_name in sheets_to_process[1:]:
        sheet = src_wb[sheet_name]
        row_count = _copy_sheet_content(sheet, tgt_sheet, preserve_formulas, include_header=False)
        print(f"已复制sheet: {sheet_name} (跳过表头，共{row_count}行数据)")

    # 自动调整列宽
    # adjust_column_width(tgt_sheet)

    # 保存目标工作簿
    tgt_wb.save(target_file)
    print(f"\n操作完成！数据已复制到 {target_file} 的 [{target_sheet_name}] sheet")
    print(f"共处理 {len(sheets_to_process)} 个sheet")

    # 删除源文件中已复制的sheet
    delete_count = _delete_processed_sheets(src_wb, sheets_to_process)

    # 保存修改后的源工作簿
    if delete_count > 0:
        src_wb.save(source_file)
        print(f"\n已从源文件删除 {delete_count} 个sheet，源文件已更新")
    else:
        print("\n没有删除任何sheet")


def _copy_sheet_content(source_sheet, target_sheet, preserve_formulas, include_header=True):
    """复制工作表内容并调整公式引用"""
    start_row = 1 if include_header else 2
    row_count = 0

    # 计算行偏移量（目标表当前行与源表起始行之间的差值）
    row_offset = target_sheet.max_row - start_row + 1

    for source_row_idx, row in enumerate(source_sheet.iter_rows(min_row=start_row), start_row):
        # 目标行位置
        target_row_idx = target_sheet.max_row + 1

        for col_idx, cell in enumerate(row, 1):
            new_cell = target_sheet.cell(row=target_row_idx, column=col_idx)

            if preserve_formulas and cell.data_type == 'f':
                # 获取公式并调整相对引用
                formula = cell.value
                adjusted_formula = adjust_formula_references(formula, row_offset)
                new_cell.value = adjusted_formula
            else:
                # 复制值
                new_cell.value = cell.value

            # 可选：复制样式
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.alignment = copy(cell.alignment)

        row_count += 1

    return row_count


def adjust_formula_references(formula, row_offset):
    """
    调整公式中的单元格引用，使其相对移动指定行数

    参数:
        formula (str): Excel公式字符串
        row_offset (int): 行偏移量（正数表示向下移动）

    返回:
        str: 调整后的公式
    """
    # 正则表达式匹配Excel单元格引用（如A1、$B$2、C$3、$D4等）
    pattern = r'([A-Za-z]{1,3})(\$?)(\d+)'

    def adjust_match(match):
        col_ref = match.group(1)  # 列引用（如"A"、"AB"）
        row_lock = match.group(2)  # 行锁定符（"$"或空）
        row_num = int(match.group(3))  # 行号

        # 如果行没有被锁定（$），则调整行号
        if not row_lock:
            row_num += row_offset

        return f"{col_ref}{row_lock}{row_num}"

    # 应用调整
    return re.sub(pattern, adjust_match, formula)


def _get_sheets_to_process(workbook, sheets_to_copy, target_sheet_name):
    """获取要处理的sheet列表"""
    if sheets_to_copy is None:
        sheets_to_process = workbook.sheetnames.copy()
    else:
        sheets_to_process = [s for s in sheets_to_copy if s in workbook.sheetnames]

    # 移除目标sheet（如果存在）
    if target_sheet_name in sheets_to_process:
        sheets_to_process.remove(target_sheet_name)

    if not sheets_to_process:
        raise ValueError("没有找到任何匹配的sheet进行复制！")

    return sheets_to_process


def _delete_processed_sheets(workbook, sheets_to_process):
    """删除已处理的sheet"""
    delete_count = 0
    for sheet_name in sheets_to_process:
        if sheet_name in workbook.sheetnames and len(workbook.sheetnames) > 1:
            del workbook[sheet_name]
            print(f"已删除sheet: {sheet_name}")
            delete_count += 1
        else:
            print(f"警告：无法删除 {sheet_name}，因为它是文件中最后一个sheet")
    return delete_count


def adjust_column_width(sheet):
    """自动调整列宽"""
    for col in range(1, sheet.max_column + 1):
        max_length = 0
        column = get_column_letter(col)
        for cell in sheet[column]:
            if cell.value:
                try:
                    # 处理可能的编码问题
                    max_length = max(max_length, len(str(cell.value)))
                except Exception as e:
                    print(f"调整列宽时出错: {e}")
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column].width = adjusted_width



def build_formula_config(cells: List[str],
                         start_cell: str = "F5",
                         target_start: str = "F10",
                         loop_cnt: int = 3,
                         param_row_shift: int = 0,
                         param_col_shift: int = 1,
                         target_row_shift: int = 0,
                         target_col_shift: int = 1,
                         sheet: str = "",
                         target_sheet: str = "") -> Dict[str, Any]:
    """
    根据单元格列表动态生成 JSON 配置
    :param cells: 参与运算的单元格偏移列表，如 ["F5", "G5", "H5"]
    :param start_cell: 第一个参数实际对应的单元格（会被映射到 cells[0]）
    :param target_start: 公式写入起始单元格
    :param loop_cnt: 循环次数
    :param param_row_shift: 每次循环参数的 row 偏移
    :param param_col_shift: 每次循环参数的 col 偏移
    :param target_row_shift: 每次循环目标的 row 偏移
    :param target_col_shift: 每次循环目标的 col 偏移
    :param sheet: 工作表名
    :return: 可直接丢给 ExcelFormulaGenerator.generate_formulas 的 dict
    """

    # 1. 生成参数名 B_val, C_val, D_val ...
    param_names = [f"{chr(66 + i)}_val" for i in range(len(cells))]
    # 2. 生成 params 区块
    params = {
        name: {"sheet": sheet, "cell": right_n_cells(cell , 2)}
        for name, cell in zip(param_names, cells)
    }
    # 3. 生成公式模板  =({B_val}+{C_val}+...)
    template = "=(" + "+".join([f"{{{p}}}" for p in param_names]) + ")"
    # 4. 生成 param_offsets 区块
    param_offsets = {
        name: {"row_shift": param_row_shift, "col_shift": param_col_shift}
        for name in param_names
    }

    return {
        "operation": "custom",
        "formula_template": template,
        "params": params,
        "target": {"sheet": target_sheet, "cell": right_n_cells(target_start, 2)},
        "loop": {
            "count": loop_cnt,
            "param_offsets": param_offsets,
            "target_offset": {
                "row_shift": target_row_shift,
                "col_shift": target_col_shift
            }
        }
    }


def is_first_half_of_year(serial):
    """
    判断Excel日期序列值属于上半年(1-6月)还是下半年(7-12月)

    参数:
    serial: Excel日期序列值（可以是数字序列或字符串）

    返回:
    "上半年" 或 "下半年"
    """
    try:
        # 1. 处理数值类型的Excel序列日期
        if isinstance(serial, (int, float)):
            base_date = datetime.datetime(1899, 12, 30)
            date = base_date + datetime.timedelta(days=serial)
            month = date.month
            return "上半年" if month <= 6 else "下半年"

    except Exception as e:
        # 错误处理
        print(f"处理日期 {serial} 时出错: {e}")

    # 4. 默认返回（无法确定时）
    return "上半年"  # 或根据业务需求返回"下半年"



def loan_summary(file_path, sheet_name, year_num):
    # file_path = "融资信息表3.xlsx"
    # sheet_name = "c借款还本付息计划表"
    # year_num = 21
    value = "期初借款余额"
    C_1_sheet_name = "C.1项目融资信息"

    formula_generator = ExcelFormulaGenerator(
        data_file=file_path,  # 包含所有数据的单一文件
        output_file=file_path  # 输出到同一个文件
    )

    # address = find_all_coords_co(file_path, value, sheet_name)  #address = ['C3', 'F7', 'A10']
    # target = address[-1]
    # print("target地址为：", target)
    # print("address地址为：", address)
    # print("address[:-1]为：", address[:-1])
    # summary = [build_formula_config(address[:-1], address[0], address[-1], year_num, target_sheet=sheet_name)]
    # print("summary配置为：", summary)
    #

    # formula_generator.generate_formulas(summary)

    address2 = find_all_cells(file_path, value, sheet_name)   #address2 = [('C', 3), ('F', 7), ('A', 10)]


    # 计算借款总结总计借款金额
    r, c = find_cell(file_path, "借款总结", sheet_name)
    summary_target = c + str(r)
    # 借款总结金额单元格地址
    summary_target = right_n_cells(summary_target, 1)

    r, c = find_cell(file_path, "借款金额\n（万元）", C_1_sheet_name)
    summary_address = []
    summary_address.extend(f"${c}${r + i + 2}" for i in range(len(address2) - 1))
    print("summary_address:", summary_address)


    # 1. 生成参数名 B_val, C_val, D_val ...
    param_names = [f"{chr(66 + i)}_val" for i in range(len(summary_address))]
    # 2. 生成 params 区块
    params = {
        name: {"sheet": C_1_sheet_name, "cell": cell}
        for name, cell in zip(param_names, summary_address)
    }
    # 3. 生成公式模板  =({B_val}+{C_val}+...)
    template = "=(" + "+".join([f"{{{p}}}" for p in param_names]) + ")"


    summary_config = [{
        "operation": "custom",
        "formula_template": template,
        "params": params,
        "target": {"sheet": sheet_name, "cell": summary_target},
    }]

    print("summary_config:", summary_config)

    formula_generator.generate_formulas(summary_config)

    # 填充借款总结其他项目
    target2 = address2[-1]
    target2 = target2[1] + str(target2[0])
    print("target2:", target2)
    total_target1 = down_n_cells(right_n_cells(target2, 1), 1)
    print("total_target1:", total_target1)
    total_target_start1 = right_n_cells(total_target1, 1)
    print("total_target_start1:", total_target_start1)

    total_target2 = down_n_cells(total_target1, 4)
    print("total_target2:", total_target2)
    total_target_start2 = down_n_cells(total_target_start1, 4)
    print("total_target_start2:", total_target_start2)

    for i in range(8):
        address2_co = [c + str(r + i) for r, c in address2]
        print("address2_co的地址为:", address2_co)
        summary = [build_formula_config(address2_co[:-1], address2_co[0], address2_co[-1], year_num, target_sheet=sheet_name)]
        print("summary配置为：", summary)
        formula_generator.generate_formulas(summary)

    # 建设期应付利息 获得建设期的起始和结束年份
    table_B_sheet = "B项目信息"
    table_c_sheet = "c借款还本付息计划表"
    start_year, end_year = struct_years(file_path, table_B_sheet)
    print("建设期的区间为：", start_year, "-", end_year)
    _, c1 = find_cell(file_path, str(start_year), table_c_sheet)
    _, c2 = find_cell(file_path, str(end_year), table_c_sheet)
    total_target3 = down_n_cells(total_target2, 2)
    # 使用正则表达式分离列字母和行数字
    match = re.match(r"([A-Za-z]+)(\d+)", total_target3)
    if not match:
        raise ValueError(f"无效的单元格地址格式: {total_target3}")

    row_str = match.group(2)  # 行数字部分
    row_str = int(row_str)
    print("c1:", c1)
    print("c2:", c2)
    print("row_str:", row_str)



    summary_total = [
    # 合计1
        {
        "operation": "custom",
        "formula_template": "=(SUM({A_val}:{B_val}))",
        "params": {
            "A_val": {"sheet": "", "cell": total_target_start1},  # A表的y (向右移动)
            "B_val": {"sheet": "", "cell": right_n_cells(total_target_start1, year_num-1)},  # A表的y (向右移动)

        },
        "target": {
            "sheet": sheet_name,  # 结果工作表
            "cell": total_target1  # 目标起始位置 (向右移动)
        },
        "loop": {
            "count": 3,  # 循环3次

            # 参数独立偏移设置
            "param_offsets": {
                "A_val": {"row_shift": 1, "col_shift": 0},  # 每次向下移动一行
                "B_val": {"row_shift": 1, "col_shift": 0},  # 每次向下移动一行
            },

            # 目标单元格偏移设置
            "target_offset": {"row_shift": 1, "col_shift": 0}  # 每次向右移动一列且向下移动一行
        }
    }

    # 合计2
    , {
        "operation": "custom",
        "formula_template": "=(SUM({A_val}:{B_val}))",
        "params": {
            "A_val": {"sheet": "", "cell": total_target_start2},  # A表的y (向右移动)
            "B_val": {"sheet": "", "cell": right_n_cells(total_target_start2, year_num-1)},  # A表的y (向右移动)

        },
        "target": {
            "sheet": sheet_name,  # 结果工作表
            "cell": total_target2  # 目标起始位置 (向右移动)
        },
        "loop": {
            "count": 2,  # 循环3次

            # 参数独立偏移设置
            "param_offsets": {
                "A_val": {"row_shift": 1, "col_shift": 0},  # 每次向下移动一行
                "B_val": {"row_shift": 1, "col_shift": 0},  # 每次向下移动一行
            },

            # 目标单元格偏移设置
            "target_offset": {"row_shift": 1, "col_shift": 0}  # 每次向右移动一列且向下移动一行
        }
    }

    # # 合计3
    # , {
    #     "operation": "custom",
    #     "formula_template": "=(SUM({A_val}:{B_val}))",
    #     "params": {
    #         "A_val": {"sheet": "", "cell": str(c1) + str(row_str)},  # A表的y (向右移动)
    #         "B_val": {"sheet": "", "cell": str(c2) + str(row_str)},  # A表的y (向右移动)
    #
    #     },
    #     "target": {
    #         "sheet": sheet_name,  # 结果工作表
    #         "cell": total_target3  # 目标起始位置 (向右移动)
    #     }
    # }
    ]

    formula_generator.generate_formulas(summary_total)


if __name__ == "__main__":

    file_path = "融资信息表3.xlsx"
    sheet_name = "c借款还本付息计划表"
    year_num = 21
    C_1_sheet_name = "C.1项目融资信息"

    # address2 = ['C3', 'F7', 'A10']
    #
    # # 计算借款总结总计借款金额
    # r, c = find_cell(file_path, "借款总结", sheet_name)
    # summary_target = c + str(r)
    # # 借款总结金额单元格地址
    # summary_target = right_n_cells(summary_target, 1)
    #
    # r, c = find_cell(file_path, "借款金额（万元）", C_1_sheet_name)
    # summary_address = []
    # summary_address.extend(f"${c}${r + i + 2}" for i in range(len(address2) - 1))
    # print("summary_address:", summary_address)
    #
    #
    # # 1. 生成参数名 B_val, C_val, D_val ...
    # param_names = [f"{chr(66 + i)}_val" for i in range(len(summary_address))]
    # # 2. 生成 params 区块
    # params = {
    #     name: {"sheet": C_1_sheet_name, "cell": cell}
    #     for name, cell in zip(param_names, summary_address)
    # }
    # # 3. 生成公式模板  =({B_val}+{C_val}+...)
    # template = "=(" + "+".join([f"{{{p}}}" for p in param_names]) + ")"
    #
    #
    # summary_config = [{
    #     "operation": "custom",
    #     "formula_template": template,
    #     "params": params,
    #     "target": {"sheet": sheet_name, "cell": summary_target},
    # }]
    #
    # print("summary_config:", summary_config)

    # loan_summary(file_path, sheet_name, year_num)
    year_number = 46082
    year = is_first_half_of_year(year_number)
    print(year)

    print("done")

