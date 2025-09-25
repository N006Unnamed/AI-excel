from openpyxl.utils import get_column_letter
import re
from collections import OrderedDict
# from openpyxl import load_workbook
from typing import List, Tuple
import json
from openpyxl_vba import load_workbook


def find_cell(file_path, value ,sheet_name):
    """
    找到Excel表中特定值的单元格的行、列

    参数:
    file_path: Excel文件路径
    value：特定值
    sheet_name: 工作表名称(可选)
    """
    # 加载工作簿
    wb = load_workbook(file_path, read_only=True)  # 使用只读模式提高性能

    if sheet_name:
        sheet = wb[sheet_name]
    else:
        sheet = wb.active

    # 查找特定值的单元格
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == value:
                # 获取行号和列号
                row_number = cell.row
                column_number = cell.column
                column_letter = get_column_letter(column_number)

                wb.close()  # 提前关闭工作簿

                return row_number, column_letter

    wb.close()
    return None, None


# def find_cell(file_path, value ,sheet_name):
#     """
#     找到Excel表中特定值的单元格的行、列
#
#     参数:
#     file_path: Excel文件路径
#     value：特定值
#     sheet_name: 工作表名称(可选)
#     """
#     # 加载工作簿
#     wb = load_workbook(file_path)  # 替换为你的文件路径
#
#     if sheet_name:
#         sheet = wb[sheet_name]
#     else:
#         sheet = wb.active
#
#     # 查找特定值的单元格
#     target_value = value  # 替换为你要查找的值
#
#     row_number = None
#     column_letter = None
#
#     for row in sheet.iter_rows():
#         for cell in row:
#             if cell.value == target_value:
#                 # 获取行号和列号
#                 row_number = cell.row
#                 column_number = cell.column
#                 column_letter = get_column_letter(column_number)
#
#                 print(f"找到值 '{target_value}' 在:")
#                 print(f" - 行号: {row_number}")
#                 print(f" - 列号: {column_number}")
#                 print(f" - 列字母: {column_letter}")
#                 print(f" - Excel坐标: {column_letter}{row_number}")
#
#                 # # 如果需要获取整行数据
#                 # row_data = [c.value for c in sheet[row_number]]
#                 # print(f" - 整行数据: {row_data}")
#                 #
#                 # # 如果需要获取整列数据
#                 # col_data = [r[column_number - 1].value for r in
#                 #             sheet.iter_rows(min_col=column_number, max_col=column_number)]
#                 # print(f" - 整列数据: {col_data}")
#
#     return row_number, column_letter


def find_all_cells(file_path: str,
                   value,
                   sheet_name: str = None) -> List[Tuple[int, str]]:
    """
    找到 Excel 表中**所有**等于给定值的单元格（行号, 列字母）
    返回：[(row1, col1), (row2, col2), ...]
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    addresses = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == value:
                addresses.append((cell.row, get_column_letter(cell.column)))

    return addresses


def find_all_coords_co(file_path: str,
                    value,
                    sheet_name: str = None) -> List[str]:
    """
    找到 Excel 表中**所有**等于给定值的单元格（行号, 列字母）
    返回：['D4', 'F5', ...]
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    coords = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == value:
                coords.append(f"{get_column_letter(cell.column)}{cell.row}")
    return coords


def read_cell_formula(file_path, sheet_name, cell_address):
    """
    读取指定单元格的公式

    参数:
    file_path: Excel文件路径
    sheet_name: 工作表名称
    cell_address: 单元格地址(如"A1"、"D5")
    """
    # 加载工作簿，注意必须设置data_only=False才能获取公式
    wb = load_workbook(file_path, data_only=False)

    try:
        # 选择工作表
        sheet = wb[sheet_name]

        # 获取单元格
        cell = sheet[cell_address]

        # 检查单元格是否包含公式
        if cell.data_type == 'f':  # 'f'表示公式
            formula = cell.value
            print(f"单元格 {sheet_name}!{cell_address} 的公式: {formula}")
            return formula
        else:
            print(f"单元格 {sheet_name}!{cell_address} 不包含公式")
            return None
    except KeyError:
        print(f"错误: 工作表 '{sheet_name}' 不存在")
        return None
    finally:
        wb.close()


def set_cell_value(file_path, row_number, column_number, new_value, sheet_name=None):
    """
    给Excel中特定行和列的单元格赋值

    参数:
    file_path: Excel文件路径
    row_number: 行号(整数)
    column_number: 列号(整数，1=A, 2=B,...)
    new_value: 要设置的新值
    sheet_name: 工作表名称(可选)
    """
    # 加载工作簿
    wb = load_workbook(file_path)

    # 选择工作表
    if sheet_name:
        sheet = wb[sheet_name]
    else:
        sheet = wb.active

    # 获取单元格
    cell = sheet.cell(row=row_number, column=column_number)

    # 获取旧值
    old_value = cell.value

    # 设置新值
    cell.value = new_value

    # 获取列字母
    col_letter = get_column_letter(column_number)

    # 保存文件
    wb.save(file_path)

    print(f"成功更新单元格 {col_letter}{row_number} 的值:")
    print(f"旧值: {old_value}")
    print(f"新值: {new_value}")

    return cell


def extract_formula_parameters(formula):
    """
    提取公式中的所有参数（包含详细的工作表和单元格信息）
    """
    # 正则表达式模式，支持跨工作表引用
    pattern = r"""
        (?:                     # 匹配工作表名称（可选）
            (                   # 捕获组1：工作表名称
                (?:             # 两种形式的工作表名称
                    '([^']+)'   # 带引号的工作表名（捕获组2）
                    |           # 或
                    ([^\s!='"+*-][^!']*)    # 不带引号的工作表名（捕获组3）
                )
            )!
        )?                      
        (\$?[A-Z]{1,3})        # 捕获组4：列部分
        (\$?\d+)               # 捕获组5：行部分
    """

    # 使用OrderedDict保持顺序并去重
    unique_params = OrderedDict()

    # 查找所有匹配
    for match in re.finditer(pattern, formula, flags=re.VERBOSE | re.IGNORECASE):
        # 提取工作表名称和引用格式
        sheet_name = ""
        sheet_ref = ""
        if match.group(1):
            # 获取工作表引用字符串（保留原始格式）
            sheet_ref = match.group(1) + "!"

            # 获取纯工作表名称
            if match.group(2):
                sheet_name = match.group(2)
            elif match.group(3):
                sheet_name = match.group(3)

        # 提取原始列和行部分
        orig_col = match.group(4)
        orig_row = match.group(5)

        # 获取不带$的纯列名和行号
        pure_col = orig_col.replace('$', '')
        pure_row = orig_row.replace('$', '')
        cell_ref = f"{pure_col}{pure_row}"

        # 创建参数标识（保留工作表名称）
        param_key = (sheet_name, cell_ref)

        # 如果参数已存在，跳过
        if param_key in unique_params:
            continue

        # 创建详细参数信息
        param_info = {
            "sheet_name": sheet_name,  # 工作表名（无引号）
            "sheet_ref": sheet_ref,  # 工作表引用（含引号和!）
            "cell_ref": cell_ref,  # 单元格引用（无$）
            "full_ref": sheet_ref + orig_col + orig_row,  # 完整原始引用
            "is_absolute": '$' in orig_col or '$' in orig_row,  # 是否为绝对引用
            "column": pure_col,  # 列字母
            "row": pure_row,  # 行号
            "orig_column": orig_col,  # 原始列（含$）
            "orig_row": orig_row  # 原始行（含$）
        }

        # 添加到有序字典（自动去重）
        unique_params[param_key] = param_info

    return list(unique_params.values())


def replace_formula_parameters(formula, cell_replacement_dict=None, sheet_replacement_dict=None):
    """
    增强版：替换公式中的参数和工作表名

    参数:
    formula: 原始公式
    cell_replacement_dict: 单元格替换映射字典 {(工作表名, 单元格引用): 新参数}
    sheet_replacement_dict: 工作表名替换映射字典 {原工作表名: 新工作表名}
    """
    # 正则表达式模式
    pattern = r"""
        (?:                     
            (                   # 捕获组1：工作表名称
                (?:             
                    '([^']+)'   # 带引号的工作表名（捕获组2）
                    |           # 或
                    ([^\s!='"+*-][^!']*)    # 不带引号的工作表名（捕获组3）
                )
            )!
        )?                      
        (\$?[A-Z]{1,3})        # 捕获组4：列部分
        (\$?\d+)               # 捕获组5：行部分
    """

    # 替换函数
    def replace_match(match):
        # 提取工作表名称
        sheet_ref = ""
        sheet_name = ""
        if match.group(1):
            # 获取工作表引用字符串（保留原始格式）
            sheet_ref = match.group(1) + "!"

            # 获取纯工作表名称
            if match.group(2):
                sheet_name = match.group(2)
            elif match.group(3):
                sheet_name = match.group(3)

        # 提取原始列和行部分
        orig_col = match.group(4)
        orig_row = match.group(5)

        # 获取不带$的纯列名和行号
        pure_col = orig_col.replace('$', '')
        pure_row = orig_row.replace('$', '')
        cell_ref = f"{pure_col}{pure_row}"

        # 1. 工作表名替换
        new_sheet_ref = sheet_ref
        if sheet_replacement_dict and sheet_name in sheet_replacement_dict:
            new_sheet_name = sheet_replacement_dict[sheet_name]

            # 确定是否需要引号（如果新表名包含特殊字符）
            if any(char in new_sheet_name for char in [' ', '!', "'", '"', ':', '-']):
                new_sheet_ref = f"'{new_sheet_name}'!"
            else:
                new_sheet_ref = f"{new_sheet_name}!"

        # 2. 单元格引用替换
        cell_replacement = None
        if cell_replacement_dict:
            # 创建参数键
            param_key = (sheet_name, cell_ref)

            # 获取替换值
            if param_key in cell_replacement_dict:
                cell_replacement = cell_replacement_dict[param_key]

        # 组合结果
        if cell_replacement:
            # 替换单元格部分
            return f"{new_sheet_ref}{cell_replacement}"
        else:
            # 保留原始单元格
            return f"{new_sheet_ref}{orig_col}{orig_row}"

    # 执行替换
    return re.sub(pattern, replace_match, formula, flags=re.VERBOSE | re.IGNORECASE)


def print_parameters(parameters, detailed=False):
    """
    打印参数信息（美观格式）
    """
    print("\n提取的参数详情:")
    print("-" * 60)
    print(f"{'序号':<5} {'工作表名':<15} {'单元格':<8} {'完整引用':<25} {'绝对引用':<8}")
    print("-" * 60)

    for i, param in enumerate(parameters, 1):
        full_ref = param["sheet_ref"] + param["orig_column"] + param["orig_row"]
        print(
            f"{i:<5} {param['sheet_name']:<15} {param['cell_ref']:<8} {full_ref:<25} {'是' if param['is_absolute'] else '否':<8}")

    if detailed:
        print("\n详细参数信息:")
        for i, param in enumerate(parameters, 1):
            print(f"\n参数 #{i}:")
            for key, value in param.items():
                print(f"  {key}: {value}")



# 使用示例
if __name__ == "__main__":
    # set_cell_value(
    #     file_path="销售数据.xlsx",
    #     row_number=10,
    #     column_number=5,  # E列
    #     new_value=25000,
    #     sheet_name="第一季度"
    # )
    file_path = "财务分析套表(修改ver2).xlsx"
    value = "现金流入"
    sheet_name = "a财务现金流量表"
    row, col = find_cell(file_path, value, sheet_name)
    column_letter = get_column_letter(col+2)
    letter = column_letter+str(row)
    print(letter)

    formula = read_cell_formula(file_path, sheet_name, letter)
    print(formula)

    print(f"\n原始公式: {formula}")

    # 1. 提取参数（包含详细的工作表和单元格信息）
    parameters = extract_formula_parameters(formula)

    print(parameters)

    # 提取所有 sheet_name 值
    sheet_names = [item['sheet_name'] for item in parameters]
    cell_address = [item['cell_ref'] for item in parameters]

    # 输出结果
    print(sheet_names)
    print(cell_address)

    # 2. 打印参数信息
    print_parameters(parameters)

    sheet_mapping = {
        sheet_names[i]: f"{{project{i + 1}}}"
        for i in range(len(sheet_names))
    }

    print("sheet_mapping:", sheet_mapping)

    cell_mapping = {
        (sheet_names[i], cell_address[i]): f"{{cell{i + 1}}}"
        for i in range(len(sheet_names))
    }

    print("cell_mapping:", cell_mapping)


    # 5. 同时替换工作表名和单元格引用
    combined_formula = replace_formula_parameters(
        formula,
        cell_replacement_dict=cell_mapping,
        sheet_replacement_dict=sheet_mapping
    )
    print(f"同时替换两者: {combined_formula}")
    #
    #
    # # 3. 只替换工作表名
    # sheet_only_formula = replace_formula_parameters(
    #     formula,
    #     sheet_replacement_dict=sheet_replacements
    # )
    # print(f"\n仅替换工作表名: {sheet_only_formula}")
    #
    # # 4. 只替换单元格引用
    # cell_only_formula = replace_formula_parameters(
    #     formula,
    #     cell_replacement_dict=cell_replacements
    # )
    # print(f"仅替换单元格: {cell_only_formula}")
    #
    #
    #
    # # 6. 自定义替换（使用参数列表）
    # param_keys = [(param["sheet_name"], param["cell_ref"]) for param in parameters]
    # custom_cell_params = {
    #     key: f"PARAM_{i + 1}" for i, key in enumerate(param_keys)}
    # custom_formula = replace_formula_parameters(
    #     formula,
    #     cell_replacement_dict=custom_cell_params
    # )
    # print(f"自定义参数替换: {custom_formula}")
    #
    # # 7. 输出JSON格式的参数信息
    # print("\nJSON格式参数信息:")
    # print(json.dumps(parameters, indent=2, ensure_ascii=False))

