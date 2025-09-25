import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

from circulate_formula import ExcelFormulaGenerator
from copy_formula import find_last_used_row, find_last_used_column, expand_excel_header, ExcelAutoFiller
from findAndSet import find_cell
from loan_assignment import right_n_cells, up_n_cells, down_n_cells
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl_vba import load_workbook

def table_III(input_path, sheet_name, last_col):
    """
    填充III表，经营性收入，对应F表的6年度合计(不含税)和4补贴收入
             政府补贴收入，对应F表的4补贴收入
    :param input_path: 文件路径
    :param sheet_name: III表的名称
    :param last_col: III表扩充表头后的最后一列
    :return:
    """

    # 经营性收入 第一个单元格地址
    target_address1 = "B4"

    # 政府补贴收入 第一个单元格地址
    target_address2 = "B5"

    # 表F 年度合计（不含税） 地址
    table_F_sheet = "F项目收入"
    r = find_last_used_row(input_path, table_F_sheet)
    address_F1 = "D" + str(r)

    # 表F 补贴收入 地址
    value = "补贴收入"
    r, _ = find_cell(input_path, value, table_F_sheet)
    address_F2 = "D" + str(r)

    # 循环次数
    loop_num = column_index_from_string(last_col) - column_index_from_string("B") - 1

    formula_generator = ExcelFormulaGenerator(
        data_file=input_path,  # 包含所有数据的单一文件
        output_file=input_path  # 输出到同一个文件
    )

    config = [
        # 1.1 建设投资
        {
            "operation": "custom",
            "formula_template": "=({A_val}-{B_val})",
            "params": {
                "A_val": {"sheet": table_F_sheet, "cell": address_F1},  # A表的y (向右移动)
                "B_val": {"sheet": table_F_sheet, "cell": address_F2},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": target_address1  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                    "B_val": {"row_shift": 0, "col_shift": 1}  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        },
        # 1.2 建设期利息
        {
            "operation": "custom",
            "formula_template": "=({A_val})",
            "params": {
                "A_val": {"sheet": table_F_sheet, "cell": address_F2},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": target_address2  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        }
    ]

    formula_generator.generate_formulas(config)


def table_VI(input_path, sheet_name, last_col):
    """
    表 VI 的填充，行 3457，分别对应c的n.1， n.2.1， n.3， n.2.2

    :param input_path:
    :param sheet_name:
    :param last_col:
    :return:
    """

    # 本期新增 第一个单元格地址
    target_address1 = "B3"

    # 本期偿还 第一个单元格地址
    target_address2 = "B4"

    # 期末本金 第一个单元格地址
    target_address3 = "B5"

    # 应付利息 第一个单元格地址
    target_address4 = "B7"

    # 表c 的n.1 地址
    table_c_sheet = "c借款还本付息计划表"
    r = find_last_used_row(input_path, table_c_sheet)
    address_c1 = "D" + str(r - 7)

    # 表c 的n.2.1 地址
    address_c2 = "D" + str(r - 5)

    # 表c 的n.2.2 地址
    address_c3 = "D" + str(r - 4)

    # 表c 的n.3 地址
    address_c4 = "D" + str(r - 3)

    _, c = find_last_used_column(input_path, table_c_sheet)

    # 循环次数
    loop_num = column_index_from_string(last_col) - column_index_from_string("B") - 1

    formula_generator = ExcelFormulaGenerator(
        data_file=input_path,  # 包含所有数据的单一文件
        output_file=input_path  # 输出到同一个文件
    )

    config = [
        # 本期新增
        {
            "operation": "custom",
            "formula_template": "=({B_val}-{A_val})",
            "params": {
                "A_val": {"sheet": table_c_sheet, "cell": address_c1},  # A表的y (向右移动)
                "B_val": {"sheet": table_c_sheet, "cell": right_n_cells(address_c1, 1)},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": target_address1  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                    "B_val": {"row_shift": 0, "col_shift": 1}  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        },
        # 本期偿还
        {
            "operation": "custom",
            "formula_template": "=({A_val})",
            "params": {
                "A_val": {"sheet": table_c_sheet, "cell": address_c2},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": target_address2  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        },
        # 期末本金
        {
            "operation": "custom",
            "formula_template": "=({A_val})",
            "params": {
                "A_val": {"sheet": table_c_sheet, "cell": address_c4}  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": target_address3  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1}  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        },
        # 应付利息
        {
            "operation": "custom",
            "formula_template": "=({A_val})",
            "params": {
                "A_val": {"sheet": table_c_sheet, "cell": address_c3},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": target_address4  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        }
    ]

    formula_generator.generate_formulas(config)
    print("done")
    print("c:", c)
    temp_address = c + str(r - 3)
    table_c_formula = f"={table_c_sheet}!{temp_address}"
    print("table_c_formula:", table_c_formula)
    return table_c_formula


def final_copy(input_path, target_year):
    # 用户输入
    output_path = input_path
    sheet_names = ["Ⅲ项目分年度收入合计表", "Ⅵ专项债券应付本息情况表"]
    table_c_formula = ""

    for sheet_name in sheet_names:
        _, start_col = find_last_used_column(input_path, sheet_name)
        start_col_plus = column_index_from_string(start_col) + 1
        start_col_plus = get_column_letter(start_col_plus)
        # start_col = "H"
        print(start_col)
        print(start_col_plus)

        if sheet_name == "Ⅲ项目分年度收入合计表":
            header_row = 3
        else:
            header_row = 2

        # 执行扩展
        try:
            expand_excel_header(
                input_file=input_path,
                sheet_name=sheet_name,
                output_file=output_path,
                target_year=target_year,
                header_row=header_row,
                start_col=start_col
            )
        except Exception as e:
            print(f"\n错误: {str(e)}")
            print("操作未完成，请检查输入参数是否正确")

        if sheet_name == "Ⅲ项目分年度收入合计表":
            table_III(input_path, sheet_name, start_col)
        else:
            table_c_formula = table_VI(input_path, sheet_name, start_col)

        r, c = find_cell(input_path, target_year, sheet_name)

        filler = ExcelAutoFiller(input_path)
        filler.set_active_sheet(sheet_name)

        last_row = find_last_used_row(input_path, sheet_name)

        for i in range(r + 1, last_row + 1):
            filler.auto_fill(f"{start_col}{i}", f"{start_col_plus}{i}:{c}{i}")

        # 修改表 VI 本期新增 的最后一格
        if sheet_name == "Ⅵ专项债券应付本息情况表":
            target_address = c + str(3)
            wb = load_workbook(input_path)
            ws = wb[sheet_name]
            ws[target_address].value = table_c_formula
            wb.save(input_path)


def format_existing_excel(file_path, sheet_name, output_path=None):
    """
    修改现有Excel文件的格式：微软雅黑字体、居中对齐、全边框

    参数:
    file_path: 原始Excel文件路径
    output_path: 输出文件路径（如果为None，则覆盖原文件）
    """
    # 如果没有指定输出路径，则覆盖原文件
    if output_path is None:
        output_path = file_path

    # 加载工作簿
    workbook = load_workbook(file_path)
    # 处理所有工作表
    sheet = workbook[sheet_name]

    # 设置微软雅黑字体和居中对齐
    font = Font(name='微软雅黑', size=11)
    alignment = Alignment(horizontal='center', vertical='center')

    # 设置全边框
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 应用到所有有数据的单元格
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:  # 只处理有内容的单元格
                cell.font = font
                cell.alignment = alignment
                cell.border = border

    # 设置矢车菊蓝背景 (RGB: 100, 149, 237)
    cornflower_light_fill = PatternFill(
        start_color="ADC6E6",  # 矢车菊蓝浅色60%的近似十六进制代码
        end_color="ADC6E6",
        fill_type="solid"
    )

    # 设置红色字体
    red_font = Font(name='微软雅黑', size=11, color="FF0000", bold=True)

    # 设置居中对齐
    alignment = Alignment(horizontal='center', vertical='center')

    # 设置全边框
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 获取第二行中的所有单元格
    second_row = 2  # Excel行号从1开始，所以第二行是2

    # 检查第二行是否有数据
    if sheet.max_row >= second_row:
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=second_row, column=col)

            # 应用格式
            cell.fill = cornflower_light_fill
            cell.font = red_font
            cell.alignment = alignment
            cell.border = border

    # 可选：保持其他行的格式不变，或者应用基本格式
    # 这里我们保持其他行不变，只修改第二行

    # 自动调整列宽以适应内容
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        # # 设置列宽，留出一些边距
        # adjusted_width = min(max_length + 2, 50)  # 限制最大列宽为50
        # sheet.column_dimensions[column_letter].width = adjusted_width

    # 保存工作簿
    workbook.save(output_path)
    print(f"Excel格式已成功更新并保存到: {output_path}")

def table_c_last(input_path):

    table_c_sheet = "c借款还本付息计划表"

    r = find_last_used_row(input_path, table_c_sheet)
    target_address1 = "A" + str(r + 1)
    target_address2 = "A" + str(r + 2)

    merge_address1 = target_address1
    merge_address2 = target_address2



    # 利息备付率 与 偿债备付率
    target_address1 = right_n_cells(target_address1, 1)
    target_address2 = right_n_cells(target_address2, 1)
    final_address1 = target_address1
    final_address2 = target_address2



    # 第一个单元格
    target_address1 = right_n_cells(target_address1, 2)
    target_address2 = right_n_cells(target_address2, 2)

    # 初始化公式生成器（输入输出为同一个文件）
    formula_generator = ExcelFormulaGenerator(
        data_file=input_path,  # 包含所有数据的单一文件
        output_file=input_path  # 输出到同一个文件
    )

    _, last_col = find_last_used_column(input_path, table_c_sheet)
    loop_num = column_index_from_string(last_col) - column_index_from_string("D")

    table_b_sheet = "b利润与利润分配表（损益和利润分配表）"
    # r, c = find_cell(input_path, value, table_b_sheet)
    table_b_address1 = "D25"
    table_c_address1 = up_n_cells(target_address1, 5)

    config = [
        # 利息备付率（%）
        {
            "operation": "custom",
            "formula_template": "=IF({B_val}<>0,{A_val}/{B_val},0)",
            "params": {
                "A_val": {"sheet": table_b_sheet, "cell": table_b_address1},  # A表的y (向右移动)
                "B_val": {"sheet": '', "cell": table_c_address1}  # A表的y (向右移动)

            },
            "target": {
                "sheet": table_c_sheet,  # 结果工作表
                "cell": target_address1  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num+1,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                    "B_val": {"row_shift": 0, "col_shift": 1}  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        },

        # 偿债备付率（%）
        {
            "operation": "custom",
            "formula_template": "=(IF({C_val}<>0,({A_val}-{B_val})/{C_val},""))",
            "params": {
                "A_val": {"sheet": table_b_sheet, "cell": down_n_cells(table_b_address1, 1)},  # A表的y (向右移动)
                "B_val": {"sheet": table_b_sheet, "cell": up_n_cells(table_b_address1, 14)},  # A表的y (向右移动)
                "C_val": {"sheet": '', "cell": up_n_cells(table_c_address1, 2)}  # A表的y (向右移动)

            },
            "target": {
                "sheet": table_c_sheet,  # 结果工作表
                "cell": target_address2  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num+1,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                    "B_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                    "C_val": {"row_shift": 0, "col_shift": 1}  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        }
    ]
    print("开始写入公式")
    formula_generator.generate_formulas(config)

    wb = load_workbook(input_path)
    ws = wb[table_c_sheet]
    # 计算指标 合并单元格
    ws[merge_address1].value = "计算指标"
    ws[merge_address2].value = "计算指标"

    ws[final_address1].value = "利息备付率（%）"
    ws[final_address2].value = "偿债备付率（%）"
    ws[right_n_cells(final_address1, 1)].value = "/"
    ws[right_n_cells(final_address2, 1)].value = "/"
    wb.save(input_path)
    wb.close()

    format_existing_excel(input_path, table_c_sheet)

    wb = load_workbook(input_path)
    ws = wb[table_c_sheet]
    ws.merge_cells(range_string=f"{merge_address1}:{merge_address2}")
    wb.save(input_path)
    wb.close()


    print("done")


if __name__ == "__main__":
    input_path = "财务分析套表自做模板编程用.xlsm"
    target_year = 2047
    table_c_last(input_path)
    print("done")
