import os
from openpyxl import load_workbook
import re


class ExcelReporter:
    def __init__(self, data_file, template_path, output_dir):
        self.data_file = data_file
        self.template_path = template_path
        self.output_dir = output_dir

        # 加载数据工作簿
        self.data_wb = load_workbook(data_file, read_only=True, data_only=True)
        print(f"数据源加载成功! 工作表: {', '.join(self.data_wb.sheetnames)}")

    def get_cell_value(self, sheet_name, cell_ref):
        """获取指定工作表的单元格值"""
        if sheet_name not in self.data_wb.sheetnames:
            raise ValueError(f"工作表 '{sheet_name}' 不存在")
        return self.data_wb[sheet_name][cell_ref].value

    def shift_cell(self, cell_ref, col_shift=0, row_shift=0):
        """将单元格引用向右/向下移动指定位置"""
        # 解析单元格引用（如"A1"）
        col_part = re.match(r'[A-Z]+', cell_ref).group()
        row_part = re.search(r'\d+', cell_ref).group()

        # 转换列字母为数字
        col_num = 0
        for char in col_part:
            col_num = col_num * 26 + (ord(char) - ord('A') + 1)

        # 应用偏移
        new_col_num = col_num + col_shift
        new_row_num = int(row_part) + row_shift

        # 将列数字转换回字母
        new_col_ref = ""
        while new_col_num > 0:
            new_col_num, remainder = divmod(new_col_num - 1, 26)
            new_col_ref = chr(65 + remainder) + new_col_ref

        return f"{new_col_ref}{new_row_num}"

    def calculate(self, formula, params):
        """执行自定义计算公式"""
        if formula == "MULTI_OPERATION":
            # 实现 (A表x * B表y) + C表z 的计算
            return (params['A_x'] * params['B_y']) + params['C_z']
        else:
            raise ValueError(f"未知公式: {formula}")

    def generate_report(self, output_name, operations):
        """生成报表"""
        # 加载模板文件
        report_wb = load_workbook(self.template_path)

        # 执行所有填表操作
        for op in operations:
            target_sheet_name = op['target']['sheet']

            # 检查目标工作表是否存在，不存在则创建
            if target_sheet_name not in report_wb.sheetnames:
                print(f"创建新工作表: {target_sheet_name}")
                report_wb.create_sheet(title=target_sheet_name)

            # 获取目标工作表
            target_sheet = report_wb[target_sheet_name]

            # 处理循环操作
            if 'loop' in op:
                loop_count = op['loop']['count']
                for i in range(loop_count):
                    # 计算当前循环的偏移量
                    col_shift = i * op['loop'].get('col_shift', 1)

                    # 准备参数
                    params = {}
                    for param_name, param_def in op['params'].items():
                        # 应用列偏移
                        shifted_cell = self.shift_cell(
                            param_def['cell'],
                            col_shift=col_shift
                        )
                        # 获取参数值
                        params[param_name] = self.get_cell_value(
                            param_def['sheet'],
                            shifted_cell
                        )

                    # 计算目标单元格位置（应用相同的偏移）
                    target_cell = self.shift_cell(
                        op['target']['cell'],
                        col_shift=col_shift
                    )

                    # 执行计算并写入结果
                    result = self.calculate(op['formula'], params)
                    target_sheet[target_cell] = result
                    print(f"写入结果: 工作表={target_sheet_name}, 单元格={target_cell}, 值={result}")
            else:
                # 非循环操作处理
                target_cell = op['target']['cell']

                if op['type'] == 'direct':
                    # 直接复制数据
                    value = self.get_cell_value(op['source']['sheet'], op['source']['cell'])
                    target_sheet[target_cell] = value

                elif op['type'] == 'formula':
                    # 执行公式计算
                    params = {}
                    for param_name, param_def in op['params'].items():
                        # 获取参数值
                        params[param_name] = self.get_cell_value(
                            param_def['sheet'],
                            param_def['cell']
                        )

                    # 执行计算并写入结果
                    result = self.calculate(op['formula'], params)
                    target_sheet[target_cell] = result

        # 保存报表
        os.makedirs(self.output_dir, exist_ok=True)
        output_path = os.path.join(self.output_dir, output_name)
        report_wb.save(output_path)
        print(f"报表已生成: {output_path}")
        return output_path


# 使用示例 =============================================
if __name__ == "__main__":
    # 初始化报表生成器
    reporter = ExcelReporter(
        data_file="多表循环计算.xlsx",  # 包含所有数据的单一文件
        template_path="财务参数表_20250801_1639.xlsx",   #需要输出的excel表的模板报表
        output_dir="./输出报表"
    )

    # 定义填表操作 - 循环处理多组数据
    operations = [
        # 循环操作：处理5组数据，每组数据右移一列
        {
            "type": "formula",
            "formula": "MULTI_OPERATION",
            "params": {
                "A_x": {  # A表的x值
                    "sheet": "Sheet1",
                    "cell": "F3"  # 起始位置
                },
                "B_y": {  # B表的y值
                    "sheet": "Sheet2",
                    "cell": "C3"  # 起始位置
                },
                "C_z": {  # C表的z值
                    "sheet": "Sheet3",
                    "cell": "E4"  # 起始位置
                }
            },
            "target": {
                "sheet": "Sheet4",  # 目标工作excel表的sheet名
                "cell": "F5"  # 目标起始位置
            },
            "loop": {
                "count": 3,  # 循环5次
                "col_shift": 1  # 每次右移一列
            }
        }
    ]

    # 生成报表
    reporter.generate_report(
        output_name="循环计算结果.xlsx",
        operations=operations
    )