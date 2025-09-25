from openpyxl import load_workbook as original_load_workbook


def load_workbook(filename, read_only=False, keep_vba=True,
                  data_only=False, keep_links=True):
    """默认保留 VBA 宏的 load_workbook 版本"""
    return original_load_workbook(
        filename,
        read_only=read_only,
        keep_vba=keep_vba,
        data_only=data_only,
        keep_links=keep_links
    )
