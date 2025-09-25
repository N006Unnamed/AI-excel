import re
import os
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, Protection
from findAndSet import find_cell, find_all_cells
from circulate_formula import ExcelFormulaGenerator
from loan_assignment import down_n_cells, struct_years
import datetime
from openpyxl_vba import load_workbook

# ====================== 公式处理工具 ======================
def column_to_index(col_letters):
    index = 0
    for char in col_letters.upper():
        index = index * 26 + (ord(char) - ord('A') + 1)
    return index


def index_to_column(index):
    letters = []
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        letters.append(chr(65 + remainder))
    return ''.join(reversed(letters))


def parse_cell_reference(ref):
    if '!' in ref:
        ref = ref.split('!')[-1]
    parts = re.match(r"^(\$?)([A-Za-z]+)(\$?)(\d+)$", ref)
    if not parts:
        return None
    col_abs = bool(parts.group(1))
    col_letters = parts.group(2)
    row_abs = bool(parts.group(3))
    row_num = int(parts.group(4))
    col_num = column_to_index(col_letters)
    return (row_num, col_num, row_abs, col_abs)


def format_cell_reference(row, col, row_abs, col_abs):
    col_letter = index_to_column(col)
    return f"{'$' if col_abs else ''}{col_letter}{'$' if row_abs else ''}{row}"


def update_formula(formula: str, row_offset: int, col_offset: int,
                   current_sheet_name: str = None) -> str:
    """
    正确解析并更新公式中的相对引用
      - 支持跨工作表（含中文、点号、空格）
      - 支持区域引用 A1:B2
      - 绝对引用 $A$1 保持不变
    """
    # 改进的正则表达式：更精确地匹配工作表名
    pattern = re.compile(
        r"""
        (?:  # 匹配工作表前缀（可选）
            '([^']+)'  # 1. 带引号的工作表名（单引号内任何字符）
            |           # 或
            ([a-zA-Z\u4e00-\u9fa5][\w\u4e00-\u9fa5\.]*)  # 2. 不带引号的工作表名（字母/中文开头，可含点号）
        )
        !    # 工作表名后的感叹号
        (    # 3. 单元格或区域引用
            (\$?[A-Z]+\$?\d+)       # 起始单元格
            (?:
                :(\$?[A-Z]+\$?\d+)  # 区域结束单元格（可选）
            )?
        )
        """,
        re.IGNORECASE | re.VERBOSE,
    )

    def shift(cell_ref: str) -> str:
        """偏移单个单元格地址"""
        parsed = parse_cell_reference(cell_ref)
        if not parsed:
            return cell_ref
        r, c, r_abs, c_abs = parsed
        new_r = (r if r_abs else r + row_offset)
        new_c = (c if c_abs else c + col_offset)
        return format_cell_reference(max(1, new_r), max(1, new_c), r_abs, c_abs)

    def repl(match):
        # 获取工作表名（带引号或不带引号的匹配组）
        sheet_name = match.group(1) or match.group(2)
        ref_full = match.group(3)  # 完整的引用部分
        start_ref = match.group(4)  # 起始单元格
        end_ref = match.group(5)  # 结束单元格（如果是区域）

        # 处理单元格引用部分
        if end_ref:  # 区域引用
            new_start = shift(start_ref)
            new_end = shift(end_ref)
            new_ref = f"{new_start}:{new_end}"
        else:  # 单格引用
            new_ref = shift(start_ref)

        # 重新组装引用（包含括号的工作表名也需要加单引号）
        # 注意：无论是否有特殊字符，都使用单引号格式
        return f"'{sheet_name}'!{new_ref}"

    # 处理所有工作表引用格式
    new_formula = pattern.sub(repl, formula)

    # 处理当前工作表的引用（无工作表前缀）
    current_sheet_pattern = re.compile(
        r"""
        (^|[(,=+*/-])  # 确保是公式中的独立引用（前面是运算符或开始）
        (               # 单元格或区域引用
            (\$?[A-Z]+\$?\d+)       # 起始单元格
            (?:
                :(\$?[A-Z]+\$?\d+)  # 区域结束单元格（可选）
            )?
        )
        (?=[),+*/-]|$)  # 后面是运算符或结束
        """,
        re.IGNORECASE | re.VERBOSE,
    )

    def repl_current(match):
        prefix = match.group(1)  # 前面的运算符
        ref_full = match.group(2)  # 完整的引用部分
        start_ref = match.group(3)  # 起始单元格
        end_ref = match.group(4)  # 结束单元格（如果是区域）

        # 处理单元格引用部分
        if end_ref:  # 区域引用
            new_start = shift(start_ref)
            new_end = shift(end_ref)
            new_ref = f"{new_start}:{new_end}"
        else:  # 单格引用
            new_ref = shift(start_ref)

        return prefix + new_ref

    new_formula = current_sheet_pattern.sub(repl_current, new_formula)

    return new_formula


# def copy_cell_style(source_cell, target_cell):
#     """
#     复制单元格样式从源单元格到目标单元格
#     包括字体、对齐、边框、填充等所有样式属性
#     """
#     # 复制字体
#     if source_cell.font:
#         target_cell.font = Font(
#             name=source_cell.font.name,
#             size=source_cell.font.size,
#             bold=source_cell.font.bold,
#             italic=source_cell.font.italic,
#             vertAlign=source_cell.font.vertAlign,
#             underline=source_cell.font.underline,
#             strike=source_cell.font.strike,
#             color=source_cell.font.color
#         )
#
#     # 复制对齐方式
#     if source_cell.alignment:
#         target_cell.alignment = Alignment(
#             horizontal=source_cell.alignment.horizontal,
#             vertical=source_cell.alignment.vertical,
#             text_rotation=source_cell.alignment.text_rotation,
#             wrap_text=source_cell.alignment.wrap_text,
#             shrink_to_fit=source_cell.alignment.shrink_to_fit,
#             indent=source_cell.alignment.indent
#         )
#
#     # 复制边框
#     if source_cell.border:
#         target_cell.border = Border(
#             left=source_cell.border.left,
#             right=source_cell.border.right,
#             top=source_cell.border.top,
#             bottom=source_cell.border.bottom,
#             diagonal=source_cell.border.diagonal,
#             diagonal_direction=source_cell.border.diagonal_direction,
#             outline=source_cell.border.outline,
#             vertical=source_cell.border.vertical,
#             horizontal=source_cell.border.horizontal
#         )
#
#     # 复制填充
#     if source_cell.fill:
#         target_cell.fill = PatternFill(
#             fill_type=source_cell.fill.fill_type,
#             start_color=source_cell.fill.start_color,
#             end_color=source_cell.fill.end_color
#         )
#
#     # 复制数字格式
#     target_cell.number_format = source_cell.number_format
#
#     # 复制保护设置
#     if source_cell.protection:
#         target_cell.protection = Protection(
#             locked=source_cell.protection.locked,
#             hidden=source_cell.protection.hidden
#         )


# ====================== Excel 自动填充工具 ======================
class ExcelAutoFiller:
    def __init__(self, file_path):
        self.file_path = file_path
        self.wb = load_workbook(file_path)
        self.sheet = None

    def set_active_sheet(self, sheet_name=None):
        if sheet_name:
            self.sheet = self.wb[sheet_name]
        else:
            self.sheet = self.wb.active

    def get_cell_formula(self, cell_address):
        if not self.sheet:
            raise ValueError("No active sheet selected")
        cell = self.sheet[cell_address]
        return cell.value if cell.data_type == 'f' else None

    def auto_fill(self, source_range, target_start, save=True, copy_style=True):
        if not self.sheet:
            raise ValueError("No active sheet selected")

        # 解析源区域
        if ':' in source_range:
            src_start, src_end = source_range.split(':')
            src_start_ref = parse_cell_reference(src_start)
            src_end_ref = parse_cell_reference(src_end)
            src_rows = src_end_ref[0] - src_start_ref[0] + 1
            src_cols = src_end_ref[1] - src_start_ref[1] + 1

            # 批量读取源区域数据
            src_data = []
            for row in range(src_start_ref[0], src_end_ref[0] + 1):
                row_data = []
                for col in range(src_start_ref[1], src_end_ref[1] + 1):
                    cell = self.sheet.cell(row=row, column=col)
                    row_data.append({
                        'value': cell.value,
                        'formula': cell.value if cell.data_type == 'f' else None,
                        'style': cell._style if copy_style else None
                    })
                src_data.append(row_data)
        else:
            # 单单元格处理
            cell_ref = parse_cell_reference(source_range)
            cell = self.sheet.cell(row=cell_ref[0], column=cell_ref[1])
            src_data = [[{
                'value': cell.value,
                'formula': cell.value if cell.data_type == 'f' else None,
                'style': cell._style if copy_style else None
            }]]
            src_rows, src_cols = 1, 1

        # 解析目标区域
        if ':' in target_start:
            tgt_start, tgt_end = target_start.split(':')
            tgt_start_ref = parse_cell_reference(tgt_start)
            tgt_end_ref = parse_cell_reference(tgt_end)
            tgt_rows = tgt_end_ref[0] - tgt_start_ref[0] + 1
            tgt_cols = tgt_end_ref[1] - tgt_start_ref[1] + 1
        else:
            tgt_start_ref = parse_cell_reference(target_start)
            tgt_rows = src_rows
            tgt_cols = src_cols
            tgt_end_ref = (tgt_start_ref[0] + tgt_rows - 1,
                           tgt_start_ref[1] + tgt_cols - 1)

        # 检查形状是否匹配（区域到区域填充时）
        if src_rows != 1 or src_cols != 1:
            if tgt_rows != src_rows or tgt_cols != src_cols:
                raise ValueError(
                    f"源区域({src_rows}×{src_cols})与目标区域({tgt_rows}×{tgt_cols})形状不一致"
                )

        # 批量处理目标区域
        for row_idx in range(tgt_rows):
            for col_idx in range(tgt_cols):
                # 计算源数据索引（考虑广播）
                src_row_idx = row_idx % src_rows
                src_col_idx = col_idx % src_cols

                src_cell_data = src_data[src_row_idx][src_col_idx]
                tgt_row = tgt_start_ref[0] + row_idx
                tgt_col = tgt_start_ref[1] + col_idx
                tgt_cell = self.sheet.cell(row=tgt_row, column=tgt_col)

                # 处理公式
                if src_cell_data['formula']:
                    formula = src_cell_data['formula']
                    if formula.startswith('='):
                        formula = formula[1:]

                    # 计算行和列偏移量
                    src_base_row = src_start_ref[0] if ':' in source_range else cell_ref[0]
                    src_base_col = src_start_ref[1] if ':' in source_range else cell_ref[1]
                    row_delta = tgt_row - src_base_row
                    col_delta = tgt_col - src_base_col

                    new_formula = update_formula(
                        formula,
                        row_delta,
                        col_delta,
                        current_sheet_name=self.sheet.title
                    )
                    tgt_cell.value = f"={new_formula}"
                else:
                    # 非公式单元格：直接复制值
                    tgt_cell.value = src_cell_data['value']

                # 复制样式
                if copy_style and src_cell_data['style']:
                    tgt_cell._style = src_cell_data['style']

        if save:
            self.save()

    # def auto_fill(self, source_range, target_start, save=True, copy_style=True):
    #     """
    #     自动填充公式，支持：
    #     1) 单格 → 单格
    #     2) 单格 → 区域（自动广播）
    #     3) 区域 → 区域（形状必须一致）
    #     """
    #     if not self.sheet:
    #         raise ValueError("No active sheet selected")
    #
    #     # ---------------- 解析源区域 ----------------
    #     if ':' in source_range:
    #         src_start_cell, src_end_cell = source_range.split(':')
    #         src_cells = self._get_cells_in_range(src_start_cell, src_end_cell)
    #     else:
    #         src_start_cell = src_end_cell = source_range
    #         src_cells = [source_range]
    #
    #     src_r1, src_c1, _, _ = parse_cell_reference(src_start_cell)
    #     src_r2, src_c2, _, _ = parse_cell_reference(src_end_cell)
    #     src_rows = src_r2 - src_r1 + 1
    #     src_cols = src_c2 - src_c1 + 1
    #
    #     # ---------------- 解析目标区域 ----------------
    #     if ':' in target_start:
    #         tgt_start_cell, tgt_end_cell = target_start.split(':')
    #         tgt_cells = self._get_cells_in_range(tgt_start_cell, tgt_end_cell)
    #         tgt_r1, tgt_c1, _, _ = parse_cell_reference(tgt_start_cell)
    #         tgt_r2, tgt_c2, _, _ = parse_cell_reference(tgt_end_cell)
    #         tgt_rows = tgt_r2 - tgt_r1 + 1
    #         tgt_cols = tgt_c2 - tgt_c1 + 1
    #     else:
    #         tgt_start_cell = target_start
    #         tgt_r1, tgt_c1, _, _ = parse_cell_reference(tgt_start_cell)
    #         tgt_r2 = tgt_r1 + src_rows - 1
    #         tgt_c2 = tgt_c1 + src_cols - 1
    #         tgt_cells = self._get_cells_in_range(
    #             tgt_start_cell, f"{index_to_column(tgt_c2)}{tgt_r2}"
    #         )
    #         tgt_rows, tgt_cols = src_rows, src_cols
    #
    #     # ---------- 关键：单格→区域广播 ----------
    #     if src_rows == 1 and src_cols == 1 and (tgt_rows > 1 or tgt_cols > 1):
    #         src_cells = [src_cells[0]] * len(tgt_cells)
    #
    #     # ---------- 形状校验（仅区域↔区域时） ----------
    #     if len(src_cells) != len(tgt_cells):
    #         raise ValueError(
    #             f"源区域({src_rows}×{src_cols})与目标区域({tgt_rows}×{tgt_cols})形状不一致"
    #         )
    #
    #     # ---------- 逐格写入 ----------
    #     for src_cell, tgt_cell in zip(src_cells, tgt_cells):
    #         formula = self.get_cell_formula(src_cell)
    #         src_cell_obj = self.sheet[src_cell]
    #         tgt_cell_obj = self.sheet[tgt_cell]
    #
    #         if formula:
    #             # 处理公式
    #             src_p = parse_cell_reference(src_cell)
    #             tgt_p = parse_cell_reference(tgt_cell)
    #             row_delta = tgt_p[0] - src_p[0]
    #             col_delta = tgt_p[1] - src_p[1]
    #             if formula.startswith('='):
    #                 formula = formula[1:]
    #             new_formula = update_formula(
    #                 formula,
    #                 row_delta,
    #                 col_delta,
    #                 current_sheet_name=self.sheet.title
    #             )
    #             self.sheet[tgt_cell] = f"={new_formula}"
    #         else:
    #             # 非公式单元格：直接复制值
    #             self.sheet[tgt_cell] = self.sheet[src_cell].value
    #         # 复制单元格样式
    #         if copy_style:
    #             copy_cell_style(src_cell_obj, tgt_cell_obj)
    #
    #     if save:
    #         self.save()

    # ----------------------------------------------------------
    def save(self, new_file_path=None):
        save_path = new_file_path or self.file_path
        self.wb.save(save_path)
        print(f"文件已保存: {save_path}")

    def _cell_offset(self, cell_address, row_offset, col_offset):
        parsed = parse_cell_reference(cell_address)
        if not parsed:
            return cell_address
        row, col, _, _ = parsed
        new_row = row + row_offset
        new_col = col + col_offset
        col_letter = index_to_column(new_col)
        return f"{col_letter}{new_row}"

    def _get_cells_in_range(self, start_cell, end_cell):
        start_p = parse_cell_reference(start_cell)
        end_p = parse_cell_reference(end_cell)
        if not start_p or not end_p:
            return [start_cell]
        sr, sc, _, _ = start_p
        er, ec, _, _ = end_p
        cells = []
        for r in range(sr, er + 1):
            for c in range(sc, ec + 1):
                cells.append(f"{get_column_letter(c)}{r}")
        return cells


# def copy_cell_style(source_cell, target_cell):
#     """复制源单元格的样式到目标单元格"""
#     # 复制字体样式
#     if source_cell.font:
#         target_cell.font = Font(
#             name=source_cell.font.name,
#             size=source_cell.font.size,
#             bold=source_cell.font.bold,
#             italic=source_cell.font.italic,
#             color=source_cell.font.color
#         )
#
#     # 复制填充（背景色）
#     if source_cell.fill:
#         target_cell.fill = PatternFill(
#             start_color=source_cell.fill.start_color,
#             end_color=source_cell.fill.end_color,
#             fill_type=source_cell.fill.fill_type
#         )
#
#     # 复制边框
#     if source_cell.border:
#         border = Border(
#             left=Side(border_style=source_cell.border.left.border_style, color=source_cell.border.left.color),
#             right=Side(border_style=source_cell.border.right.border_style, color=source_cell.border.right.color),
#             top=Side(border_style=source_cell.border.top.border_style, color=source_cell.border.top.color),
#             bottom=Side(border_style=source_cell.border.bottom.border_style, color=source_cell.border.bottom.color)
#         )
#         target_cell.border = border
#
#     # 复制对齐方式
#     if source_cell.alignment:
#         target_cell.alignment = Alignment(
#             horizontal=source_cell.alignment.horizontal,
#             vertical=source_cell.alignment.vertical,
#             wrap_text=source_cell.alignment.wrap_text,
#             shrink_to_fit=source_cell.alignment.shrink_to_fit,
#             indent=source_cell.alignment.indent
#         )
#
#     # 复制数字格式
#     target_cell.number_format = source_cell.number_format

# 创建样式对象缓存
_style_cache = {
    'fonts': {},
    'fills': {},
    'borders': {},
    'alignments': {}
}


def _get_font_key(font):
    """生成字体对象的唯一键"""
    if not font:
        return None
    return (font.name, font.size, font.bold, font.italic, str(font.color))


def _get_fill_key(fill):
    """生成填充对象的唯一键"""
    if not fill:
        return None
    return (fill.start_color, fill.end_color, fill.fill_type)


def _get_border_key(border):
    """生成边框对象的唯一键"""
    if not border:
        return None
    left = (border.left.border_style, str(border.left.color)) if border.left else None
    right = (border.right.border_style, str(border.right.color)) if border.right else None
    top = (border.top.border_style, str(border.top.color)) if border.top else None
    bottom = (border.bottom.border_style, str(border.bottom.color)) if border.bottom else None
    return (left, right, top, bottom)


def _get_alignment_key(alignment):
    """生成对齐对象的唯一键"""
    if not alignment:
        return None
    return (alignment.horizontal, alignment.vertical,
            alignment.wrap_text, alignment.shrink_to_fit, alignment.indent)


def copy_cell_style(source_cell, target_cell):
    """复制源单元格的样式到目标单元格（优化版）"""
    # 复制字体样式（使用缓存）
    if source_cell.font:
        font_key = _get_font_key(source_cell.font)
        if font_key in _style_cache['fonts']:
            target_cell.font = _style_cache['fonts'][font_key]
        else:
            new_font = Font(
                name=source_cell.font.name,
                size=source_cell.font.size,
                bold=source_cell.font.bold,
                italic=source_cell.font.italic,
                color=source_cell.font.color
            )
            _style_cache['fonts'][font_key] = new_font
            target_cell.font = new_font
    elif target_cell.font:  # 清除目标单元格的字体样式
        target_cell.font = None

    # 复制填充（使用缓存）
    if source_cell.fill:
        fill_key = _get_fill_key(source_cell.fill)
        if fill_key in _style_cache['fills']:
            target_cell.fill = _style_cache['fills'][fill_key]
        else:
            new_fill = PatternFill(
                start_color=source_cell.fill.start_color,
                end_color=source_cell.fill.end_color,
                fill_type=source_cell.fill.fill_type
            )
            _style_cache['fills'][fill_key] = new_fill
            target_cell.fill = new_fill
    elif target_cell.fill:  # 清除目标单元格的填充样式
        target_cell.fill = None

    # 复制边框（使用缓存）
    if source_cell.border:
        border_key = _get_border_key(source_cell.border)
        if border_key in _style_cache['borders']:
            target_cell.border = _style_cache['borders'][border_key]
        else:
            # 安全地处理可能为None的边框边
            left_side = source_cell.border.left
            right_side = source_cell.border.right
            top_side = source_cell.border.top
            bottom_side = source_cell.border.bottom

            new_border = Border(
                left=Side(
                    border_style=left_side.border_style,
                    color=left_side.color
                ) if left_side else None,
                right=Side(
                    border_style=right_side.border_style,
                    color=right_side.color
                ) if right_side else None,
                top=Side(
                    border_style=top_side.border_style,
                    color=top_side.color
                ) if top_side else None,
                bottom=Side(
                    border_style=bottom_side.border_style,
                    color=bottom_side.color
                ) if bottom_side else None
            )
            _style_cache['borders'][border_key] = new_border
            target_cell.border = new_border
    elif target_cell.border:  # 清除目标单元格的边框样式
        target_cell.border = None

    # 复制对齐方式（使用缓存）
    if source_cell.alignment:
        alignment_key = _get_alignment_key(source_cell.alignment)
        if alignment_key in _style_cache['alignments']:
            target_cell.alignment = _style_cache['alignments'][alignment_key]
        else:
            new_alignment = Alignment(
                horizontal=source_cell.alignment.horizontal,
                vertical=source_cell.alignment.vertical,
                wrap_text=source_cell.alignment.wrap_text,
                shrink_to_fit=source_cell.alignment.shrink_to_fit,
                indent=source_cell.alignment.indent
            )
            _style_cache['alignments'][alignment_key] = new_alignment
            target_cell.alignment = new_alignment
    elif target_cell.alignment:  # 清除目标单元格的对齐样式
        target_cell.alignment = None

    # 复制数字格式
    target_cell.number_format = source_cell.number_format


# 可选：添加一个函数来清空缓存，以防内存占用过大
def clear_style_cache():
    """清空样式缓存"""
    for cache in _style_cache.values():
        cache.clear()


def find_last_used_column(excel_file_path, sheet_name=0):
    """
    在Excel工作表中找到最后一列有数据的列

    参数:
    excel_file_path (str): Excel文件的路径
    sheet_name (str/int): 工作表名称或索引，默认为第一个工作表

    返回:
    int: 最后一列的索引（从1开始）
    str: 最后一列的字母表示（如"A", "B", "C"）
    """
    try:
        # 使用openpyxl直接加载工作簿以获得更精确的单元格信息
        wb = load_workbook(excel_file_path, data_only=True)

        # 根据sheet_name参数选择工作表
        if isinstance(sheet_name, int):
            sheet = wb.worksheets[sheet_name]
        else:
            sheet = wb[sheet_name]

        # 初始化最后一列为0
        last_column = 0

        # 遍历每一行，查找有数据的最后一列
        for row in sheet.iter_rows():
            for cell in reversed(row):  # 从右向左检查以提高效率
                if cell.value is not None and str(cell.value).strip() != '':
                    if cell.column > last_column:
                        last_column = cell.column
                    break  # 找到该行最后一个有数据的单元格后跳出内层循环

        # 将列索引转换为字母表示
        column_letter = openpyxl.utils.get_column_letter(last_column)

        return last_column, column_letter

    except Exception as e:
        print(f"处理Excel文件时出错: {e}")
        return 0, ""


def find_last_used_row(excel_file_path, sheet_name=0):
    """
    在Excel工作表中找到最后一行有数据的行

    参数:
    excel_file_path (str): Excel文件的路径
    sheet_name (str/int): 工作表名称或索引，默认为第一个工作表

    返回:
    int: 最后一行的索引（从1开始）
    """
    try:
        # 使用openpyxl直接加载工作簿以获得更精确的单元格信息
        wb = load_workbook(excel_file_path, data_only=True)

        # 根据sheet_name参数选择工作表
        if isinstance(sheet_name, int):
            sheet = wb.worksheets[sheet_name]
        else:
            sheet = wb[sheet_name]

        # 初始化最后一行为0
        last_row = 0

        # 遍历每一列，查找有数据的最后一行
        for col in sheet.iter_cols():
            for cell in reversed(col):  # 从下向上检查以提高效率
                if cell.value is not None and str(cell.value).strip() != '':
                    if cell.row > last_row:
                        last_row = cell.row
                    break  # 找到该列最后一个有数据的单元格后跳出内层循环

        return last_row

    except Exception as e:
        print(f"处理Excel文件时出错: {e}")
        return 0


def expand_excel_header(input_file, sheet_name, output_file, target_year, header_row=2, start_col=None):
    """
    扩展Excel表头的年份到指定目标年份，并保持原始格式

    参数:
    input_file: 输入Excel文件路径
    output_file: 输出Excel文件路径
    target_year: 目标年份（整数）
    header_row: 表头所在行号（默认为2）
    start_col: 起始列位置（列字母或列号），None表示自动检测最后一列
    """
    # 加载工作簿
    wb = load_workbook(input_file)
    ws = wb[sheet_name]

    # 确定起始列索引
    if start_col is None:
        # 自动检测最后一列
        current_col, _ = find_last_used_column(input_file, sheet_name)
        print(f"自动检测到最后一列: {get_column_letter(current_col)}")
    else:
        # 处理手动指定的列
        if isinstance(start_col, str):
            current_col = openpyxl.utils.column_index_from_string(start_col)
        else:
            current_col = int(start_col)
        print(f"使用指定列: {get_column_letter(current_col)}")

    # 获取当前年份和源单元格
    source_cell = ws.cell(row=header_row, column=current_col)
    try:
        current_year = int(source_cell.value)
        print(f"当前最后一列的年份: {current_year}")
    except (ValueError, TypeError):
        raise ValueError(f"起始列的值不是有效年份: {source_cell.value}")

    # 验证年份范围
    if current_year >= target_year:
        raise ValueError(f"目标年份({target_year})必须大于当前年份({current_year})")

    # 扩展年份到目标年份
    print("正在扩展年份并复制格式...")
    for year in range(current_year + 1, target_year + 1):
        current_col += 1
        new_cell = ws.cell(row=header_row, column=current_col, value=year)

        # 复制源单元格的样式
        copy_cell_style(source_cell, new_cell)

        print(f"添加年份: {year} 到列 {get_column_letter(current_col)} (已复制格式)")

    # 保存结果
    wb.save(output_file)
    print(f"\n扩展完成! 文件已保存至: {os.path.abspath(output_file)}")
    print(f"共添加了 {target_year - current_year} 个年份")


# ====================== 使用示例 ======================
def just_copy(input_path, target_year):
    # 用户输入
    output_path = input_path
    header_row = 2
    sheet_names = ["b利润与利润分配表（损益和利润分配表）", "d财务计划现金流量表", "e资产负债表"]
    # sheet_name = "a.1财务现金流量表"
    # sheet_name = "b利润与利润分配表（损益和利润分配表）"

    for sheet_name in sheet_names:
        _, start_col = find_last_used_column(input_path, sheet_name)
        start_col_plus = column_index_from_string(start_col) + 1
        start_col_plus = get_column_letter(start_col_plus)
        # start_col = "H"
        print(start_col)
        print(start_col_plus)

        # 执行扩展
        try:
            expand_excel_header(
                input_file=input_path,
                sheet_name=sheet_name,
                output_file=output_path,
                target_year=target_year,
                header_row=header_row,
                start_col=start_col
            )
        except Exception as e:
            print(f"\n错误: {str(e)}")
            print("操作未完成，请检查输入参数是否正确")

        r, c = find_cell(input_path, target_year, sheet_name)

        filler = ExcelAutoFiller(input_path)
        filler.set_active_sheet(sheet_name)

        if sheet_name == "b利润与利润分配表（损益和利润分配表）":
            last_row = find_last_used_row(input_path, sheet_name)
        else:
            last_row = find_last_used_row(input_path, sheet_name) + 1

        for i in range(r + 1, last_row):
            filler.auto_fill(f"{start_col}{i}", f"{start_col_plus}{i}:{c}{i}")

        # 填充合计
        summary(input_path, sheet_name, last_row, c)

        if sheet_name == "b利润与利润分配表（损益和利润分配表）":
            table_b(input_path, sheet_name, c)
        elif sheet_name == "d财务计划现金流量表":
            table_d(input_path, sheet_name, c)
        else:
            table_e(input_path, sheet_name, c)

    # print("将 G6 的公式填充到 J6:K6...")
    # filler.auto_fill("G6", "J6:K6")


def special_copy(input_path, target_year):
    # 用户输入
    output_path = input_path
    header_row = 2
    sheet_names = ["a.1财务现金流量表", "a.2项目资本金现金流量表"]
    # sheet_name = "a.1财务现金流量表"
    # sheet_name = "b利润与利润分配表（损益和利润分配表）"

    for sheet_name in sheet_names:
        _, start_col = find_last_used_column(input_path, sheet_name)
        start_col_plus = column_index_from_string(start_col) + 1
        start_col_plus = get_column_letter(start_col_plus)
        start_col_minus = column_index_from_string(start_col) - 1
        start_col_minus = get_column_letter(start_col_minus)
        # start_col = "H"
        print(start_col)
        print(start_col_plus)
        print(start_col_minus)

        # 获取 “回收固定资产余值” 和 “回收流动资产” 的地址，用于取公式
        wb = load_workbook(input_path)
        ws = wb[sheet_name]
        value1 = "回收固定资产余值"
        value2 = "回收流动资金"
        r1, _ = find_cell(input_path, value1, sheet_name)
        r2, _ = find_cell(input_path, value2, sheet_name)
        value1_address = start_col + str(r1)
        value2_address = start_col + str(r2)
        # 获取 “回收固定资产余值” 的公式，用于在最后一列中替换
        value1_formula = ws[value1_address]
        value1_formula = value1_formula.value if value1_formula.data_type == 'f' else None
        print("value1_formula:", value1_formula)
        # 获取 “回收流动资金” 的公式，用于在最后一列中替换
        value2_formula = ws[value2_address]
        value2_formula = value2_formula.value if value2_formula.data_type == 'f' else None
        print("value2_formula:", value2_formula)
        wb.close()

        # start_col = input("请输入起始列(列字母或列号，直接回车自动检测): ")

        # 处理可选参数
        # header_row = int(header_row) if header_row.strip() else 2
        # start_col = start_col.strip() or None

        # 执行扩展
        try:
            expand_excel_header(
                input_file=input_path,
                sheet_name=sheet_name,
                output_file=output_path,
                target_year=target_year,
                header_row=header_row,
                start_col=start_col_minus
            )
        except Exception as e:
            print(f"\n错误: {str(e)}")
            print("操作未完成，请检查输入参数是否正确")

        r, c = find_cell(input_path, target_year, sheet_name)

        filler = ExcelAutoFiller(input_path)
        filler.set_active_sheet(sheet_name)
        last_row = 0
        if sheet_name == "a.1财务现金流量表":
            last_row = find_last_used_row(input_path, sheet_name) - 7
        else:
            last_row = find_last_used_row(input_path, sheet_name)

        # 横向复制公式
        for i in range(r + 1, last_row):
            filler.auto_fill(f"{start_col_minus}{i}", f"{start_col}{i}:{c}{i}")

        # 填写最后一列的特殊项
        wb = load_workbook(input_path)
        ws = wb[sheet_name]
        # 找到扩展后的最后一年的列序号
        _, c = find_cell(input_path, target_year, sheet_name)
        value1_address = c + str(r1)
        value2_address = c + str(r2)
        table_D4_sheet = "D.4流动资金估算表"
        _, temp_c = find_last_used_column(input_path, table_D4_sheet)
        value = "流动资金（1-2）"
        r, _ = find_cell(input_path, value, table_D4_sheet)
        address_D4 = temp_c + str(r)

        # 填入之前的公式
        ws[value1_address].value = value1_formula
        # ws[value2_address].value = value2_formula
        ws[value2_address].value = f"={table_D4_sheet}!{address_D4}"
        wb.save(input_path)
        wb.close()

        # 填充合计
        summary(input_path, sheet_name, last_row, c)

        if sheet_name == "a.1财务现金流量表":
            table_a1(input_path, sheet_name, c)
        else:
            table_a2(input_path, sheet_name, c)


def summary(file_path, sheet_name, last_row, last_col):
    value = "合计"
    r, c = find_cell(file_path, value, sheet_name)
    if r and c:
        # 初始化公式生成器（输入输出为同一个文件）
        formula_generator = ExcelFormulaGenerator(
            data_file=file_path,  # 包含所有数据的单一文件
            output_file=file_path  # 输出到同一个文件
        )

        target_address = c + str(r + 1)
        start_address = get_column_letter(column_index_from_string(c) + 1) + str(r + 1)
        end_address = last_col + str(r + 1)

        config = [
            {
                "operation": "custom",
                "formula_template": "=(SUM({A_val}:{B_val}))",
                "params": {
                    "A_val": {"sheet": "", "cell": start_address},  # A表的y (向右移动)
                    "B_val": {"sheet": "", "cell": end_address},  # A表的y (向右移动)

                },
                "target": {
                    "sheet": sheet_name,  # 结果工作表
                    "cell": target_address  # 目标起始位置 (向右移动)
                },
                "loop": {
                    "count": last_row - r - 1,  # 循环3次

                    # 参数独立偏移设置
                    "param_offsets": {
                        "A_val": {"row_shift": 1, "col_shift": 0},  # 每次向下移动一行
                        "B_val": {"row_shift": 1, "col_shift": 0},  # 每次向下移动一行
                    },

                    # 目标单元格偏移设置
                    "target_offset": {"row_shift": 1, "col_shift": 0}  # 每次向右移动一列且向下移动一行
                }
            }
        ]

        formula_generator.generate_formulas(config)


def table_b(file_path, sheet_name, last_col):
    value = "当期还本付息"
    table_c_sheet = "c借款还本付息计划表"
    address = find_all_cells(file_path, value, table_c_sheet)
    address = address[-1]
    address = "C" + str(address[0])
    target_address = "C27"
    target_end = last_col + str(27)

    # 找到表b的 息税前利润 起始位置
    table_B_sheet = "B项目信息"
    start_year, end_year = struct_years(file_path, table_B_sheet)
    print("建设期的区间为：", start_year, "-", end_year)
    _, c1 = find_cell(file_path, start_year, sheet_name)
    _, c2 = find_cell(file_path, end_year, sheet_name)

    # 第一段与第二段的起始位置
    struct_target = c1 + str(25)
    c2 = get_column_letter(column_index_from_string(c2) + 1)
    struct_target2 = c2 + str(25)
    loop_num = column_index_from_string(last_col) - column_index_from_string(c2)

    # 找到表b中引用的表c的单元格
    value1 = "债券发行及服务费"
    address1 = find_all_cells(file_path, value1, table_c_sheet)
    address1 = address1[-1]
    address1 = "D" + str(address1[0])

    value2 = "应付利息"
    address2 = find_all_cells(file_path, value2, table_c_sheet)
    address2 = address2[-1]
    address2 = "D" + str(address2[0])

    value3 = "还本付息兑付手续费"
    address3 = find_all_cells(file_path, value3, table_c_sheet)
    address3 = address3[-1]
    address3 = "D" + str(address3[0])

    # 找到表b中引用的表E的单元格
    table_E_sheet = "E总成本费用估算表"
    _, c3 = find_cell(file_path, end_year + 1, table_E_sheet)

    # 1 营业收入
    table_F_sheet = "F项目收入"
    table_F_value = "年度合计（不含税）"
    r, c = find_cell(file_path, table_F_value, table_F_sheet)
    c = get_column_letter(column_index_from_string(c) + 2)
    table_F_address = c + str(r)

    # 4 补贴收入 (默认表F中存在"补贴收入")
    table_F_value2 = "补贴收入"
    r, c = find_cell(file_path, table_F_value2, table_F_sheet)
    c = get_column_letter(column_index_from_string(c) + 2)
    table_F_address2 = c + str(r)

    # 2 营业税金及附加
    table_G_sheet = "G.税金及附加测算表"
    table_G_value = "城建税及教育费附加"
    r, c = find_cell(file_path, table_G_value, table_G_sheet)
    c = get_column_letter(column_index_from_string(c) + 2)
    table_G_address = c + str(r)

    # 初始化公式生成器（输入输出为同一个文件）
    formula_generator = ExcelFormulaGenerator(
        data_file=file_path,  # 包含所有数据的单一文件
        output_file=file_path  # 输出到同一个文件
    )

    config = [
        {
            "operation": "custom",
            "formula_template": "=((C26-C11)/{A_val})",
            "params": {
                "A_val": {"sheet": table_c_sheet, "cell": address}  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": target_address  # 目标起始位置 (向右移动)
            }
        },
        # 息税前利润 (第一段引用表c部分)
        {
            "operation": "custom",
            "formula_template": "=({A_val}+{B_val}+{C_val}+{D_val})",
            "params": {
                "A_val": {"sheet": "", "cell": "D7"},  # A表的y (向右移动)
                "B_val": {"sheet": table_c_sheet, "cell": address1},  # A表的y (向右移动)
                "C_val": {"sheet": table_c_sheet, "cell": address2},  # A表的y (向右移动)
                "D_val": {"sheet": table_c_sheet, "cell": address3}  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": struct_target  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": end_year - start_year + 1,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                    "B_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                    "C_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                    "D_val": {"row_shift": 0, "col_shift": 1}  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        },
        # 息税前利润 (第二段)
        {
            "operation": "custom",
            "formula_template": "=({A_val}+{B_val})",
            "params": {
                "A_val": {"sheet": "", "cell": c2 + str(7)},  # A表的y (向右移动)
                "B_val": {"sheet": table_E_sheet, "cell": c3 + str(14)}  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": struct_target2  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                    "B_val": {"row_shift": 0, "col_shift": 1}  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        },

        # 营业收入（不含增值税）
        {
            "operation": "custom",
            "formula_template": "=({A_val})",
            "params": {
                "A_val": {"sheet": table_F_sheet, "cell": table_F_address},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": "D3"  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        },

        # 营业税金及附加
        {
            "operation": "custom",
            "formula_template": "=({A_val})",
            "params": {
                "A_val": {"sheet": table_G_sheet, "cell": table_G_address},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": "D4"  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        },

        # 补贴收入
        {
            "operation": "custom",
            "formula_template": "=({A_val})",
            "params": {
                "A_val": {"sheet": table_F_sheet, "cell": table_F_address2},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": "D6"  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        },
    ]

    formula_generator.generate_formulas(config)

    # 合并单元格

    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    # 创建居中对齐样式
    alignment_center = Alignment(horizontal='center', vertical='center')
    # 设置全边框
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    ws.merge_cells(range_string=f"{target_address}:{target_end}")
    # 设置指定区域单元格居中对齐
    for row in ws[f"{target_address}:{target_end}"]:
        for cell in row:
            cell.alignment = alignment_center
            cell.border = border

    wb.save(file_path)
    wb.close()


def table_d(file_path, sheet_name, last_col):
    value = "其中：还本"
    table_c_sheet = "c借款还本付息计划表"
    address = find_all_cells(file_path, value, table_c_sheet)
    address = address[-1]
    address1 = "D" + str(address[0] + 1)
    address2 = "D" + str(address[0])
    target_address = "D28"

    loop_num = column_index_from_string(last_col) - column_index_from_string("D")

    # 1.1.1 营业收入
    table_F_sheet = "F项目收入"
    table_F_value = "年度合计（不含税）"
    r, c = find_cell(file_path, table_F_value, table_F_sheet)
    c = get_column_letter(column_index_from_string(c) + 2)
    table_F_address = c + str(r)

    # 1.1.2 增值税销项税额
    table_G_sheet = "G.税金及附加测算表"
    table_G_value3 = "当期销项税额"
    r, c = find_cell(file_path, table_G_value3, table_G_sheet)
    c = get_column_letter(column_index_from_string(c) + 2)
    table_G_address3 = c + str(r)

    # 1.1.3 补贴收入 (默认表F中存在"补贴收入")
    table_F_value2 = "补贴收入"
    r, c = find_cell(file_path, table_F_value2, table_F_sheet)
    c = get_column_letter(column_index_from_string(c) + 2)
    table_F_address2 = c + str(r)

    # 1.2.2 增值税进项税额
    table_G_sheet = "G.税金及附加测算表"
    table_G_value4 = "当期进项税额"
    r, c = find_cell(file_path, table_G_value4, table_G_sheet)
    c = get_column_letter(column_index_from_string(c) + 2)
    table_G_address4 = c + str(r)

    # 1.2.3 营业税金及附加
    table_G_sheet = "G.税金及附加测算表"
    table_G_value = "城建税及教育费附加"
    r, c = find_cell(file_path, table_G_value, table_G_sheet)
    c = get_column_letter(column_index_from_string(c) + 2)
    table_G_address = c + str(r)

    # 1.2.4 增值税
    table_G_value2 = "当期应缴增值税"
    r, c = find_cell(file_path, table_G_value2, table_G_sheet)
    c = get_column_letter(column_index_from_string(c) + 2)
    table_G_address2 = c + str(r)

    # 2.2.1 建设投资
    table_E4_sheet = "E.4建设投资税后金额表"
    table_E4_value = "增值税率"
    r, c = find_cell(file_path, table_E4_value, table_E4_sheet)
    c = get_column_letter(column_index_from_string(c) + 1)
    r = find_last_used_row(file_path, table_E4_sheet)
    table_E4_address = c + str(r)

    # 初始化公式生成器（输入输出为同一个文件）
    formula_generator = ExcelFormulaGenerator(
        data_file=file_path,  # 包含所有数据的单一文件
        output_file=file_path  # 输出到同一个文件
    )

    config = [
        # 各种利息支出
        {
            "operation": "custom",
            "formula_template": "=({A_val}+{B_val}+{C_val})",
            "params": {
                "A_val": {"sheet": table_c_sheet, "cell": address1},  # A表的y (向右移动)
                "B_val": {"sheet": table_c_sheet, "cell": down_n_cells(address1, 2)},  # A表的y (向右移动)
                "C_val": {"sheet": table_c_sheet, "cell": down_n_cells(address1, 3)}  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": target_address  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num + 1,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                    "B_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                    "C_val": {"row_shift": 0, "col_shift": 1}  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        },
        # 偿还债务本金
        {
            "operation": "custom",
            "formula_template": "=({A_val})",
            "params": {
                "A_val": {"sheet": table_c_sheet, "cell": address2}  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": down_n_cells(target_address, 1)  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num + 1,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        },

        # 营业收入
        {
            "operation": "custom",
            "formula_template": "=({A_val})",
            "params": {
                "A_val": {"sheet": table_F_sheet, "cell": table_F_address},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": "D5"  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        },

        # 增值税销项税额
        {
            "operation": "custom",
            "formula_template": "=({A_val})",
            "params": {
                "A_val": {"sheet": table_G_sheet, "cell": table_G_address3},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": "D6"  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        },

        # 补贴收入
        {
            "operation": "custom",
            "formula_template": "=({A_val})",
            "params": {
                "A_val": {"sheet": table_F_sheet, "cell": table_F_address2},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": "D7"  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        },

        # 增值税进项税额
        {
            "operation": "custom",
            "formula_template": "=({A_val})",
            "params": {
                "A_val": {"sheet": table_G_sheet, "cell": table_G_address4},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": "D11"  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        },

        # 营业税金及附加
        {
            "operation": "custom",
            "formula_template": "=({A_val})",
            "params": {
                "A_val": {"sheet": table_G_sheet, "cell": table_G_address},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": "D12"  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        },

        # 增值税
        {
            "operation": "custom",
            "formula_template": "=({A_val})",
            "params": {
                "A_val": {"sheet": table_G_sheet, "cell": table_G_address2},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": "D13"  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        },

        # 建设投资
        {
            "operation": "custom",
            "formula_template": "=({A_val})",
            "params": {
                "A_val": {"sheet": table_E4_sheet, "cell": table_E4_address},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": "D19"  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        }
    ]

    formula_generator.generate_formulas(config)


def table_e(file_path, sheet_name, last_col):
    value = "期末借款余额"
    table_c_sheet = "c借款还本付息计划表"
    address = find_all_cells(file_path, value, table_c_sheet)
    address = address[-1]
    address = "D" + str(address[0])
    target_address = "C19"

    loop_num = column_index_from_string(last_col) - column_index_from_string("C")

    # 1.3 固定资产净值
    table_E2_sheet = "E.2固定资产折旧费估算表"
    table_E2_value = "净值合计"
    r, c = find_cell(file_path, table_E2_value, table_E2_sheet)
    c = get_column_letter(column_index_from_string(c) + 5)
    table_E2_address = c + str(r)

    # 1.4 无形及其他资产净值
    table_E3_sheet = "E.3无形资产和其他资产摊销费估算表"
    table_E3_value = "净值合计"
    r, c = find_cell(file_path, table_E3_value, table_E3_sheet)
    c = "F"
    table_E3_address = c + str(r)

    # 初始化公式生成器（输入输出为同一个文件）
    formula_generator = ExcelFormulaGenerator(
        data_file=file_path,  # 包含所有数据的单一文件
        output_file=file_path  # 输出到同一个文件
    )

    config = [
        # 建设投资借款
        {
            "operation": "custom",
            "formula_template": "=({A_val})",
            "params": {
                "A_val": {"sheet": table_c_sheet, "cell": address}  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": target_address  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num + 1,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        },

        # 固定资产净值
        {
            "operation": "custom",
            "formula_template": "=({A_val})",
            "params": {
                "A_val": {"sheet": table_E2_sheet, "cell": table_E2_address}  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": "C11"  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num + 1,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        },

        # 无形及其他资产净值
        {
            "operation": "custom",
            "formula_template": "=({A_val})",
            "params": {
                "A_val": {"sheet": table_E3_sheet, "cell": table_E3_address}  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": "C12"  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num + 1,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        }
    ]

    formula_generator.generate_formulas(config)


def table_a1(file_path, sheet_name, last_col):
    # 初始化公式生成器（输入输出为同一个文件）
    formula_generator = ExcelFormulaGenerator(
        data_file=file_path,  # 包含所有数据的单一文件
        output_file=file_path  # 输出到同一个文件
    )

    value = "Ⅰ"
    r, _ = find_cell(file_path, value, sheet_name)

    start_address1 = "D15"
    end_address1 = last_col + str(15)

    start_address2 = "D18"
    end_address2 = last_col + str(18)

    target_start = "D" + str(r)
    target_end = last_col + str(r)

    loop_num = column_index_from_string(last_col) - column_index_from_string("D")

    # 1.1 营业收入
    table_F_sheet = "F项目收入"
    table_F_value = "年度合计（不含税）"
    r, c = find_cell(file_path, table_F_value, table_F_sheet)
    c = get_column_letter(column_index_from_string(c) + 2)
    table_F_address = c + str(r)

    # 1.2 补贴收入 (默认表F中存在"补贴收入")
    table_F_value2 = "补贴收入"
    r, c = find_cell(file_path, table_F_value2, table_F_sheet)
    c = get_column_letter(column_index_from_string(c) + 2)
    table_F_address2 = c + str(r)

    # 2.1 建设投资
    table_E4_sheet = "E.4建设投资税后金额表"
    table_E4_value = "增值税率"
    r, c = find_cell(file_path, table_E4_value, table_E4_sheet)
    c = get_column_letter(column_index_from_string(c) + 1)
    r = find_last_used_row(file_path, table_E4_sheet)
    table_E4_address = c + str(r)

    # 2.4 营业税金及附加
    table_G_sheet = "G.税金及附加测算表"
    table_G_value = "城建税及教育费附加"
    r, c = find_cell(file_path, table_G_value, table_G_sheet)
    c = get_column_letter(column_index_from_string(c) + 2)
    table_G_address = c + str(r)

    # 2.5 增值税
    table_G_value2 = "当期应缴增值税"
    r, c = find_cell(file_path, table_G_value2, table_G_sheet)
    c = get_column_letter(column_index_from_string(c) + 2)
    table_G_address2 = c + str(r)

    config = [
        {
            "operation": "custom",
            "formula_template": "=(IRR({A_val}:{B_val}))",
            "params": {
                "A_val": {"sheet": "", "cell": start_address1},  # A表的y (向右移动)
                "B_val": {"sheet": "", "cell": end_address1},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": target_start  # 目标起始位置 (向右移动)
            }
        },

        {
            "operation": "custom",
            "formula_template": "=(IRR({A_val}:{B_val}))",
            "params": {
                "A_val": {"sheet": "", "cell": start_address2},  # A表的y (向右移动)
                "B_val": {"sheet": "", "cell": end_address2},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": down_n_cells(target_start, 1)  # 目标起始位置 (向右移动)
            }
        },

        {
            "operation": "custom",
            "formula_template": "=(NPV(A财务假设!$D$3,{A_val}:{B_val}))",
            "params": {
                "A_val": {"sheet": "", "cell": start_address1},  # A表的y (向右移动)
                "B_val": {"sheet": "", "cell": end_address1},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": down_n_cells(target_start, 2)  # 目标起始位置 (向右移动)
            }
        },

        {
            "operation": "custom",
            "formula_template": "=(NPV(A财务假设!$D$4,{A_val}:{B_val}))",
            "params": {
                "A_val": {"sheet": "", "cell": start_address2},  # A表的y (向右移动)
                "B_val": {"sheet": "", "cell": end_address2},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": down_n_cells(target_start, 3)  # 目标起始位置 (向右移动)
            }
        },

        # 营业收入
        {
            "operation": "custom",
            "formula_template": "=({A_val})",
            "params": {
                "A_val": {"sheet": table_F_sheet, "cell": table_F_address},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": "D4"  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        },

        # 补贴收入
        {
            "operation": "custom",
            "formula_template": "=({A_val})",
            "params": {
                "A_val": {"sheet": table_F_sheet, "cell": table_F_address2},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": "D5"  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        },

        # 建设投资
        {
            "operation": "custom",
            "formula_template": "=({A_val})",
            "params": {
                "A_val": {"sheet": table_E4_sheet, "cell": table_E4_address},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": "D9"  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        },

        # 营业税金及附加
        {
            "operation": "custom",
            "formula_template": "=({A_val})",
            "params": {
                "A_val": {"sheet": table_G_sheet, "cell": table_G_address},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": "D12"  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        },

        # 增值税
        {
            "operation": "custom",
            "formula_template": "=({A_val})",
            "params": {
                "A_val": {"sheet": table_G_sheet, "cell": table_G_address2},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": "D13"  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        }
    ]

    formula_generator.generate_formulas(config)

    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    # 创建居中对齐样式
    alignment_center = Alignment(horizontal='center', vertical='center')
    # 设置全边框
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for i in range(4):
        merge_start = down_n_cells(target_start, i)
        merge_end = down_n_cells(target_end, i)
        ws.merge_cells(range_string=f"{merge_start}:{merge_end}")
        # 设置指定区域单元格居中对齐
        for row in ws[f"{merge_start}:{merge_end}"]:
            for cell in row:
                cell.alignment = alignment_center
                cell.border = border

    wb.save(file_path)
    wb.close()


def table_a2(file_path, sheet_name, last_col):
    # 初始化公式生成器（输入输出为同一个文件）
    formula_generator = ExcelFormulaGenerator(
        data_file=file_path,  # 包含所有数据的单一文件
        output_file=file_path  # 输出到同一个文件
    )

    value = "资本金财务内部收益率（%）"
    r, _ = find_cell(file_path, value, sheet_name)

    start_address = "D17"
    end_address = last_col + str(17)

    target_start = "D" + str(r)
    target_end = last_col + str(r)

    value = "其中：还本"
    table_c_sheet = "c借款还本付息计划表"
    address = find_all_cells(file_path, value, table_c_sheet)
    address = address[-1]
    # 利息
    address1 = "D" + str(address[0] + 1)
    # 本金
    address2 = "D" + str(address[0])

    loop_num = column_index_from_string(last_col) - column_index_from_string("D")
    target_address = "D10"

    # 1.1 营业收入
    table_F_sheet = "F项目收入"
    table_F_value = "年度合计（不含税）"
    r, c = find_cell(file_path, table_F_value, table_F_sheet)
    c = get_column_letter(column_index_from_string(c) + 2)
    table_F_address = c + str(r)

    # 1.2 补贴收入 (默认表F中存在"补贴收入")
    table_F_value2 = "补贴收入"
    r, c = find_cell(file_path, table_F_value2, table_F_sheet)
    c = get_column_letter(column_index_from_string(c) + 2)
    table_F_address2 = c + str(r)

    # 2.5 营业税金及附加
    table_G_sheet = "G.税金及附加测算表"
    table_G_value = "城建税及教育费附加"
    r, c = find_cell(file_path, table_G_value, table_G_sheet)
    c = get_column_letter(column_index_from_string(c) + 2)
    table_G_address = c + str(r)

    # 2.6 增值税
    table_G_value2 = "当期应缴增值税"
    r, c = find_cell(file_path, table_G_value2, table_G_sheet)
    c = get_column_letter(column_index_from_string(c) + 2)
    table_G_address2 = c + str(r)

    config = [
        {
            "operation": "custom",
            "formula_template": "=(IRR({A_val}:{B_val}))",
            "params": {
                "A_val": {"sheet": "", "cell": start_address},  # A表的y (向右移动)
                "B_val": {"sheet": "", "cell": end_address},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": target_start  # 目标起始位置 (向右移动)
            }
        },
        # 借款本金偿还
        {
            "operation": "custom",
            "formula_template": "=({A_val})",
            "params": {
                "A_val": {"sheet": table_c_sheet, "cell": address2}  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": target_address  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num + 1,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        },
        # 借款利息支付
        {
            "operation": "custom",
            "formula_template": "=({A_val}+{B_val}+{C_val})",
            "params": {
                "A_val": {"sheet": table_c_sheet, "cell": address1},  # A表的y (向右移动)
                "B_val": {"sheet": table_c_sheet, "cell": down_n_cells(address1, 2)},  # A表的y (向右移动)
                "C_val": {"sheet": table_c_sheet, "cell": down_n_cells(address1, 3)}  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": down_n_cells(target_address, 1)  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num + 1,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                    "B_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                    "C_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        },

        # 营业收入
        {
            "operation": "custom",
            "formula_template": "=({A_val})",
            "params": {
                "A_val": {"sheet": table_F_sheet, "cell": table_F_address},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": "D4"  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        },

        # 补贴收入
        {
            "operation": "custom",
            "formula_template": "=({A_val})",
            "params": {
                "A_val": {"sheet": table_F_sheet, "cell": table_F_address2},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": "D5"  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        },

        # 营业税金及附加
        {
            "operation": "custom",
            "formula_template": "=({A_val})",
            "params": {
                "A_val": {"sheet": table_G_sheet, "cell": table_G_address},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": "D13"  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        },

        # 增值税
        {
            "operation": "custom",
            "formula_template": "=({A_val})",
            "params": {
                "A_val": {"sheet": table_G_sheet, "cell": table_G_address2},  # A表的y (向右移动)

            },
            "target": {
                "sheet": sheet_name,  # 结果工作表
                "cell": "D14"  # 目标起始位置 (向右移动)
            },
            "loop": {
                "count": loop_num,  # 循环3次

                # 参数独立偏移设置
                "param_offsets": {
                    "A_val": {"row_shift": 0, "col_shift": 1},  # 每次向下移动一行
                },

                # 目标单元格偏移设置
                "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
            }
        }
    ]

    formula_generator.generate_formulas(config)

    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    # 创建居中对齐样式
    alignment_center = Alignment(horizontal='center', vertical='center')
    # 设置全边框
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    ws.merge_cells(range_string=f"{target_start}:{target_end}")
    # 设置指定区域单元格居中对齐
    for row in ws[f"{target_start}:{target_end}"]:
        for cell in row:
            cell.alignment = alignment_center
            cell.border = border

    wb.save(file_path)
    wb.close()


if __name__ == "__main__":
    # main()
    input_path = "财务分析套表自做模板编程用.xlsm"
    target_year = 2048
    sheet_name = "a.1财务现金流量表"
    table_B_sheet = "B项目信息"

    just_copy(input_path, target_year)
    special_copy(input_path, target_year)

    # print(struct_years(input_path, table_B_sheet))

    # last_row = find_last_used_row(input_path, sheet_name) - 7
    # _, last_col = find_last_used_column(input_path, sheet_name)
    # summary(input_path, sheet_name, last_row, last_col)

    # table_a1(input_path, sheet_name, "I")
    print("done")
