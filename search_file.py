import pandas as pd
import openpyxl
from openpyxl import load_workbook
import json
import requests
from typing import Dict, List, Any, Optional
import re
from datetime import datetime
import time


class ExcelIntelligentValidator:
    def __init__(self, max_rows_per_sheet: int = 50, max_columns_per_sheet: int = 27):
        """
        初始化 Excel 智能校验器

        参数:
            max_rows_per_sheet: 每个工作表最大处理行数
            max_columns_per_sheet: 每个工作表最大处理列数
        """
        self.max_rows = max_rows_per_sheet
        self.max_columns = max_columns_per_sheet
        self.structured_data = {
            "metadata": {},
            "sheets": [],
            "cross_references": [],
            "key_metrics": {}
        }
        self.validation_results = {}
        self.modification_report = {}

    # ==================== 数据提取方法 ====================

    def extract_formulas(self, workbook, sheet_name: str) -> Dict[str, str]:
        """提取指定工作表中的所有公式"""
        formulas = {}
        sheet = workbook[sheet_name]

        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == 'f' and cell.value:  # 公式类型且有值
                    # 清理公式中的多余空格
                    formula = str(cell.value).strip()
                    if formula.startswith('='):
                        formula = formula[1:]  # 去掉等号
                    formulas[cell.coordinate] = formula

        return formulas

    def extract_cross_references(self, formulas: Dict[str, str]) -> List[Dict[str, str]]:
        """从公式中提取跨表引用"""
        references = []
        pattern = r'(\[.*?\])?\'?([^\!]+?)\'?!\$?[A-Z]+\$?[0-9]+'

        for cell, formula in formulas.items():
            matches = re.findall(pattern, formula)
            for match in matches:
                ref_workbook, ref_sheet = match
                ref_workbook = ref_workbook[1:-1] if ref_workbook else ""  # 去掉方括号

                references.append({
                    "source_cell": cell,
                    "target_sheet": ref_sheet,
                    "target_workbook": ref_workbook,
                    "formula": formula
                })

        return references

    def detect_data_types(self, df: pd.DataFrame) -> Dict[str, str]:
        """检测每列的数据类型"""
        type_mapping = {
            'int64': 'integer',
            'float64': 'float',
            'bool': 'boolean',
            'datetime64[ns]': 'date',
            'object': 'text'
        }

        column_types = {}
        for column in df.columns:
            dtype = str(df[column].dtype)
            column_types[column] = type_mapping.get(dtype, 'text')

        return column_types

    def extract_key_metrics(self, df: pd.DataFrame, sheet_name: str) -> Dict[str, Any]:
        """提取关键指标和统计数据"""
        metrics = {}

        # 数值列的统计信息
        numeric_cols = df.select_dtypes(include=['number']).columns
        for col in numeric_cols:
            metrics[f"{sheet_name}_{col}_stats"] = {
                "sum": float(df[col].sum()),
                "mean": float(df[col].mean()),
                "min": float(df[col].min()),
                "max": float(df[col].max()),
                "count": int(df[col].count())
            }

        # 文本列的统计信息
        text_cols = df.select_dtypes(include=['object']).columns
        for col in text_cols:
            unique_vals = df[col].unique()
            metrics[f"{sheet_name}_{col}_stats"] = {
                "unique_count": len(unique_vals),
                "sample_values": [str(val) for val in unique_vals[:5]] if len(unique_vals) > 0 else []
            }

        return metrics

    def read_excel_to_structured_data(self, file_path: str,
                                      include_formulas: bool = True,
                                      include_data_stats: bool = True) -> Dict[str, Any]:
        """
        读取 Excel 文件并转换为结构化数据

        参数:
            file_path: Excel 文件路径
            include_formulas: 是否包含公式信息
            include_data_stats: 是否包含数据统计信息

        返回:
            结构化数据字典
        """
        try:
            # 读取文件基本信息
            self.structured_data["metadata"] = {
                "file_name": file_path.split("/")[-1],
                "processed_at": datetime.now().isoformat(),
                "total_sheets": 0
            }

            # 使用 pandas 读取数据
            excel_file = pd.ExcelFile(file_path)
            self.structured_data["metadata"]["total_sheets"] = len(excel_file.sheet_names)

            # 使用 openpyxl 读取公式和详细信息
            workbook = load_workbook(file_path, data_only=False)

            for sheet_name in excel_file.sheet_names:
                # 读取数据
                df = pd.read_excel(excel_file, sheet_name=sheet_name)

                # 限制处理的行列数
                df = df.iloc[:self.max_rows, :self.max_columns]

                # 处理数据
                sheet_data = {
                    "sheet_name": sheet_name,
                    "dimensions": {
                        "rows": len(df),
                        "columns": len(df.columns)
                    },
                    "headers": list(df.columns),
                    "data": df.where(pd.notnull(df), None).values.tolist(),
                    "data_types": self.detect_data_types(df)
                }

                # 提取公式
                if include_formulas:
                    formulas = self.extract_formulas(workbook, sheet_name)
                    sheet_data["formulas"] = formulas

                    # 提取跨表引用
                    cross_refs = self.extract_cross_references(formulas)
                    self.structured_data["cross_references"].extend(cross_refs)

                # 提取关键指标
                if include_data_stats:
                    metrics = self.extract_key_metrics(df, sheet_name)
                    self.structured_data["key_metrics"].update(metrics)

                self.structured_data["sheets"].append(sheet_data)

            return self.structured_data

        except Exception as e:
            print(f"读取 Excel 文件时出错: {e}")
            return {"error": str(e)}

    def to_llm_prompt(self, max_tokens: int = 4000) -> str:
        """
        将结构化数据转换为适合大模型的提示文本格式

        参数:
            max_tokens: 最大 token 数量限制

        返回:
            格式化后的提示文本
        """
        prompt_parts = []

        # 添加元数据
        prompt_parts.append(f"# Excel 文件分析: {self.structured_data['metadata']['file_name']}")
        prompt_parts.append(f"处理时间: {self.structured_data['metadata']['processed_at']}")
        prompt_parts.append(f"工作表数量: {self.structured_data['metadata']['total_sheets']}")
        prompt_parts.append("")

        # 添加每个工作表的信息
        for sheet in self.structured_data["sheets"]:
            prompt_parts.append(f"## 工作表: {sheet['sheet_name']}")
            prompt_parts.append(f"维度: {sheet['dimensions']['rows']} 行 × {sheet['dimensions']['columns']} 列")
            prompt_parts.append("")

            # 添加表头
            prompt_parts.append("### 表头")
            prompt_parts.append("| " + " | ".join(sheet['headers']) + " |")
            prompt_parts.append("|" + "|".join(["---"] * len(sheet['headers'])) + "|")

            # 添加前几行数据
            prompt_parts.append("### 数据示例")
            for i, row in enumerate(sheet['data'][:5]):  # 只显示前5行
                row_str = "| " + " | ".join([str(cell) if cell is not None else "" for cell in row]) + " |"
                prompt_parts.append(row_str)

            # 添加公式信息
            if "formulas" in sheet and sheet["formulas"]:
                prompt_parts.append("")
                prompt_parts.append("### 关键公式")
                for cell, formula in list(sheet["formulas"].items())[:10]:  # 只显示前10个公式
                    prompt_parts.append(f"- `{cell}: ={formula}`")

            prompt_parts.append("")

        # 添加跨表引用信息
        if self.structured_data["cross_references"]:
            prompt_parts.append("## 跨表引用")
            for ref in self.structured_data["cross_references"][:10]:  # 只显示前10个引用
                prompt_parts.append(f"- `{ref['source_cell']}` → `{ref['target_sheet']}`: {ref['formula']}")
            prompt_parts.append("")

        # 添加关键指标
        if self.structured_data["key_metrics"]:
            prompt_parts.append("## 关键指标摘要")
            for key, value in list(self.structured_data["key_metrics"].items())[:5]:  # 只显示前5个指标
                prompt_parts.append(f"- {key}: {value}")

        # 合并所有部分
        full_prompt = "\n".join(prompt_parts)

        # 如果内容过长，进行截断
        if len(full_prompt) > max_tokens * 3:  # 粗略估计，1 token ≈ 3-4 字符
            # 保留开头和结尾部分
            half_tokens = max_tokens * 3 // 2
            full_prompt = full_prompt[:half_tokens] + "\n\n...\n\n" + full_prompt[-half_tokens:]

        return full_prompt

    # ==================== 大模型交互方法 ====================

    def call_llm_api(self, prompt: str, api_key: str, api_url: str, model: str = "deepseek-chat") -> Dict[str, Any]:
        """
        调用大模型 API 进行 Excel 验证

        参数:
            prompt: 输入提示
            api_key: API 密钥
            api_url: API 端点
            model: 模型名称

        返回:
            大模型响应结果
        """
        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json"
        }

        payload = {
            "model": model,
            "messages": [
                {"role": "system", "content": "你是一个Excel专家，擅长财务模型验证和公式审计。"},
                {"role": "user", "content": prompt}
            ],
            "temperature": 0.1,
            "max_tokens": 2000
        }

        try:
            print("正在调用大模型 API...")
            response = requests.post(api_url, headers=headers, json=payload)
            response.raise_for_status()

            # 解析响应
            result = response.json()
            content = result['choices'][0]['message']['content']

            # 尝试解析 JSON 格式的响应
            try:
                return json.loads(content)
            except json.JSONDecodeError:
                # 如果不是 JSON 格式，返回原始内容
                return {"raw_response": content}

        except Exception as e:
            print(f"API调用错误: {e}")
            return {"error": str(e)}

    def validate_with_llm(self, file_path: str, api_key: str, api_url: str) -> Dict[str, Any]:
        """
        使用大模型验证 Excel 文件

        参数:
            file_path: Excel 文件路径
            api_key: 大模型 API 密钥
            api_url: 大模型 API 端点

        返回:
            验证结果
        """
        # 1. 读取 Excel 数据
        print("正在读取 Excel 数据...")
        self.read_excel_to_structured_data(file_path)

        # 2. 生成大模型提示
        print("正在生成大模型提示...")
        prompt = self.to_llm_prompt(max_tokens=4000)

        # 3. 调用大模型 API
        print("正在调用大模型进行验证...")
        self.validation_results = self.call_llm_api(prompt, api_key, api_url)

        # 4. 保存验证结果
        with open(file_path.replace('.xlsx', '_validation.json'), 'w', encoding='utf-8') as f:
            json.dump(self.validation_results, f, indent=2, ensure_ascii=False)

        return self.validation_results

    # ==================== Excel 修复方法 ====================

    def apply_fixes(self, file_path: str, validation_results: Dict[str, Any]) -> int:
        """
        应用修复建议到 Excel 文件

        参数:
            file_path: Excel 文件路径
            validation_results: 验证结果

        返回:
            应用的修复数量
        """
        try:
            # 加载工作簿
            wb = load_workbook(file_path)
            modifications = []

            # 应用问题修复
            issues = validation_results.get('issues', [])
            for issue in issues:
                sheet_name = issue.get('sheet')
                cell_ref = issue.get('cell')
                recommended_value = issue.get('recommended_value')

                if not all([sheet_name, cell_ref, recommended_value]):
                    continue

                try:
                    sheet = wb[sheet_name]
                    cell = sheet[cell_ref]

                    # 记录原始值
                    original_value = cell.value

                    # 应用修复
                    cell.value = recommended_value

                    # 记录修改
                    modifications.append({
                        'type': 'fix',
                        'sheet': sheet_name,
                        'cell': cell_ref,
                        'original': str(original_value),
                        'new': recommended_value,
                        'issue_id': issue.get('issue_id', ''),
                        'description': issue.get('description', '')
                    })

                except Exception as e:
                    print(f"修复失败 {sheet_name}!{cell_ref}: {e}")

            # 应用优化建议
            suggestions = validation_results.get('suggestions', [])
            for suggestion in suggestions:
                if "添加首年计息月份列" in suggestion.get('description', ''):
                    self.add_first_year_months_column(wb, modifications)

            # 保存修改后的文件
            new_file_path = file_path.replace('.xlsx', '_fixed.xlsx')
            wb.save(new_file_path)

            # 生成修改报告
            self.modification_report = {
                'original_file': file_path,
                'modified_file': new_file_path,
                'modifications': modifications,
                'summary': {
                    'total_fixes': len([m for m in modifications if m['type'] == 'fix']),
                    'total_improvements': len([m for m in modifications if m['type'] != 'fix']),
                    'processed_at': datetime.now().isoformat()
                }
            }

            # 保存修改报告
            with open(file_path.replace('.xlsx', '_modification_report.json'), 'w', encoding='utf-8') as f:
                json.dump(self.modification_report, f, indent=2, ensure_ascii=False)

            return len(modifications)

        except Exception as e:
            print(f"应用修复时出错: {e}")
            return 0

    def add_first_year_months_column(self, wb, modifications: List[Dict]) -> bool:
        """添加首年计息月份列"""
        try:
            sheet_name = "项目融资信息"
            if sheet_name not in wb.sheetnames:
                return False

            sheet = wb[sheet_name]

            # 确定插入位置（在最后一列之后）
            new_col_idx = sheet.max_column + 1

            # 设置列标题
            sheet.cell(row=1, column=new_col_idx, value="首年计息月数")

            # 查找开始时间列（假设是第5列，E列）
            start_date_col = 5

            # 填充数据
            for row_idx in range(2, sheet.max_row + 1):
                start_date = sheet.cell(row=row_idx, column=start_date_col).value
                if start_date and isinstance(start_date, (int, float)):
                    # 计算首年计息月数
                    from datetime import datetime
                    from openpyxl.utils import from_excel

                    try:
                        date_obj = from_excel(start_date)
                        months = 13 - date_obj.month  # 当年剩余月数+1
                        sheet.cell(row=row_idx, column=new_col_idx, value=months)
                    except:
                        sheet.cell(row=row_idx, column=new_col_idx, value=12)
                else:
                    sheet.cell(row=row_idx, column=new_col_idx, value=12)

            # 记录修改
            modifications.append({
                'type': 'add_column',
                'sheet': sheet_name,
                'column': '首年计息月数',
                'position': openpyxl.utils.get_column_letter(new_col_idx),
                'description': '添加首年计息月份列简化计算'
            })

            return True

        except Exception as e:
            print(f"添加首年计息月份列失败: {e}")
            return False

    # ==================== 完整流程方法 ====================

    def full_validation_pipeline(self, file_path: str, api_key: str, api_url: str) -> Dict[str, Any]:
        """
        执行完整的 Excel 验证和修复流程

        参数:
            file_path: Excel 文件路径
            api_key: 大模型 API 密钥
            api_url: 大模型 API 端点

        返回:
            完整流程结果
        """
        start_time = time.time()

        print("=" * 50)
        print("开始 Excel 智能验证与修复流程")
        print("=" * 50)

        # 1. 使用大模型验证 Excel
        validation_results = self.validate_with_llm(file_path, api_key, api_url)

        if "error" in validation_results:
            print("验证失败:", validation_results["error"])
            return {"status": "error", "message": validation_results["error"]}

        print("大模型验证完成，发现 {} 个问题".format(
            len(validation_results.get('issues', []))
        ))

        # 2. 应用修复
        fix_count = self.apply_fixes(file_path, validation_results)

        print("应用了 {} 处修复".format(fix_count))

        # 3. 生成最终报告
        end_time = time.time()
        processing_time = end_time - start_time

        final_report = {
            "status": "success",
            "processing_time_seconds": round(processing_time, 2),
            "validation_results": validation_results,
            "modification_report": self.modification_report,
            "files_generated": [
                file_path.replace('.xlsx', '_validation.json'),
                file_path.replace('.xlsx', '_fixed.xlsx'),
                file_path.replace('.xlsx', '_modification_report.json')
            ]
        }

        # 保存最终报告
        with open(file_path.replace('.xlsx', '_final_report.json'), 'w', encoding='utf-8') as f:
            json.dump(final_report, f, indent=2, ensure_ascii=False)

        print("=" * 50)
        print("Excel 智能验证与修复流程完成")
        print("=" * 50)

        return final_report


# 使用示例
def main():
    # 初始化校验器
    validator = ExcelIntelligentValidator(max_rows_per_sheet=100, max_columns_per_sheet=15)

    # 配置大模型 API
    api_key = "sk-9fc773ee795847d1b71d5fd2ea3546fd"  # 替换为您的 API 密钥
    api_url = "https://api.deepseek.com/v1/chat/completions"  # 替换为实际的 API 端点

    # 执行完整流程
    file_path = "工作簿2.xlsx"
    result = validator.full_validation_pipeline(file_path, api_key, api_url)

    # 打印结果摘要
    if result["status"] == "success":
        print("\n流程结果摘要:")
        print(f"- 处理时间: {result['processing_time_seconds']} 秒")
        print(f"- 发现问题: {len(result['validation_results'].get('issues', []))} 个")
        print(f"- 应用修复: {result['modification_report']['summary']['total_fixes']} 处")
        print(f"- 生成文件: {', '.join(result['files_generated'])}")
    else:
        print("流程执行失败:", result.get("message", "未知错误"))


if __name__ == "__main__":
    main()