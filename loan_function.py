from typing import List, Tuple
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl_vba import load_workbook

def right_n_cells(addr: str, n: int) -> str:
    """
    addr: 形如 "D5" 的单元格地址
    n   : 向右偏移量
    return: 偏移后的地址，例如 "I5"
    """
    col_str, row_str = addr[0], addr[1:]  # 拆分列字母与行号
    col_idx = column_index_from_string(col_str)  # 字母 → 数字
    new_col_idx = col_idx + n  # 向右移动 n 格
    new_col = get_column_letter(new_col_idx)  # 数字 → 字母
    return f"{new_col}{row_str}"


def to_absolute_address(cell_ref):
    """
    将 Excel 单元格地址转换为绝对引用格式

    参数:
    cell_ref (str): 单元格地址（如 "A1", "B3", "C.2项目每年借款信息!D4"）

    返回:
    str: 绝对引用格式的地址（如 "$A$1", "C.2项目每年借款信息!$D$4"）
    """
    # 检查是否包含工作表名称
    if '!' in cell_ref:
        # 分离工作表名称和单元格地址
        sheet_name, cell_address = cell_ref.split('!', 1)

        # 处理带引号的工作表名（如 'Sheet Name'!A1）
        if sheet_name.startswith("'") and sheet_name.endswith("'"):
            sheet_name = f"'{sheet_name.strip("'")}'"

        # 转换单元格部分为绝对引用
        return f"{sheet_name}!{_convert_cell_part(cell_address)}"
    else:
        # 直接转换单元格地址
        return _convert_cell_part(cell_ref)


def _convert_cell_part(cell_address):
    """处理单元格地址部分（不含工作表名称）"""
    # 分离列字母和行数字
    col_part = ''.join(filter(str.isalpha, cell_address))
    row_part = ''.join(filter(str.isdigit, cell_address))

    # 添加$符号创建绝对引用
    return f"${col_part}${row_part}"


def repay_method_cal(file_path, sheet_name, target, C_1_sheet_name, C_1_money, year_num, method):
    """
    7 种还款方式（利息按年计算）
    1) 到期一次性还清
    2) 贷款期内本期等额本金还款
    3) 贷款期内每年等额本息  (未完成, 用的 2）的计算公式)
    4) 后五年每年还本 20%
    5) 后十年每年还本 10%
    6) 后二十年每年还本 5%
    7) 自定义还款  (未完成, 用的1）的计算公式)
    """

    C_1_year = right_n_cells(C_1_money, 3)

    # 加载 Excel 文件（确保路径正确）
    wb = load_workbook(file_path)
    ws = wb[C_1_sheet_name]  # 或 ws = wb['Sheet1']

    # 读取 G4 单元格的值
    C_1_year_num = ws[C_1_year].value
    print(C_1_year, "的值是：", C_1_year_num)

    # 更改为绝对地址
    C_1_money = to_absolute_address(C_1_money)
    C_1_year = to_absolute_address(C_1_year)
    repay_config_method = []



    # 1) 到期一次性还清
    if method == "到期一次性还清":
        repay_config_method = [
            {
                "operation": "custom",
                "formula_template": "=({B_val})",
                "params": {
                    "B_val": {"sheet": C_1_sheet_name, "cell": C_1_money},  # B表的y (向右移动)

                },
                "target": {
                    "sheet": sheet_name,  # 结果工作表
                    "cell": right_n_cells(target, year_num)  # 目标起始位置 (向右移动)
                },

            }
        ]

    # 2) 贷款期内每年等额本金
    elif method == "贷款期内本期等额本金还款":
        repay_config_method = [
            {
                "operation": "custom",
                "formula_template": "=({B_val}/{C_val})",
                "params": {
                    "B_val": {"sheet": C_1_sheet_name, "cell": C_1_money},  # B表的y (向右移动)
                    "C_val": {"sheet": C_1_sheet_name, "cell": C_1_year},  # C表的z (向右移动)
                },
                "target": {
                    "sheet": sheet_name,  # 结果工作表
                    "cell": target  # 目标起始位置 (向右移动)
                },
                "loop": {
                    "count": int(C_1_year_num),  # 循环3次

                    # 参数独立偏移设置
                    "param_offsets": {
                        "B_val": {"row_shift": 0, "col_shift": 0},  # 每次向右移动一列
                        "C_val": {"row_shift": 0, "col_shift": 0},  # 每次向右移动一列
                    },

                    # 目标单元格偏移设置
                    "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
                }
            }
        ]

    # 3) 贷款期内每年等额本息
    elif method == "贷款期内本期等额本息还款":
        repay_config_method = [
            {
                "operation": "custom",
                "formula_template": "=({B_val}/{C_val})",
                "params": {
                    "B_val": {"sheet": C_1_sheet_name, "cell": C_1_money},  # B表的y (向右移动)
                    "C_val": {"sheet": C_1_sheet_name, "cell": C_1_year},  # C表的z (向右移动)
                },
                "target": {
                    "sheet": sheet_name,  # 结果工作表
                    "cell": target  # 目标起始位置 (向右移动)
                },
                "loop": {
                    "count": int(C_1_year_num),  # 循环3次

                    # 参数独立偏移设置
                    "param_offsets": {
                        "B_val": {"row_shift": 0, "col_shift": 0},  # 每次向右移动一列
                        "C_val": {"row_shift": 0, "col_shift": 0},  # 每次向右移动一列
                    },

                    # 目标单元格偏移设置
                    "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
                }
            }
        ]

    # 4) 后五年每年还本 20%
    elif method == "后五年每年还本20%":
        repay_config_method = [
            {
                "operation": "custom",
                "formula_template": "=({B_val}/5)",
                "params": {
                    "B_val": {"sheet": C_1_sheet_name, "cell": C_1_money},  # B表的y (向右移动)
                },
                "target": {
                    "sheet": sheet_name,  # 结果工作表
                    "cell": right_n_cells(target, year_num - 5)  # 目标起始位置 (向右移动)
                },
                "loop": {
                    "count": 5,  # 循环3次

                    # 参数独立偏移设置
                    "param_offsets": {
                        "B_val": {"row_shift": 0, "col_shift": 0},  # 每次向右移动一列
                    },

                    # 目标单元格偏移设置
                    "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
                }
            }
        ]

    # 5) 后十年每年还本 10%
    elif method == "后十年每年还本10%":
        repay_config_method = [
            {
                "operation": "custom",
                "formula_template": "=({B_val}/10)",
                "params": {
                    "B_val": {"sheet": C_1_sheet_name, "cell": C_1_money},  # B表的y (向右移动)
                },
                "target": {
                    "sheet": sheet_name,  # 结果工作表
                    "cell": right_n_cells(target, year_num - 10)  # 目标起始位置 (向右移动)
                },
                "loop": {
                    "count": 10,  # 循环3次

                    # 参数独立偏移设置
                    "param_offsets": {
                        "B_val": {"row_shift": 0, "col_shift": 0},  # 每次向右移动一列
                    },

                    # 目标单元格偏移设置
                    "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
                }
            }
        ]

    # 6) 后二十年每年还本 5%
    elif method == "后二十年每年还本5%":
        repay_config_method = [
            {
                "operation": "custom",
                "formula_template": "=({B_val}/20)",
                "params": {
                    "B_val": {"sheet": C_1_sheet_name, "cell": C_1_money},  # B表的y (向右移动)
                },
                "target": {
                    "sheet": sheet_name,  # 结果工作表
                    "cell": right_n_cells(target, year_num - 20)  # 目标起始位置 (向右移动)
                },
                "loop": {
                    "count": 20,  # 循环3次

                    # 参数独立偏移设置
                    "param_offsets": {
                        "B_val": {"row_shift": 0, "col_shift": 0},  # 每次向右移动一列
                    },

                    # 目标单元格偏移设置
                    "target_offset": {"row_shift": 0, "col_shift": 1}  # 每次向右移动一列且向下移动一行
                }
            }
        ]

    # 7) 自定义还款
    elif method == "自定义还款":
        repay_config_method = [
            {
                "operation": "custom",
                "formula_template": "=({B_val})",
                "params": {
                    "B_val": {"sheet": C_1_sheet_name, "cell": C_1_money},  # B表的y (向右移动)

                },
                "target": {
                    "sheet": sheet_name,  # 结果工作表
                    "cell": right_n_cells(target, year_num)  # 目标起始位置 (向右移动)
                },

            }
        ]

    return repay_config_method


# ------------------------- Demo ------------------------- #
if __name__ == '__main__':
    sheet_name = "c借款还本付息计划表"
    target = "D5"
    C_1_sheet_name = "C.1项目融资信息"
    C_1_money = "D4"
    year_num = 5
    method = "贷款期内本期等额本金还款"

    config = repay_method_cal(sheet_name, target, C_1_sheet_name, C_1_money, year_num, method)
    print("config:", config)
    print("done")
